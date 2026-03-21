import os, json, re, time, threading
import urllib.request, urllib.parse
from flask import Flask, request, jsonify, send_from_directory, session
from dotenv import load_dotenv
from functools import wraps
import openpyxl
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash

load_dotenv()
app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-change-me')

# ---------------------------------------------------------------------------
# LANGUAGE PROMPT  (loaded once at startup; used as AI system prompt)
# ---------------------------------------------------------------------------
def _load_language_prompt():
    path = os.path.join(os.path.dirname(__file__), 'prompts', 'lane4_language_prompt.txt')
    try:
        with open(path, encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        return ''

LANE4_LANGUAGE_PROMPT = _load_language_prompt()

def _load_deep_dive_prompt():
    path = os.path.join(os.path.dirname(__file__), 'prompts', 'lane4_deep_dive_prompt.txt')
    try:
        with open(path, encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        return ''

LANE4_DEEP_DIVE_PROMPT = _load_deep_dive_prompt()

# ---------------------------------------------------------------------------
# DATABASE
# ---------------------------------------------------------------------------
def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def _init_db():
    """Create tables if they don't exist (safe to run on every startup)."""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id            SERIAL PRIMARY KEY,
                    email         TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    created_at    TIMESTAMPTZ DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS sync_data (
                    id         SERIAL PRIMARY KEY,
                    user_id    INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    data_key   TEXT NOT NULL,
                    data_value JSONB,
                    updated_at TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE (user_id, data_key)
                )
            """)
        conn.commit()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        return f(*args, **kwargs)
    return decorated

# ---------------------------------------------------------------------------
# AUTH ENDPOINTS
# ---------------------------------------------------------------------------
@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    body = request.get_json(silent=True) or {}
    email    = (body.get('email') or '').strip().lower()
    password = (body.get('password') or '').strip()
    if not email or '@' not in email:
        return jsonify({'error': 'A valid email is required.'}), 400
    if len(password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters.'}), 400
    pw_hash = generate_password_hash(password)
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'INSERT INTO users (email, password_hash) VALUES (%s, %s) RETURNING id',
                    (email, pw_hash)
                )
                user_id = cur.fetchone()[0]
        session['user_id'] = user_id
        session['email']   = email
        return jsonify({'ok': True, 'email': email})
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'An account with that email already exists.'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    body = request.get_json(silent=True) or {}
    email    = (body.get('email') or '').strip().lower()
    password = (body.get('password') or '').strip()
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute('SELECT id, password_hash FROM users WHERE email = %s', (email,))
                row = cur.fetchone()
        if not row or not check_password_hash(row['password_hash'], password):
            return jsonify({'error': 'Incorrect email or password.'}), 401
        session['user_id'] = row['id']
        session['email']   = email
        return jsonify({'ok': True, 'email': email})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    if 'user_id' not in session:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'email': session.get('email'), 'user_id': session['user_id']})

# ---------------------------------------------------------------------------
# DATA SYNC ENDPOINTS
# ---------------------------------------------------------------------------
_ALLOWED_KEYS = {'swimmer', 'my_list', 'crm_data', 'vibe_state', 'other_prefs', 'preferences'}

@app.route('/api/data/load', methods=['GET'])
@login_required
def data_load():
    user_id = session['user_id']
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    'SELECT data_key, data_value FROM sync_data WHERE user_id = %s',
                    (user_id,)
                )
                rows = cur.fetchall()
        result = {r['data_key']: r['data_value'] for r in rows}
        return jsonify({'ok': True, 'data': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/save', methods=['POST'])
@login_required
def data_save():
    user_id = session['user_id']
    body    = request.get_json(silent=True) or {}
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                for key, value in body.items():
                    if key not in _ALLOWED_KEYS:
                        continue
                    cur.execute(
                        '''INSERT INTO sync_data (user_id, data_key, data_value, updated_at)
                           VALUES (%s, %s, %s::jsonb, NOW())
                           ON CONFLICT (user_id, data_key)
                           DO UPDATE SET data_value = EXCLUDED.data_value, updated_at = NOW()''',
                        (user_id, key, json.dumps(value))
                    )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# TEAM NAME NORMALIZATIONS
# Workbook cell widths truncate long names and add trailing qualifiers.
# These are the ONLY 9 places where workbook text differs from canonical names.
# Flagged explicitly per guardrail — no silent guessing.
# ---------------------------------------------------------------------------
TEAM_NAME_MAP = {
    "McDaniel College Swim Team":     ("McDaniel College",                 "trailing qualifier"),
    "Rochester Institute of Technol": ("Rochester Institute of Technology", "truncated at 30 chars"),
    "Rensselaer Polytechnic Institu": ("Rensselaer Polytechnic Institute",  "truncated at 30 chars"),
    "Union College (New York)":        ("Union College",                     "parenthetical qualifier"),
    "Massachusetts Institute of Tec": ("MIT",                               "truncated — schema uses 'MIT'"),
    "Worcester Polytechnic Institut": ("Worcester Polytechnic Institute",   "truncated at 30 chars"),
    "Wheaton College (Ma)":            ("Wheaton College (MA)",              "wrong casing"),
    "Whitworth University Swim Team": ("Whitworth University",              "trailing qualifier"),
    "California Institute of Techno": ("Caltech",                           "truncated — schema uses 'Caltech'"),
    "Saint Johns University":          ("Saint John's University",           "missing apostrophe"),
    "Harvard Men's Swimming":          ("Harvard University",                 "gendered PDF team name artifact"),
    # UAA schools — snapshot uses full institutional names; SCHOOL_META uses short forms
    "Brandeis University":             ("Brandeis",                          "SCHOOL_META uses short name"),
    "Carnegie Mellon University":      ("Carnegie Mellon",                   "SCHOOL_META drops University"),
    "Case Western Reserve Universit":  ("Case Western",                      "truncated + SCHOOL_META short name"),
    "Case Western Reserve University": ("Case Western",                      "SCHOOL_META uses short name"),
    "Emory University":                ("Emory",                             "SCHOOL_META drops University"),
    "New York University":             ("NYU",                               "SCHOOL_META uses abbreviation"),
    "University of Chicago":           ("Chicago",                           "SCHOOL_META uses city name"),
    "University of Rochester":         ("Rochester",                         "SCHOOL_META uses city name"),
    "Washington University St Louis":  ("Washington (Mo)",                   "SCHOOL_META uses location suffix"),
    "Washington University in St. Lou":("Washington (Mo)",                   "SCHOOL_META uses location suffix"),
}

# ---------------------------------------------------------------------------
# JAMES — hardcoded swimmer profile (source of truth for all scoring runs)
# ---------------------------------------------------------------------------
JAMES = {
    "name":             "James",
    "gpa":              4.0,
    "sat":              1460,
    "satProjected":     1500,
    "actScore":         0,
    "apCount":          0,
    "mathSat":          720,
    "mathSatProjected": 760,
    "times": {
        "1650 Free":              "16:06",
        "1000 Free":              "9:30",
        "500 Free":               "4:37",
        "200 Free":               "1:43",
        "400 IM":                 "4:09",
        "200 IM":                 "1:56",
        "100 Breast":             "59.5",
        "50 Breast (Relay Split)": "25.68",
    },
    "vibe": {
        "campus":   "Small and tight-knit — everyone knows everyone",
        "friday":   "Library with 2–3 close friends",
        "academic": "Genuinely want to be well-rounded",
        "compete":  "Love pushing myself inside a team environment",
        "location": None,
        "career":   None,
    },
}

# ---------------------------------------------------------------------------
# Conference → NCAA division label (used by build_school_universe)
CONF_DIVISION: dict[str, str] = {
    # ── D1 ────────────────────────────────────────────────────────────────────
    "ACC":            "D1", "ASUN":          "D1", "America East":  "D1",
    "Atlantic 10":    "D1", "Big 12":        "D1", "Big East":      "D1",
    "Big Ten":        "D1", "Big West":      "D1", "CAA":           "D1",
    "Horizon League": "D1", "Ivy League":    "D1", "MAAC":          "D1",
    "MPSF":           "D1", "Patriot":       "D1", "SEC":           "D1",
    "Summit League":  "D1", "WAC":           "D1",
    # ── D2 ────────────────────────────────────────────────────────────────────
    "GLIAC":          "D2", "PSAC":          "D2", "SAC":           "D2",
    # ── D3 ────────────────────────────────────────────────────────────────────
    "CCIW":           "D3", "Centennial":    "D3", "Colorado College": "D3",
    "Landmark":       "D3", "Liberty League":"D3", "MAC":           "D3",
    "MIAC":           "D3", "NCAC":          "D3", "NESCAC":        "D3",
    "NEWMAC":         "D3", "NWC":           "D3", "ODAC":          "D3",
    "SCIAC":          "D3", "UAA":           "D3",
    # ── NAIA ──────────────────────────────────────────────────────────────────
    "PCSC":           "NAIA",
}

# SCHOOL_META — per-school metadata for all 76 programs
# Fields: accept (int %), satMedian (int), hiddenIvy (bool), stem (bool),
#         merit ("none"|"moderate"|"high"), location (str), vibe (str),
#         moonshot (bool, optional)
# Keys must match canonical names after TEAM_NAME_MAP normalization.
# ---------------------------------------------------------------------------
SCHOOL_META = {
    # ── CENTENNIAL ───────────────────────────────────────────────────────────
    "Johns Hopkins University": {
        "accept": 7, "satMedian": 1510, "sat25": 1530, "sat75": 1560, "gpaMean": 3.93, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Baltimore, MD",
        "vibe": "Research powerhouse where pre-med and STEM culture run the campus",
    },
    "Gettysburg College": {
        "accept": 43, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Gettysburg, PA",
        "vibe": "Historic campus with strong Greek life and leadership culture",
    },
    "Swarthmore College": {
        "accept": 7, "satMedian": 1505, "sat25": 1490, "sat75": 1550, "gpaMean": 3.91, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Swarthmore, PA",
        "vibe": "Academically intense and collaborative with Quaker roots",
    },
    "Dickinson College": {
        "accept": 48, "satMedian": 1235, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Carlisle, PA",
        "vibe": "Sustainability-focused with strong international programs",
    },
    "Franklin & Marshall College": {
        "accept": 34, "satMedian": 1280, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Lancaster, PA",
        "vibe": "Pre-professional culture with strong pre-law and alumni network",
    },
    "Ursinus College": {
        "accept": 67, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Collegeville, PA",
        "vibe": "Warm undergraduate-focused campus with strong research access",
    },
    "Washington College": {
        "accept": 75, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Chestertown, MD",
        "vibe": "Small waterfront campus; close community and strong writing tradition",
    },
    "McDaniel College": {
        "accept": 82, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Westminster, MD",
        "vibe": "Friendly and personal campus with strong teacher education programs",
    },
    # ── LIBERTY LEAGUE ───────────────────────────────────────────────────────
    "Ithaca College": {
        "accept": 68, "satMedian": 1180, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Ithaca, NY",
        "vibe": "Creative and artsy; media, music, and performance everywhere",
    },
    "Rochester Institute of Technology": {
        "accept": 73, "satMedian": 1295, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Rochester, NY",
        "vibe": "Tech and project-driven culture built around co-ops and real careers",
    },
    "Rensselaer Polytechnic Institute": {
        "accept": 63, "satMedian": 1410, "sat25": 1280, "sat75": 1480, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Troy, NY",
        "vibe": "Pure engineering culture; hard-working students who live for problems",
    },
    "Clarkson University": {
        "accept": 80, "satMedian": 1205, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Potsdam, NY",
        "vibe": "Tight-knit STEM community in the North Country; outdoorsy and close",
    },
    "Union College": {
        "accept": 38, "satMedian": 1335, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Schenectady, NY",
        "vibe": "Liberal arts meets engineering; theme houses and strong traditions",
    },
    "Skidmore College": {
        "accept": 29, "satMedian": 1300, "sat25": 1230, "sat75": 1390, "gpaMean": 3.76, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Saratoga Springs, NY",
        "vibe": "Creative and arts-forward campus in a lively upstate NY town",
    },
    "Vassar College": {
        "accept": 18, "satMedian": 1455, "sat25": 1400, "sat75": 1540, "gpaMean": 3.82, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Poughkeepsie, NY",
        "vibe": "Progressive intellectual culture; students who love big ideas",
    },
    "St. Lawrence University": {
        "accept": 60, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Canton, NY",
        "vibe": "Outdoorsy North Country campus with a close community and athletics",
    },
    "Hobart and William Smith": {
        "accept": 58, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Geneva, NY",
        "vibe": "Lakeside dual-college campus with strong social scene and sailing",
    },
    "Bard College": {
        "accept": 64, "satMedian": 1265, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Annandale-on-Hudson, NY",
        "vibe": "Artsy and progressive with discussion-heavy classes and bohemian feel",
    },
    # ── NCAC ─────────────────────────────────────────────────────────────────
    "Denison University": {
        "accept": 28, "satMedian": 1325, "sat25": 1200, "sat75": 1400, "gpaMean": 3.73, "hiddenIvy": True, "stem": False,
        "merit": "high", "location": "Granville, OH",
        "vibe": "Beautiful hilltop campus; ambitious academics and a strong social scene",
    },
    "Kenyon College": {
        "accept": 33, "satMedian": 1370, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Gambier, OH",
        "vibe": "Literary and deeply intellectual; famous writing program in rural Ohio",
    },
    "John Carroll University": {
        "accept": 80, "satMedian": 1135, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "University Heights, OH",
        "vibe": "Jesuit values; service-oriented with strong business programs",
    },
    "DePauw University": {
        "accept": 67, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Greencastle, IN",
        "vibe": "Greek life-heavy with strong communications and music programs",
    },
    "Wabash College": {
        "accept": 69, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Crawfordsville, IN",
        "vibe": "All-male liberal arts with intense brotherhood and strong traditions",
    },
    "College of Wooster": {
        "accept": 57, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Wooster, OH",
        "vibe": "Every senior writes a thesis; close-knit with strong independent study",
    },
    "Oberlin College": {
        "accept": 33, "satMedian": 1385, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Oberlin, OH",
        "vibe": "Progressive and artistic; famous conservatory and engaged campus politics",
    },
    "Ohio Wesleyan University": {
        "accept": 83, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Delaware, OH",
        "vibe": "Emphasis on global experience; internship-focused with civic engagement",
    },
    "Wittenberg University": {
        "accept": 86, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Springfield, OH",
        "vibe": "Lutheran-rooted campus; personal and strong in teacher education",
    },
    # ── NESCAC ───────────────────────────────────────────────────────────────
    "Williams College": {
        "accept": 9, "satMedian": 1510, "sat25": 1500, "sat75": 1560, "gpaMean": 3.94, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Williamstown, MA",
        "vibe": "Consistently ranked #1 LAC; mountain campus with elite academics",
    },
    "Tufts University": {
        "accept": 9, "satMedian": 1500, "sat25": 1390, "sat75": 1540, "gpaMean": 3.87, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Medford, MA",
        "vibe": "Globally minded near Boston; research-intensive with elite academics",
    },
    "Amherst College": {
        "accept": 9, "satMedian": 1515, "sat25": 1500, "sat75": 1560, "gpaMean": 3.93, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Amherst, MA",
        "vibe": "Open curriculum, no required courses; fiercely intellectual with 5-College access",
    },
    "Connecticut College": {
        "accept": 38, "satMedian": 1315, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "New London, CT",
        "vibe": "Student self-governance model; students run nearly everything on campus",
    },
    "Bates College": {
        "accept": 13, "satMedian": 1430, "sat25": 1330, "sat75": 1500, "gpaMean": 3.78, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Lewiston, ME",
        "vibe": "Politically engaged and outdoorsy; tight community in coastal Maine",
    },
    "Hamilton College": {
        "accept": 14, "satMedian": 1440, "sat25": 1360, "sat75": 1500, "gpaMean": 3.79, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Clinton, NY",
        "vibe": "Writing-intensive; every major leads to a thesis on a beautiful rural campus",
    },
    "Bowdoin College": {
        "accept": 9, "satMedian": 1495, "sat25": 1470, "sat75": 1540, "gpaMean": 3.86, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Brunswick, ME",
        "vibe": "Outdoorsy and intellectual in coastal Maine; sustainability and community",
    },
    "Middlebury College": {
        "accept": 13, "satMedian": 1445, "sat25": 1390, "sat75": 1520, "gpaMean": 3.83, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Middlebury, VT",
        "vibe": "Environmental passion meets rigorous academics in a beautiful Vermont setting",
    },
    "Colby College": {
        "accept": 11, "satMedian": 1435, "sat25": 1360, "sat75": 1490, "gpaMean": 3.79, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Waterville, ME",
        "vibe": "Liberal arts in the Maine wilderness; entrepreneurial with a tight community",
    },
    "Trinity College": {
        "accept": 34, "satMedian": 1310, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Hartford, CT",
        "vibe": "Classic New England campus with strong city partnerships and Greek life",
    },
    "Wesleyan University": {
        "accept": 17, "satMedian": 1455, "sat25": 1410, "sat75": 1530, "gpaMean": 3.85, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Middletown, CT",
        "vibe": "Quirky and politically active; film and social sciences define the culture",
    },
    # ── NEWMAC ───────────────────────────────────────────────────────────────
    "MIT": {
        "accept": 4, "satMedian": 1565, "sat25": 1500, "sat75": 1580, "gpaMean": 3.97, "hiddenIvy": False, "stem": True,
        "merit": "none", "moonshot": True, "location": "Cambridge, MA",
        "vibe": "The world's most famous STEM institution; unmatched resources and intensity",
    },
    "U.S. Coast Guard Academy": {
        "accept": 14, "satMedian": 1265, "sat25": 1190, "sat75": 1360, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "New London, CT",
        "vibe": "Military service academy; full scholarship, intense discipline, meaningful mission",
    },
    "Worcester Polytechnic Institute": {
        "accept": 58, "satMedian": 1370, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Worcester, MA",
        "vibe": "Project-based learning at a tech school with strong industry connections",
    },
    "Babson College": {
        "accept": 24, "satMedian": 1330, "sat25": 1260, "sat75": 1410, "gpaMean": 3.65, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Wellesley, MA",
        "vibe": "#1 entrepreneurship school; every freshman runs a real business for credit",
    },
    "Wheaton College (MA)": {
        "accept": 71, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Norton, MA",
        "vibe": "Small and personal campus reinventing itself with a bold connected curriculum",
    },
    "Springfield College": {
        "accept": 77, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Springfield, MA",
        "vibe": "Birthplace of basketball; health sciences and PT define the campus culture",
    },
    "Clark University": {
        "accept": 52, "satMedian": 1225, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Worcester, MA",
        "vibe": "Free fifth year for any grad program; research-first culture in an urban campus",
    },
    # ── NWC ──────────────────────────────────────────────────────────────────
    "Whitworth University": {
        "accept": 89, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Spokane, WA",
        "vibe": "Christian liberal arts in the Pacific Northwest with strong education programs",
    },
    "University of Puget Sound": {
        "accept": 87, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Tacoma, WA",
        "vibe": "NW outdoor culture meets classic liberal arts; tight-knit campus",
    },
    "Linfield University": {
        "accept": 88, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "McMinnville, OR",
        "vibe": "Small Oregon liberal arts in wine country; personal and community-focused",
    },
    "Whitman College": {
        "accept": 46, "satMedian": 1295, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Walla Walla, WA",
        "vibe": "Quirky Pacific NW intellectual gem with outdoor access and high grad school rates",
    },
    "Pacific Lutheran University": {
        "accept": 87, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Tacoma, WA",
        "vibe": "Lutheran values in the Pacific Northwest; strong music and education programs",
    },
    "Lewis & Clark College": {
        "accept": 65, "satMedian": 1275, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Portland, OR",
        "vibe": "Hippie-intellectual Portland culture; environmental law and social justice focus",
    },
    "Willamette University": {
        "accept": 81, "satMedian": 1190, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Salem, OR",
        "vibe": "NW LAC with strong law school connections and active civic engagement",
    },
    "George Fox University": {
        "accept": 93, "satMedian": 1090, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Newberg, OR",
        "vibe": "Christian university with strong community, athletics, and business programs",
    },
    # ── SCIAC ────────────────────────────────────────────────────────────────
    "Pomona-Pitzer": {
        "accept": 7, "satMedian": 1510, "sat25": 1470, "sat75": 1550, "gpaMean": 3.91, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Claremont, CA",
        "vibe": "Elite SoCal LAC in the Claremont Consortium; 5 colleges sharing resources",
    },
    "Claremont-Mudd-Scripps": {
        "accept": 9, "satMedian": 1490, "sat25": 1470, "sat75": 1560, "gpaMean": 3.9, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Claremont, CA",
        "vibe": "Harvey Mudd's STEM intensity meets Scripps' creative and humanistic edge",
    },
    "Chapman University": {
        "accept": 52, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Orange, CA",
        "vibe": "Film school prestige meets SoCal sunshine; entrepreneurial and media-forward",
    },
    "Caltech": {
        "accept": 3, "satMedian": 1560, "sat25": 1530, "sat75": 1580, "gpaMean": 3.97, "hiddenIvy": False, "stem": True,
        "merit": "none", "moonshot": True, "location": "Pasadena, CA",
        "vibe": "Hardest STEM school to enter in America; Nobel laureates teach undergrads",
    },
    "Whittier College": {
        "accept": 73, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Whittier, CA",
        "vibe": "Liberal arts with strong Latino heritage and close community feel in SoCal",
    },
    "Cal Lutheran University": {
        "accept": 61, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Thousand Oaks, CA",
        "vibe": "Lutheran roots in Ventura County; strong business and communications programs",
    },
    "Occidental College": {
        "accept": 37, "satMedian": 1295, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Los Angeles, CA",
        "vibe": "Progressive urban LAC in Eagle Rock; politics and international relations culture",
    },
    "University of Redlands": {
        "accept": 67, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Redlands, CA",
        "vibe": "Inland Empire LAC; strong music, environmental studies, and pre-law culture",
    },
    "University of La Verne": {
        "accept": 65, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "La Verne, CA",
        "vibe": "Close-knit SoCal campus with strong business and criminal justice programs",
    },
    # ── UAA ──────────────────────────────────────────────────────────────────
    # Names are abbreviated in the workbook — kept as-is (not truncation, just short forms)
    "Emory": {
        "accept": 12, "satMedian": 1470, "sat25": 1360, "sat75": 1530, "gpaMean": 3.87, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Atlanta, GA",
        "vibe": "Research powerhouse in Atlanta; dominant pre-med culture and strong social scene",
    },
    "NYU": {
        "accept": 12, "satMedian": 1460, "sat25": 1350, "sat75": 1530, "gpaMean": 3.79, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "New York, NY",
        "vibe": "Urban campus without borders; Greenwich Village is your quad in the heart of NYC",
    },
    "Chicago": {
        "accept": 6, "satMedian": 1530, "sat25": 1510, "sat75": 1570, "gpaMean": 3.94, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Chicago, IL",
        "vibe": "Intellectual intensity above all else; famous for taking ideas more seriously than sleep",
    },
    "Washington (Mo)": {
        "accept": 14, "satMedian": 1500, "sat25": 1480, "sat75": 1560, "gpaMean": 3.94, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "St. Louis, MO",
        "vibe": "Research powerhouse in the Midwest; strong pre-med, engineering, and business",
    },
    "Carnegie Mellon": {
        "accept": 11, "satMedian": 1535, "sat25": 1460, "sat75": 1560, "gpaMean": 3.87, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Pittsburgh, PA",
        "vibe": "Top CS and engineering with a rigorous, career-driven campus culture",
    },
    "Case Western": {
        "accept": 30, "satMedian": 1455, "sat25": 1430, "sat75": 1560, "gpaMean": 3.82, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Cleveland, OH",
        "vibe": "STEM-focused research university; pre-med and engineering define campus life",
    },
    "Rochester": {
        "accept": 29, "satMedian": 1440, "sat25": 1380, "sat75": 1530, "gpaMean": 3.83, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Rochester, NY",
        "vibe": "Research-intensive with strong engineering, optics, and pre-med programs",
    },
    "Brandeis": {
        "accept": 37, "satMedian": 1420, "sat25": 1380, "sat75": 1530, "gpaMean": 3.79, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Waltham, MA",
        "vibe": "Social justice mission and research strength near Boston with a unique founding story",
    },
    # ── MIAC ─────────────────────────────────────────────────────────────────
    "Gustavus Adolphus College": {
        "accept": 72, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Saint Peter, MN",
        "vibe": "Lutheran liberal arts with Swedish heritage and strong athletics in Minnesota",
    },
    "Carleton College": {
        "accept": 18, "satMedian": 1495, "sat25": 1470, "sat75": 1540, "gpaMean": 3.89, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Northfield, MN",
        "vibe": "One of the Midwest's best LACs; intellectual culture with top grad school placement",
    },
    "Saint John's University": {
        "accept": 75, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Collegeville, MN",
        "vibe": "Coordinate college with Saint Benedict; Benedictine tradition and strong community",
    },
    "Macalester College": {
        "accept": 28, "satMedian": 1430, "sat25": 1380, "sat75": 1520, "gpaMean": 3.85, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Saint Paul, MN",
        "vibe": "Globally focused and politically active urban LAC with high international enrollment",
    },
    "Saint Olaf College": {
        "accept": 48, "satMedian": 1270, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Northfield, MN",
        "vibe": "Norwegian Lutheran roots; world-famous music programs and close Minnesota community",
    },
    "Hamline University": {
        "accept": 82, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Saint Paul, MN",
        "vibe": "Urban liberal arts with social justice focus; personal and accessible in Twin Cities",
    },
    # ── MIAC (remaining 4 — 6 already above) ─────────────────────────────────
    "Augsburg University": {
        "accept": 72, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Minneapolis, MN",
        "vibe": "Urban Lutheran campus in Minneapolis with strong social work and nursing programs",
    },
    "College of Saint Benedict": {
        "accept": 69, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "St. Joseph, MN",
        "vibe": "Women's college partnered with Saint John's; strong Catholic identity and close community",
    },
    "Concordia College": {
        "accept": 60, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Moorhead, MN",
        "vibe": "Lutheran liberal arts on the Minnesota-North Dakota border; strong music and global programs",
    },
    "Saint Catherine University": {
        "accept": 62, "satMedian": 1110, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "St. Paul, MN",
        "vibe": "Catholic women's university in the Twin Cities with strong health sciences programs",
    },
    # ── NEWMAC (remaining 3 — 7 already above) ───────────────────────────────
    "Mount Holyoke College": {
        "accept": 49, "satMedian": 1335, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "South Hadley, MA",
        "vibe": "Historic women's college in the Pioneer Valley; Seven Sisters with strong social science traditions",
    },
    "Smith College": {
        "accept": 33, "satMedian": 1390, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Northampton, MA",
        "vibe": "Premier women's liberal arts college; consistently top-ranked with exceptional alumnae network",
    },
    "Wellesley College": {
        "accept": 14, "satMedian": 1440, "sat25": 1370, "sat75": 1510, "gpaMean": 3.87, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Wellesley, MA",
        "vibe": "Elite women's college near Boston; Hillary Clinton and Madeline Albright territory",
    },
    # ── ODAC ─────────────────────────────────────────────────────────────────
    "Bridgewater College": {
        "accept": 63, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Bridgewater, VA",
        "vibe": "Small Church of the Brethren college in the Shenandoah Valley with a close athletic community",
    },
    "Greensboro College": {
        "accept": 47, "satMedian": 1065, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Greensboro, NC",
        "vibe": "Methodist-affiliated urban college; small and personal with strong arts and teacher ed programs",
    },
    "Hampden-Sydney College": {
        "accept": 65, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Hampden Sydney, VA",
        "vibe": "All-male liberal arts college with a strong honor code tradition and close brotherhood",
    },
    "Hollins University": {
        "accept": 62, "satMedian": 1160, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Roanoke, VA",
        "vibe": "Women's university in the Blue Ridge with creative writing fame and an equestrian program",
    },
    "Randolph College": {
        "accept": 62, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Lynchburg, VA",
        "vibe": "Small coed liberal arts college with an equestrian program and strong individualized attention",
    },
    "Randolph-Macon College": {
        "accept": 58, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Ashland, VA",
        "vibe": "Methodist-affiliated LAC near Richmond with strong pre-law and business tracks",
    },
    "Roanoke College Swimming Maroo": {
        "accept": 65, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Salem, VA",
        "vibe": "Lutheran liberal arts college in the Blue Ridge foothills with a strong athletics tradition",
    },
    "Sweet Briar College": {
        "accept": 47, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Sweet Briar, VA",
        "vibe": "Women's college with an equestrian program on a stunning Virginia estate; intimate and resilient",
    },
    "University of Lynchburg": {
        "accept": 60, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Lynchburg, VA",
        "vibe": "Disciples of Christ-affiliated campus with strong health sciences and a growing athletics profile",
    },
    "Virginia Wesleyan University": {
        "accept": 74, "satMedian": 1090, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Norfolk, VA",
        "vibe": "Methodist university near Virginia Beach with a student-centered campus and growing athletics",
    },
    "Washington and Lee University": {
        "accept": 21, "satMedian": 1455, "sat25": 1390, "sat75": 1520, "gpaMean": 3.82, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Lexington, VA",
        "vibe": "Honor-code-driven LAC with an exceptional law school pipeline and Southern intellectual tradition",
    },
    # ── CCIW ─────────────────────────────────────────────────────────────────
    "Augustana College": {
        "accept": 66, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Rock Island, IL",
        "vibe": "Swedish Lutheran LAC on the Mississippi with strong pre-health and music programs",
    },
    "Carroll University": {
        "accept": 71, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Waukesha, WI",
        "vibe": "Presbyterian-rooted campus near Milwaukee with strong physical therapy and nursing programs",
    },
    "Carthage College": {
        "accept": 62, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Kenosha, WI",
        "vibe": "Lutheran-affiliated college on Lake Michigan with a semester-based calendar and strong arts",
    },
    "Illinois Wesleyan University": {
        "accept": 61, "satMedian": 1205, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Bloomington, IL",
        "vibe": "Highly selective for a Midwest LAC; strong theatre, pre-law, and business in a college town",
    },
    "Millikin University": {
        "accept": 59, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Decatur, IL",
        "vibe": "Performance-focused LAC with a nationally known entrepreneurship program and strong arts",
    },
    "North Central College": {
        "accept": 62, "satMedian": 1190, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Naperville, IL",
        "vibe": "Methodist-affiliated college in Chicago's suburbs with strong business and pre-professional programs",
    },
    "Wheaton College": {
        "accept": 64, "satMedian": 1300, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Wheaton, IL",
        "vibe": "Evangelical Christian LAC with rigorous academics; produces many graduate school attendees",
    },
    # ── LANDMARK ─────────────────────────────────────────────────────────────
    "Catholic University of America": {
        "accept": 76, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Washington, DC",
        "vibe": "The national university of the Catholic Church; DC location enables strong internship access",
    },
    "Drew University": {
        "accept": 63, "satMedian": 1205, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Madison, NJ",
        "vibe": "Methodis LAC in NJ suburbs; small and personal with strong connections to NYC and Wall Street",
    },
    "Elizabethtown College": {
        "accept": 67, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Elizabethtown, PA",
        "vibe": "Church of the Brethren college in Lancaster County; strong occupational therapy and social work",
    },
    "Goucher College": {
        "accept": 73, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Baltimore, MD",
        "vibe": "Globally focused LAC near Baltimore; every student studies abroad and faculty are highly accessible",
    },
    "Juniata College": {
        "accept": 68, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Huntingdon, PA",
        "vibe": "Student-designed majors and strong pre-health in rural Pennsylvania; personal and innovative",
    },
    "Lycoming College": {
        "accept": 72, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Williamsport, PA",
        "vibe": "Methodist LAC in north-central PA; generous merit aid and strong undergraduate research",
    },
    "Moravian University": {
        "accept": 70, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Bethlehem, PA",
        "vibe": "Historic Moravian campus in Bethlehem; strong nursing, business, and community engagement",
    },
    "Susquehanna University": {
        "accept": 81, "satMedian": 1160, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Selinsgrove, PA",
        "vibe": "Lutheran LAC on the Susquehanna River; strong communications and creative writing programs",
    },
    "University of Scranton": {
        "accept": 72, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Scranton, PA",
        "vibe": "Jesuit university in the Pocono foothills; strong business, health sciences, and service culture",
    },
    "Wilkes University": {
        "accept": 73, "satMedian": 1125, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Wilkes-Barre, PA",
        "vibe": "Engineering and health sciences focus in northeast PA; hands-on and career-oriented",
    },
    # ── MAC ───────────────────────────────────────────────────────────────────
    "Arcadia University": {
        "accept": 62, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Glenside, PA",
        "vibe": "Global-focused LAC in Philadelphia's suburbs with a stunning castle campus and strong study-abroad",
    },
    "Fairleigh Dickinson University": {
        "accept": 80, "satMedian": 1085, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Teaneck, NJ",
        "vibe": "Multi-campus metro NJ university with strong business and pharmacy programs",
    },
    "Hood College Swimming": {
        "accept": 62, "satMedian": 1170, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Frederick, MD",
        "vibe": "Small coed LAC in Frederick's historic district with strong biomedical science and education",
    },
    "King's College": {
        "accept": 71, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Wilkes-Barre, PA",
        "vibe": "Catholic LAC in northeast PA with strong physician assistant and business programs",
    },
    "Lebanon Valley College": {
        "accept": 66, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Annville, PA",
        "vibe": "United Methodist LAC in central PA; strong physical therapy, music, and athletic training",
    },
    "Messiah University": {
        "accept": 63, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Mechanicsburg, PA",
        "vibe": "Christian university with strong nursing, engineering, and social work near Harrisburg",
    },
    "Misericordia University": {
        "accept": 68, "satMedian": 1125, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Dallas, PA",
        "vibe": "Health sciences powerhouse in northeast PA; outstanding PT, OT, and nursing programs",
    },
    "Stevens Institute of Technolog": {
        "accept": 42, "satMedian": 1400, "sat25": 1310, "sat75": 1480, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Hoboken, NJ",
        "vibe": "Engineering and tech university on the Hudson with a stunning Manhattan skyline view",
    },
    "Stevenson University": {
        "accept": 57, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Owings Mills, MD",
        "vibe": "Career-focused university near Baltimore with strong forensics, nursing, and business programs",
    },
    "Widener University": {
        "accept": 64, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Chester, PA",
        "vibe": "Engineering, nursing, and law-pipeline campus near Philadelphia with strong professional programs",
    },
    "York College of Pennsylvania": {
        "accept": 61, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "York, PA",
        "vibe": "Business and nursing-focused university in south-central PA; career-oriented and affordable",
    },
    # ── MAAC ─────────────────────────────────────────────────────────────────
    "Canisius University": {
        "accept": 81, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Buffalo, NY",
        "vibe": "Jesuit university in Buffalo with strong business, health sciences, and athletics identity",
    },
    "Fairfield University": {
        "accept": 60, "satMedian": 1275, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Fairfield, CT",
        "vibe": "Jesuit university on Connecticut's Gold Coast; strong business and nursing near NYC",
    },
    "Iona University": {
        "accept": 72, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "New Rochelle, NY",
        "vibe": "Irish Christian Brothers university 20 minutes from Manhattan; strong business and communications",
    },
    "Manhattan University": {
        "accept": 76, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Riverdale, NY",
        "vibe": "Christian Brothers university in the Bronx with strong engineering and education programs",
    },
    "Marist University": {
        "accept": 49, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Poughkeepsie, NY",
        "vibe": "Scenic Hudson Valley campus with strong fashion, communications, and business programs",
    },
    "Merrimack College": {
        "accept": 77, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "North Andover, MA",
        "vibe": "Augustinian university north of Boston with strong health sciences and business programs",
    },
    "Mount Saint Mary's University": {
        "accept": 75, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Emmitsburg, MD",
        "vibe": "Catholic university in the Blue Ridge foothills with a close community and strong nursing",
    },
    "Niagara University": {
        "accept": 78, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Niagara University, NY",
        "vibe": "Vincentian Catholic university near Niagara Falls; strong education, social work, and business",
    },
    "Rider University": {
        "accept": 72, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Lawrenceville, NJ",
        "vibe": "Business and education-focused university near Princeton with Westminster Choir College",
    },
    "Sacred Heart University": {
        "accept": 72, "satMedian": 1180, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Fairfield, CT",
        "vibe": "Catholic university in affluent coastal Connecticut with strong PT, nursing, and athletics",
    },
    "Saint Peter's University": {
        "accept": 80, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Jersey City, NJ",
        "vibe": "Jesuit university across from lower Manhattan; strong business and pre-health in an urban setting",
    },
    "Siena University": {
        "accept": 71, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Loudonville, NY",
        "vibe": "Franciscan Catholic university near Albany with strong business, biology, and social work",
    },
    # ── PATRIOT ──────────────────────────────────────────────────────────────
    "American University": {
        "accept": 35, "satMedian": 1310, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Washington, DC",
        "vibe": "International relations and politics powerhouse in DC; students go straight from class to Capitol Hill",
    },
    "Boston University": {
        "accept": 14, "satMedian": 1450, "sat25": 1350, "sat75": 1500, "gpaMean": 3.73, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Boston, MA",
        "vibe": "Large research university on the Charles with elite engineering, business, and communications",
    },
    "Bucknell University": {
        "accept": 36, "satMedian": 1345, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Lewisburg, PA",
        "vibe": "Selective LAC-meets-engineering in central PA; strong alumni network and Greek life culture",
    },
    "Colgate University": {
        "accept": 21, "satMedian": 1420, "sat25": 1380, "sat75": 1510, "gpaMean": 3.82, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Hamilton, NY",
        "vibe": "Highly selective rural NY LAC with strong preprofessional culture and loyal athletic fanbase",
    },
    "College of the Holy Cross": {
        "accept": 32, "satMedian": 1360, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Worcester, MA",
        "vibe": "Jesuit LAC with rigorous core curriculum and strong Jesuit service tradition in New England",
    },
    "Lafayette College": {
        "accept": 33, "satMedian": 1325, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Easton, PA",
        "vibe": "Engineering-meets-liberal arts on a hilltop campus; strong alumni network and Division 1 rivalry",
    },
    "Lehigh University": {
        "accept": 33, "satMedian": 1385, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Bethlehem, PA",
        "vibe": "STEM-forward research university with a strong engineering and business culture in the Lehigh Valley",
    },
    "Loyola University": {
        "accept": 54, "satMedian": 1265, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Baltimore, MD",
        "vibe": "Jesuit university overlooking Baltimore; strong business, communications, and service programs",
    },
    "United States Military Academy": {
        "accept": 9, "satMedian": 1270, "sat25": 1200, "sat75": 1400, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "West Point, NY",
        "vibe": "West Point — full scholarship, intense commitment, and guaranteed career in military leadership",
    },
    "United States Navy Academy": {
        "accept": 9, "satMedian": 1290, "sat25": 1220, "sat75": 1390, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Annapolis, MD",
        "vibe": "Annapolis — full scholarship, rigorous engineering curriculum, and a Navy or Marine career ahead",
    },
    # ── IVY LEAGUE ───────────────────────────────────────────────────────────
    "Brown University": {
        "accept": 5, "satMedian": 1510, "sat25": 1440, "sat75": 1570, "gpaMean": 3.94, "hiddenIvy": False, "ivyLeague": True, "stem": False,
        "merit": "none", "location": "Providence, RI",
        "vibe": "Open Curriculum Ivy gives students unusual freedom — Brown is the most academically flexible of the eight",
    },
    "Columbia University": {
        "accept": 4, "satMedian": 1530, "sat25": 1470, "sat75": 1560, "gpaMean": 3.92, "hiddenIvy": False, "ivyLeague": True, "stem": True,
        "merit": "none", "location": "New York, NY",
        "vibe": "Ivy in Morningside Heights — the Core Curriculum meets the greatest city on earth",
    },
    "Cornell University": {
        "accept": 7, "satMedian": 1490, "sat25": 1400, "sat75": 1560, "gpaMean": 3.89, "hiddenIvy": False, "ivyLeague": True, "stem": True,
        "merit": "none", "location": "Ithaca, NY",
        "vibe": "The broadest Ivy — engineering, agriculture, hotel, and arts all on one spectacular gorge campus",
    },
    "Dartmouth College": {
        "accept": 6, "satMedian": 1510, "sat25": 1440, "sat75": 1560, "gpaMean": 3.95, "hiddenIvy": False, "ivyLeague": True, "stem": False,
        "merit": "none", "location": "Hanover, NH",
        "vibe": "Smallest Ivy with fierce alumni loyalty; outdoor culture, Greek life, and an undergrad-first focus",
    },
    "Harvard University": {
        "accept": 3, "satMedian": 1540, "sat25": 1460, "sat75": 1570, "gpaMean": 3.95, "hiddenIvy": False, "ivyLeague": True, "stem": True,
        "merit": "none", "location": "Cambridge, MA",
        "vibe": "The most recognized university brand in the world — extraordinary in every dimension",
    },
    "Princeton University": {
        "accept": 4, "satMedian": 1530, "sat25": 1460, "sat75": 1570, "gpaMean": 3.94, "hiddenIvy": False, "ivyLeague": True, "stem": True,
        "merit": "none", "location": "Princeton, NJ",
        "vibe": "No-loan financial aid and the strongest endowment per student of any university in America",
    },
    "University of Pennsylvania": {
        "accept": 6, "satMedian": 1510, "sat25": 1450, "sat75": 1560, "gpaMean": 3.9, "hiddenIvy": False, "ivyLeague": True, "stem": True,
        "merit": "none", "location": "Philadelphia, PA",
        "vibe": "Wharton, Penn Medicine, and interdisciplinary programs in one of America's greatest college cities",
    },
    "Yale University": {
        "accept": 5, "satMedian": 1535, "sat25": 1460, "sat75": 1570, "gpaMean": 3.95, "hiddenIvy": False, "ivyLeague": True, "stem": False,
        "merit": "none", "location": "New Haven, CT",
        "vibe": "Architecture, drama, law, and music in a campus that looks like what college is supposed to look like",
    },
    # ── ACC ───────────────────────────────────────────────────────────────────
    "Boston College": {
        "accept": 19, "satMedian": 1410, "sat25": 1360, "sat75": 1520, "gpaMean": 3.87, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Chestnut Hill, MA",
        "vibe": "Jesuit university with major football culture and strong pre-law, finance, and nursing programs",
    },
    "California, University of, Ber": {
        "accept": 14, "satMedian": 1415, "sat25": 1310, "sat75": 1530, "gpaMean": 3.9, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Berkeley, CA",
        "vibe": "The flagship UC — world-class research, Nobel laureates, and legendary campus activism",
    },
    "Duke University": {
        "accept": 6, "satMedian": 1530, "sat25": 1450, "sat75": 1570, "gpaMean": 3.94, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Durham, NC",
        "vibe": "Elite research university with a powerhouse basketball program and a beautiful Gothic campus",
    },
    "Florida State University": {
        "accept": 25, "satMedian": 1265, "sat25": 1230, "sat75": 1390, "gpaMean": 3.69, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Tallahassee, FL",
        "vibe": "Major public flagship with strong business, film, and performing arts in Florida's capital city",
    },
    "Georgia Institute of Technolog": {
        "accept": 17, "satMedian": 1440, "sat25": 1360, "sat75": 1530, "gpaMean": 3.87, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Atlanta, GA",
        "vibe": "Top-3 public engineering school; demanding but with a world-class alumni network and Atlanta access",
    },
    "Louisville, University of": {
        "accept": 70, "satMedian": 1170, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Louisville, KY",
        "vibe": "Urban research university in Louisville with strong health sciences, business, and engineering",
    },
    "North Carolina State Universit": {
        "accept": 47, "satMedian": 1290, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Raleigh, NC",
        "vibe": "NC's engineering and agriculture flagship in Research Triangle; career-focused and highly connected",
    },
    "North Carolina, University of": {
        "accept": 17, "satMedian": 1340, "sat25": 1300, "sat75": 1490, "gpaMean": 3.89, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Chapel Hill, NC",
        "vibe": "The birthplace of public higher education in America; research powerhouse with a legendary campus",
    },
    "Notre Dame, University of": {
        "accept": 13, "satMedian": 1500, "sat25": 1400, "sat75": 1550, "gpaMean": 3.94, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Notre Dame, IN",
        "vibe": "Elite Catholic university with a massive football identity, strong business, and an unmatched alumni network",
    },
    "Pittsburgh, University of": {
        "accept": 53, "satMedian": 1310, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Pittsburgh, PA",
        "vibe": "Major research university in a revitalized city; strong STEM, business, and health sciences",
    },
    "Southern Methodist University": {
        "accept": 45, "satMedian": 1380, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Dallas, TX",
        "vibe": "Well-heeled Dallas campus with Dedman School of Law and a strong pre-law and business culture",
    },
    "Stanford University": {
        "accept": 4, "satMedian": 1530, "sat25": 1440, "sat75": 1570, "gpaMean": 3.96, "hiddenIvy": False, "stem": True,
        "merit": "none", "moonshot": True, "location": "Stanford, CA",
        "vibe": "Silicon Valley's university — extraordinary in every dimension, and nearly impossible to enter",
    },
    "University of Miami (Florida)": {
        "accept": 23, "satMedian": 1380, "sat25": 1290, "sat75": 1440, "gpaMean": 3.75, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Coral Gables, FL",
        "vibe": "Selective private research university in South Florida; strong music, marine science, and business",
    },
    "VA Tech": {
        "accept": 57, "satMedian": 1305, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Blacksburg, VA",
        "vibe": "Engineering and architecture powerhouse in the Blue Ridge; strong Hokie athletics culture",
    },
    "Virginia, University of": {
        "accept": 20, "satMedian": 1425, "sat25": 1340, "sat75": 1520, "gpaMean": 3.88, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Charlottesville, VA",
        "vibe": "Jefferson's academical village — academically elite public flagship with strong honor culture",
    },
    # ── BIG TEN ───────────────────────────────────────────────────────────────
    "Indiana University": {
        "accept": 82, "satMedian": 1220, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Bloomington, IN",
        "vibe": "Big Ten flagship with a world-class music school, strong business, and a classic college-town feel",
    },
    "Iowa, University of": {
        "accept": 85, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Iowa City, IA",
        "vibe": "Big Ten flagship with the nation's top-ranked creative writing program and strong health sciences",
    },
    "Michigan, University of": {
        "accept": 18, "satMedian": 1445, "sat25": 1340, "sat75": 1530, "gpaMean": 3.89, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Ann Arbor, MI",
        "vibe": "The Michigan Difference — elite public university with a massive endowment and All-American campus",
    },
    "Northwestern University": {
        "accept": 7, "satMedian": 1530, "sat25": 1440, "sat75": 1550, "gpaMean": 3.94, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Evanston, IL",
        "vibe": "The Ivy of the Midwest — elite research university on Lake Michigan with a powerhouse journalism school",
    },
    "Ohio State University": {
        "accept": 52, "satMedian": 1335, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Columbus, OH",
        "vibe": "Massive flagship with world-class research, medicine, and one of America's most passionate fanbases",
    },
    "Pennsylvania State University": {
        "accept": 56, "satMedian": 1245, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "State College, PA",
        "vibe": "Penn State pride is legendary; massive STEM and business programs in a quintessential college town",
    },
    "Purdue University": {
        "accept": 68, "satMedian": 1290, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "West Lafayette, IN",
        "vibe": "Engineering powerhouse that launched more Fortune 500 leaders than almost any school in America",
    },
    "Rutgers University": {
        "accept": 67, "satMedian": 1225, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "New Brunswick, NJ",
        "vibe": "New Jersey's flagship public university with strong pharmacy, business, and public policy programs",
    },
    "University of California, Los": {
        "accept": 9, "satMedian": 1405, "sat25": 1280, "sat75": 1530, "gpaMean": 3.9, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Los Angeles, CA",
        "vibe": "UCLA — elite public research university in LA with Bruin athletics and world-class film and medicine",
    },
    "University of Illinois": {
        "accept": 62, "satMedian": 1310, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Urbana-Champaign, IL",
        "vibe": "Top-5 public engineering school; massive campus with elite CS, business, and architecture programs",
    },
    "University of Minnesota": {
        "accept": 75, "satMedian": 1275, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Minneapolis, MN",
        "vibe": "Big Ten flagship in the Twin Cities with strong medical, law, and business schools",
    },
    "University of Nebraska-Lincoln": {
        "accept": 80, "satMedian": 1190, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Lincoln, NE",
        "vibe": "Nebraska Cornhusker pride meets strong engineering and agriculture in a welcoming college town",
    },
    "University of Southern Califor": {
        "accept": 12, "satMedian": 1455, "sat25": 1360, "sat75": 1530, "gpaMean": 3.87, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Los Angeles, CA",
        "vibe": "USC — elite private research university in LA; elite alumni network, film school, and strong athletics",
    },
    "Wisconsin, University of, Madi": {
        "accept": 57, "satMedian": 1355, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Madison, WI",
        "vibe": "UW-Madison — top public research university on a beautiful lakefront campus; strong in nearly everything",
    },
    # ── SEC ───────────────────────────────────────────────────────────────────
    "Auburn University": {
        "accept": 80, "satMedian": 1255, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Auburn, AL",
        "vibe": "War Eagle — strong engineering, business, and pharmacy programs in a tight-knit college town",
    },
    "Georgia, University of": {
        "accept": 45, "satMedian": 1340, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Athens, GA",
        "vibe": "Georgia Bulldogs — elite public flagship with a legendary campus town and strong business programs",
    },
    "Kentucky, University of": {
        "accept": 95, "satMedian": 1195, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Lexington, KY",
        "vibe": "Wildcat country — strong pharmacy, business, and engineering programs in the Horse Capital of the World",
    },
    "Louisiana State University": {
        "accept": 73, "satMedian": 1215, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Baton Rouge, LA",
        "vibe": "LSU — Tiger Stadium is the loudest place in college football; strong engineering and mass comm",
    },
    "Missouri": {
        "accept": 80, "satMedian": 1225, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Columbia, MO",
        "vibe": "Mizzou — top journalism school in America and a flagship with strong agriculture and business",
    },
    "South Carolina, University of": {
        "accept": 66, "satMedian": 1250, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Columbia, SC",
        "vibe": "Gamecocks flagship with a top-ranked international business school and strong nursing programs",
    },
    "Texas A&M University": {
        "accept": 57, "satMedian": 1245, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "College Station, TX",
        "vibe": "Aggie tradition runs deep — one of America's largest engineering and agriculture programs",
    },
    "University of Alabama": {
        "accept": 80, "satMedian": 1215, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Tuscaloosa, AL",
        "vibe": "Roll Tide — massive flagship with strong engineering, business, and enormous SEC football culture",
    },
    "University of Arkansas": {
        "accept": 79, "satMedian": 1205, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Fayetteville, AR",
        "vibe": "Razorbacks flagship in the Ozarks with a strong Walton School of Business and growing research profile",
    },
    "University of Florida": {
        "accept": 23, "satMedian": 1385, "sat25": 1310, "sat75": 1470, "gpaMean": 3.89, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Gainesville, FL",
        "vibe": "Top-5 public university — UF has elite research, Gator athletics, and an outstanding value proposition",
    },
    "University of Tennessee": {
        "accept": 67, "satMedian": 1235, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Knoxville, TN",
        "vibe": "Vol Nation — Tennessee orange campus on the Tennessee River with strong business and engineering",
    },
    "University of Texas": {
        "accept": 29, "satMedian": 1360, "sat25": 1230, "sat75": 1460, "gpaMean": 3.72, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Austin, TX",
        "vibe": "UT Austin — What starts here changes the world; massive flagship in the live music capital",
    },
    "Vanderbilt University": {
        "accept": 7, "satMedian": 1530, "sat25": 1460, "sat75": 1560, "gpaMean": 3.92, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Nashville, TN",
        "vibe": "Elite private research university in Nashville; Vandy offers Ivy-caliber academics with SEC athletics",
    },
    # ── BIG 12 ────────────────────────────────────────────────────────────────
    "Arizona State University": {
        "accept": 88, "satMedian": 1215, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Tempe, AZ",
        "vibe": "America's largest innovation university; strong engineering, business, and journalism in the Sunbelt",
    },
    "Brigham Young University": {
        "accept": 65, "satMedian": 1265, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Provo, UT",
        "vibe": "LDS flagship with strong accounting, animation, and pre-law — honor code shapes daily campus life",
    },
    "Iowa State University": {
        "accept": 92, "satMedian": 1185, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Ames, IA",
        "vibe": "Cyclones — top-10 design school, strong engineering and agriculture, welcoming Midwest culture",
    },
    "Texas Christian University": {
        "accept": 45, "satMedian": 1290, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Fort Worth, TX",
        "vibe": "TCU — selective Methodist university in Fort Worth with strong business and communications",
    },
    "University of Arizona": {
        "accept": 84, "satMedian": 1195, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Tucson, AZ",
        "vibe": "Wildcats flagship in the Sonoran Desert; strong optics, astronomy, and business programs",
    },
    "University of Cincinnati": {
        "accept": 86, "satMedian": 1215, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Cincinnati, OH",
        "vibe": "Bearcats flagship with a top-10 co-op program; pioneered cooperative education in America",
    },
    "University of Houston": {
        "accept": 65, "satMedian": 1200, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Houston, TX",
        "vibe": "Cougars — urban research university in Houston with strong engineering and entrepreneurship",
    },
    "University of Kansas": {
        "accept": 93, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Lawrence, KS",
        "vibe": "KU — Jayhawks flagship with a legendary pharmacy program and top-ranked journalism school",
    },
    "University of Utah": {
        "accept": 84, "satMedian": 1235, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Salt Lake City, UT",
        "vibe": "Utes — flagship with access to world-class skiing and strong gaming, engineering, and health programs",
    },
    "West Virginia University": {
        "accept": 79, "satMedian": 1170, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Morgantown, WV",
        "vibe": "Mountaineers flagship with strong forensic science, pharmacy, and a tight-knit Appalachian community",
    },
    # ── BIG EAST ──────────────────────────────────────────────────────────────
    "Butler University": {
        "accept": 72, "satMedian": 1225, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Indianapolis, IN",
        "vibe": "Selective private university with a gorgeous campus and strong pharmacy, business, and performing arts",
    },
    "Connecticut, University of": {
        "accept": 56, "satMedian": 1280, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Storrs, CT",
        "vibe": "UConn — flagship with elite nursing, business, and pharmacy programs and a historic basketball tradition",
    },
    "Georgetown University": {
        "accept": 12, "satMedian": 1490, "sat25": 1380, "sat75": 1550, "gpaMean": 3.89, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Washington, DC",
        "vibe": "Elite Jesuit university in DC — politics, pre-law, and international relations on Capitol Hill's doorstep",
    },
    "Providence College": {
        "accept": 50, "satMedian": 1235, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Providence, RI",
        "vibe": "Dominican Catholic college with a required Western civilization core and a passionate Friar basketball culture",
    },
    "Seton Hall University": {
        "accept": 73, "satMedian": 1225, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "South Orange, NJ",
        "vibe": "Catholic university 14 miles from Manhattan; strong diplomacy, nursing, and business programs",
    },
    "Villanova University": {
        "accept": 24, "satMedian": 1415, "sat25": 1310, "sat75": 1470, "gpaMean": 3.8, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Villanova, PA",
        "vibe": "Augustinian university near Philadelphia; highly selective with elite business, nursing, and law pipeline",
    },
    "Xavier University": {
        "accept": 74, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Cincinnati, OH",
        "vibe": "Jesuit university in Cincinnati with strong business, pre-health, and a passionate Musketeer basketball culture",
    },
    # ── AMERICA EAST ──────────────────────────────────────────────────────────
    "Binghamton University": {
        "accept": 43, "satMedian": 1295, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Binghamton, NY",
        "vibe": "SUNY flagship known as the Public Ivy of New York; strong business, engineering, and nursing",
    },
    "Bryant University": {
        "accept": 71, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Smithfield, RI",
        "vibe": "Business-focused university near Providence with a strong actuarial science and data analytics track",
    },
    "Maine, University of": {
        "accept": 89, "satMedian": 1155, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Orono, ME",
        "vibe": "Maine's flagship land-grant with strong forestry, marine science, and engineering in the North Woods",
    },
    "New Hampshire, University of": {
        "accept": 82, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Durham, NH",
        "vibe": "New England's flagship with strong ocean engineering, environmental science, and a vibrant campus culture",
    },
    "New Jersey Institute of Techno": {
        "accept": 68, "satMedian": 1270, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Newark, NJ",
        "vibe": "NJIT — urban tech university with strong CS, architecture, and engineering 20 minutes from NYC",
    },
    "University of Maryland Baltimo": {
        "accept": 52, "satMedian": 1240, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Baltimore, MD",
        "vibe": "UMBC — ranked #1 for transforming undergraduate education; strong STEM and Meyerhoff Scholars program",
    },
    "Vermont, University of": {
        "accept": 63, "satMedian": 1245, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Burlington, VT",
        "vibe": "UVM — New England's outdoor university in Burlington; strong environmental, health, and agriculture",
    },
    "Virginia Military Institute": {
        "accept": 63, "satMedian": 1170, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Lexington, VA",
        "vibe": "VMI — America's oldest state military college; rigorous discipline, camaraderie, and leadership",
    },
    # ── ATLANTIC 10 ───────────────────────────────────────────────────────────
    "Davidson College": {
        "accept": 19, "satMedian": 1400, "sat25": 1350, "sat75": 1490, "gpaMean": 3.83, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Davidson, NC",
        "vibe": "Elite honor-code LAC near Charlotte; consistently tops national rankings and places students at top grad schools",
    },
    "Duquesne University": {
        "accept": 73, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Pittsburgh, PA",
        "vibe": "Spiritan Catholic university on a hilltop in Pittsburgh with strong pharmacy, business, and law pipeline",
    },
    "Fordham University": {
        "accept": 49, "satMedian": 1305, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "New York, NY",
        "vibe": "Jesuit university in the Bronx and Lincoln Center — Rose Hill campus meets the heart of Manhattan",
    },
    "George Mason University": {
        "accept": 87, "satMedian": 1235, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Fairfax, VA",
        "vibe": "Northern Virginia's flagship; tech corridor access, strong CS, policy, and economics near DC",
    },
    "George Washington University": {
        "accept": 43, "satMedian": 1365, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Washington, DC",
        "vibe": "Urban university two blocks from the White House; political science and international affairs culture",
    },
    "La Salle University": {
        "accept": 79, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Philadelphia, PA",
        "vibe": "De La Salle Brothers Catholic university in Philadelphia with strong nursing and business programs",
    },
    "Saint Louis University": {
        "accept": 65, "satMedian": 1265, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "St. Louis, MO",
        "vibe": "Jesuit university with a top aviation program and strong pre-health in the Gateway City",
    },
    "St Bonaventure University": {
        "accept": 79, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "St. Bonaventure, NY",
        "vibe": "Franciscan university in the Southern Tier with a nationally respected journalism program",
    },
    "University of Rhode Island": {
        "accept": 80, "satMedian": 1190, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Kingston, RI",
        "vibe": "URI — strong pharmacy, engineering, and ocean science programs in coastal Rhode Island",
    },
    "University of Richmond": {
        "accept": 28, "satMedian": 1380, "sat25": 1240, "sat75": 1420, "gpaMean": 3.79, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Richmond, VA",
        "vibe": "Selective private university with a beautiful campus; strong business, law pipeline, and generous aid",
    },
    # ── CAA ───────────────────────────────────────────────────────────────────
    "Campbell University": {
        "accept": 61, "satMedian": 1160, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Buies Creek, NC",
        "vibe": "Baptist university in North Carolina with a highly regarded pharmacy and law school",
    },
    "Drexel University": {
        "accept": 77, "satMedian": 1295, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Philadelphia, PA",
        "vibe": "Co-op powerhouse in Philadelphia — Drexel students graduate with up to 18 months of real work experience",
    },
    "Monmouth University": {
        "accept": 85, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "West Long Branch, NJ",
        "vibe": "NJ shore university near NYC with a stunning campus and strong communications and business programs",
    },
    "Northeastern University": {
        "accept": 7, "satMedian": 1490, "sat25": 1440, "sat75": 1550, "gpaMean": 3.84, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Boston, MA",
        "vibe": "Co-op model taken to the extreme — Northeastern's network places students at the world's top employers",
    },
    "Stony Brook University": {
        "accept": 49, "satMedian": 1295, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Stony Brook, NY",
        "vibe": "SUNY's research flagship on Long Island; strong STEM and medical programs with a large international community",
    },
    "Towson University": {
        "accept": 84, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Towson, MD",
        "vibe": "Maryland's largest university by enrollment; strong education, health professions, and business programs",
    },
    "University of North Carolina W": {
        "accept": 73, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Wilmington, NC",
        "vibe": "UNCW — coastal NC campus with outstanding marine biology, film, and business programs",
    },
    "William and Mary": {
        "accept": 33, "satMedian": 1400, "sat25": 1280, "sat75": 1470, "gpaMean": 3.79, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Williamsburg, VA",
        "vibe": "America's second-oldest university and the 'Public Ivy' of Virginia; strong law, business, and history",
    },
    # ── ASUN ──────────────────────────────────────────────────────────────────
    "Bellarmine University": {
        "accept": 62, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Louisville, KY",
        "vibe": "Catholic LAC in Louisville with strong nursing, education, and a student-focused campus community",
    },
    "Delaware": {
        "accept": 75, "satMedian": 1210, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Newark, DE",
        "vibe": "UD — flagship with a top-ranked physical therapy program and strong business and engineering",
    },
    "Florida Atlantic University": {
        "accept": 64, "satMedian": 1150, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Boca Raton, FL",
        "vibe": "Sun Belt university in South Florida with growing research profile and strong engineering programs",
    },
    "Florida Gulf Coast University": {
        "accept": 78, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Fort Myers, FL",
        "vibe": "Young and growing Southwest Florida university with strong business and health sciences programs",
    },
    "Gardner-Webb University": {
        "accept": 75, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Boiling Springs, NC",
        "vibe": "Baptist-affiliated university in the North Carolina foothills with strong nursing and education",
    },
    "Georgia Southern University": {
        "accept": 81, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Statesboro, GA",
        "vibe": "Eagle Nation — growing Sun Belt university with strong business, IT, and health sciences programs",
    },
    "Old Dominion University": {
        "accept": 91, "satMedian": 1110, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Norfolk, VA",
        "vibe": "ODU — Hampton Roads research university with strong engineering, maritime, and health sciences",
    },
    "Queens University of Charlotte": {
        "accept": 70, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Charlotte, NC",
        "vibe": "Presbyterian-affiliated urban university in Charlotte with strong nursing and business programs",
    },
    "Univ North Carolina Asheville": {
        "accept": 79, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Asheville, NC",
        "vibe": "UNCA — liberal arts focus in one of America's most vibrant mountain cities; strong humanities",
    },
    "University of North Florida": {
        "accept": 75, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Jacksonville, FL",
        "vibe": "UNF — coastal Jacksonville campus with strong business, education, and health sciences programs",
    },
    # ── BIG WEST ──────────────────────────────────────────────────────────────
    "Cal State Bakersfield": {
        "accept": 71, "satMedian": 1085, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Bakersfield, CA",
        "vibe": "CSU campus in California's Central Valley with strong nursing, criminal justice, and business",
    },
    "Grand Canyon University": {
        "accept": 73, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Phoenix, AZ",
        "vibe": "Christian university in Phoenix with rapid growth, strong nursing, and a vibrant athletics culture",
    },
    "Seattle U": {
        "accept": 79, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Seattle, WA",
        "vibe": "Jesuit university on Capitol Hill with strong business, nursing, and law pipeline in a world-class city",
    },
    "UC Davis": {
        "accept": 39, "satMedian": 1270, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Davis, CA",
        "vibe": "UC Davis — world's top agriculture and veterinary school; bike-friendly campus near Sacramento",
    },
    "UC San Diego": {
        "accept": 24, "satMedian": 1370, "sat25": 1290, "sat75": 1450, "gpaMean": 3.9, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "La Jolla, CA",
        "vibe": "UCSD — elite research campus with top-3 global ranking in oceanography and strong STEM programs",
    },
    "UC Santa Barbara": {
        "accept": 26, "satMedian": 1330, "sat25": 1270, "sat75": 1440, "gpaMean": 3.83, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Santa Barbara, CA",
        "vibe": "UCSB — stunning ocean campus with Nobel-laureate faculty and a strong surf and research culture",
    },
    "University of Hawaii": {
        "accept": 61, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Honolulu, HI",
        "vibe": "UH Manoa — flagship in paradise; strong marine biology, Asian studies, and education programs",
    },
    "University of San Diego": {
        "accept": 49, "satMedian": 1250, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "San Diego, CA",
        "vibe": "Catholic university on a stunning mesa overlooking the Pacific with strong law and business programs",
    },
    # ── HORIZON LEAGUE ────────────────────────────────────────────────────────
    "Cleveland State University": {
        "accept": 72, "satMedian": 1160, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Cleveland, OH",
        "vibe": "Urban research university in downtown Cleveland with strong engineering, law, and health sciences",
    },
    "Green Bay Phoenix": {
        "accept": 85, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Green Bay, WI",
        "vibe": "UW-Green Bay — interdisciplinary focus and strong environmental science programs in Packer country",
    },
    "IU Indianapolis": {
        "accept": 80, "satMedian": 1155, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Indianapolis, IN",
        "vibe": "IU's urban campus in Indiana's capital; strong health sciences, business, and law programs",
    },
    "Northern Kentucky University": {
        "accept": 86, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Highland Heights, KY",
        "vibe": "Growing metro-Cincinnati university with strong business, nursing, and computer science programs",
    },
    "Oakland University": {
        "accept": 76, "satMedian": 1175, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Rochester, MI",
        "vibe": "Oakland — suburban Detroit university with strong engineering, business, and health sciences programs",
    },
    "University of Wisconsin-Milwau": {
        "accept": 85, "satMedian": 1170, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Milwaukee, WI",
        "vibe": "UWM — urban research campus in Milwaukee with strong engineering, architecture, and nursing programs",
    },
    "Youngstown State University": {
        "accept": 72, "satMedian": 1125, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Youngstown, OH",
        "vibe": "Regional university with strong engineering and health sciences serving the Mahoning Valley",
    },
    # ── GLIAC ─────────────────────────────────────────────────────────────────
    "Augustana University": {
        "accept": 72, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Sioux Falls, SD",
        "vibe": "Lutheran university in the Sioux Falls metro; strong business, education, and nursing programs",
    },
    "Davenport University": {
        "accept": 71, "satMedian": 1100, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Grand Rapids, MI",
        "vibe": "Career-focused university with strong cybersecurity, business, and healthcare administration",
    },
    "Grand Valley State University": {
        "accept": 82, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Allendale, MI",
        "vibe": "GVSU — one of Michigan's fastest-growing universities; strong education, nursing, and business",
    },
    "Lake Superior State University": {
        "accept": 75, "satMedian": 1100, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Sault Ste. Marie, MI",
        "vibe": "Small STEM-focused university on the Canada border with strong robotics and fisheries programs",
    },
    "Northern Michigan University": {
        "accept": 76, "satMedian": 1135, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Marquette, MI",
        "vibe": "NMU — outdoor recreation capital of the UP; strong nursing, education, and culinary programs",
    },
    "Saginaw Valley State Universit": {
        "accept": 82, "satMedian": 1090, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "University Center, MI",
        "vibe": "Regional Michigan university with strong nursing, education, and engineering technology programs",
    },
    "St Cloud State University (M)": {
        "accept": 79, "satMedian": 1130, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "St. Cloud, MN",
        "vibe": "SCSU — one of Minnesota's largest universities; strong aviation, engineering, and business programs",
    },
    "Wayne State University": {
        "accept": 75, "satMedian": 1155, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Detroit, MI",
        "vibe": "Urban research university in Detroit's Midtown; strong medicine, law, and engineering programs",
    },
    # ── MPSF ─────────────────────────────────────────────────────────────────
    "Cal Baptist University": {
        "accept": 77, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Riverside, CA",
        "vibe": "Christian university in the Inland Empire with a growing athletics program and strong nursing",
    },
    "Idaho, University of": {
        "accept": 80, "satMedian": 1175, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Moscow, ID",
        "vibe": "Idaho's flagship land-grant with strong forestry, engineering, and agriculture in a college-town setting",
    },
    "New Mexico State University": {
        "accept": 97, "satMedian": 1100, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Las Cruces, NM",
        "vibe": "NMSU — Aggie engineering and agriculture in the Mesilla Valley near the Rio Grande",
    },
    "Northern Arizona University": {
        "accept": 76, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Flagstaff, AZ",
        "vibe": "Ponderosa pine campus at 7,000 feet; strong nursing, education, and hotel management programs",
    },
    "Northern Colorado, University": {
        "accept": 80, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Greeley, CO",
        "vibe": "UNC Bears — strong education, health sciences, and performing arts in northern Colorado",
    },
    "Pepperdine University": {
        "accept": 36, "satMedian": 1340, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Malibu, CA",
        "vibe": "Christian university on a Pacific Ocean bluff in Malibu; strong law, business, and international programs",
    },
    "US Air Force Academy": {
        "accept": 11, "satMedian": 1340, "sat25": 1240, "sat75": 1410, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Colorado Springs, CO",
        "vibe": "Full scholarship service academy with rigorous STEM curriculum and a guaranteed Air Force career",
    },
    "Univ Texas Rio Grande Valley": {
        "accept": 100, "satMedian": 1025, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Edinburg, TX",
        "vibe": "UTRGV — open-access Hispanic-serving institution along the Rio Grande with strong STEM growth",
    },
    "University of Incarnate Word": {
        "accept": 82, "satMedian": 1090, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "San Antonio, TX",
        "vibe": "Catholic university in San Antonio with strong nursing, optometry, and pharmacy programs",
    },
    "University of the Pacific": {
        "accept": 67, "satMedian": 1215, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Stockton, CA",
        "vibe": "West Coast private with a top pharmacy school, conservatory of music, and accelerated dental program",
    },
    # ── SUMMIT LEAGUE ─────────────────────────────────────────────────────────
    "Eastern Illinois University": {
        "accept": 57, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Charleston, IL",
        "vibe": "Regional Illinois university with strong education, business, and athletic training programs",
    },
    "South Dakota State University": {
        "accept": 81, "satMedian": 1190, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Brookings, SD",
        "vibe": "SDSU — land-grant flagship with strong agriculture, engineering, and pharmacy programs",
    },
    "University of Denver": {
        "accept": 68, "satMedian": 1310, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Denver, CO",
        "vibe": "Private research university in the Mile High City with strong business, law, and international programs",
    },
    "University of Nebraska Omaha": {
        "accept": 80, "satMedian": 1165, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Omaha, NE",
        "vibe": "Urban NU campus in Omaha with strong IT, business, and criminal justice programs",
    },
    "University of South Dakota": {
        "accept": 87, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Vermillion, SD",
        "vibe": "South Dakota's flagship with a top law school, strong health sciences, and a close campus community",
    },
    "University of Southern Indiana": {
        "accept": 80, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Evansville, IN",
        "vibe": "Regional university in southwest Indiana with strong dental hygiene, nursing, and engineering tech",
    },
    "University of St. Thomas MN": {
        "accept": 77, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "St. Paul, MN",
        "vibe": "Catholic university in the Twin Cities; strong business, law, and education with a large alumni network",
    },
    # ── WAC ───────────────────────────────────────────────────────────────────
    "Air Force": {
        "accept": 11, "satMedian": 1340, "sat25": 1240, "sat75": 1410, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Colorado Springs, CO",
        "vibe": "Full scholarship service academy with rigorous STEM curriculum and a guaranteed Air Force career",
    },
    "California Baptist": {
        "accept": 73, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Riverside, CA",
        "vibe": "Christian university in the Inland Empire with a growing athletics program and strong nursing",
    },
    "Grand Canyon": {
        "accept": 73, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Phoenix, AZ",
        "vibe": "Christian university in Phoenix with rapid growth, strong nursing, and a vibrant athletics culture",
    },
    "Idaho": {
        "accept": 80, "satMedian": 1175, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Moscow, ID",
        "vibe": "Idaho's flagship land-grant with strong forestry, engineering, and agriculture in a college-town setting",
    },
    "New Mexico State": {
        "accept": 97, "satMedian": 1100, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Las Cruces, NM",
        "vibe": "NMSU — Aggie engineering and agriculture in the Mesilla Valley near the Rio Grande",
    },
    "Northern Arizona": {
        "accept": 76, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Flagstaff, AZ",
        "vibe": "Ponderosa pine campus at 7,000 feet; strong nursing, education, and hotel management programs",
    },
    "Northern Colorado": {
        "accept": 80, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Greeley, CO",
        "vibe": "UNC Bears — strong education, health sciences, and performing arts in northern Colorado",
    },
    "Seattle": {
        "accept": 79, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Seattle, WA",
        "vibe": "Jesuit university on Capitol Hill with strong business, nursing, and law pipeline in a world-class city",
    },
    "University of Nevada Las Vegas": {
        "accept": 83, "satMedian": 1130, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Las Vegas, NV",
        "vibe": "UNLV — urban research university with top hospitality and hotel management programs in Las Vegas",
    },
    "University of Texas Rio Grande": {
        "accept": 100, "satMedian": 1025, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Edinburg, TX",
        "vibe": "UTRGV — open-access Hispanic-serving institution along the Rio Grande with strong STEM growth",
    },
    "University of Wyoming": {
        "accept": 96, "satMedian": 1155, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Laramie, WY",
        "vibe": "Wyoming's only four-year university; strong engineering, geology, and agriculture at 7,200 feet",
    },
    "Utah Tech University": {
        "accept": 100, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "St. George, UT",
        "vibe": "Open-access university in Utah's red rock country with strong healthcare, business, and technology",
    },
    # ── PCSC ─────────────────────────────────────────────────────────────────
    "Arizona Christian University": {
        "accept": 44, "satMedian": 1085, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Phoenix, AZ",
        "vibe": "Christian liberal arts in metro Phoenix; small and student-focused with strong ministry programs",
    },
    "Azusa Pacific University": {
        "accept": 51, "satMedian": 1155, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Azusa, CA",
        "vibe": "Evangelical Christian university in the San Gabriel Valley with strong nursing and music programs",
    },
    "Biola University": {
        "accept": 60, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "La Mirada, CA",
        "vibe": "Evangelical university near LA with a strong biblical studies core and excellent nursing program",
    },
    "California State University, E": {
        "accept": 72, "satMedian": 1055, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Hayward, CA",
        "vibe": "Cal State East Bay — diverse urban campus in the East Bay with strong business and nursing programs",
    },
    "College of Idaho": {
        "accept": 90, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Caldwell, ID",
        "vibe": "Small Idaho LAC with generous merit aid and a close-knit campus community near Boise",
    },
    "Concordia University Irvine": {
        "accept": 73, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Irvine, CA",
        "vibe": "Lutheran university in Orange County; strong business and education in a Southern California setting",
    },
    "Fresno Pacific University": {
        "accept": 49, "satMedian": 1075, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Fresno, CA",
        "vibe": "Mennonite Christian university in the Central Valley with strong teacher education programs",
    },
    "Ottawa University of Arizona": {
        "accept": 82, "satMedian": 1050, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Surprise, AZ",
        "vibe": "Baptist-affiliated university in the Phoenix metro with small class sizes and career-focused programs",
    },
    "Simpson University": {
        "accept": 53, "satMedian": 1070, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Redding, CA",
        "vibe": "Christian LAC in Northern California with strong biblical studies and teacher education programs",
    },
    "Soka University": {
        "accept": 36, "satMedian": 1240, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Aliso Viejo, CA",
        "vibe": "Unique Buddhist-inspired liberal arts university with a strong global citizenship mission and full scholarships",
    },
    "The Master's University": {
        "accept": 46, "satMedian": 1190, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Santa Clarita, CA",
        "vibe": "Conservative Christian university in the Santa Clarita Valley with a strong biblical studies emphasis",
    },
    "University of Alaska Fairbanks": {
        "accept": 77, "satMedian": 1105, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Fairbanks, AK",
        "vibe": "UAF — world-class arctic research with strong engineering, geophysics, and wildlife biology programs",
    },
    "University of California, Sant": {
        "accept": 47, "satMedian": 1240, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Santa Cruz, CA",
        "vibe": "UC Santa Cruz — redwood campus culture with strong marine science, CS, and social justice programs",
    },
    "Westmont College": {
        "accept": 71, "satMedian": 1290, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Santa Barbara, CA",
        "vibe": "Selective Christian LAC in the Santa Barbara hills; strong academics with evangelical community values",
    },
    # ── PSAC ─────────────────────────────────────────────────────────────────
    "Bloomsburg University": {
        "accept": 77, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Bloomsburg, PA",
        "vibe": "PASSHE university in the Susquehanna Valley with strong nursing, education, and business programs",
    },
    "East Stroudsburg University": {
        "accept": 79, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "East Stroudsburg, PA",
        "vibe": "Poconos campus with strong health, physical education, and tourism and hospitality programs",
    },
    "Gannon University": {
        "accept": 80, "satMedian": 1150, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Erie, PA",
        "vibe": "Catholic university in Erie with strong engineering, PA studies, and health sciences programs",
    },
    "Indiana University of PA": {
        "accept": 72, "satMedian": 1085, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Indiana, PA",
        "vibe": "IUP — western PA regional university with strong education, criminology, and culinary programs",
    },
    "Kutztown University": {
        "accept": 78, "satMedian": 1075, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Kutztown, PA",
        "vibe": "PASSHE campus in Pennsylvania Dutch Country with strong education, art, and communication programs",
    },
    "Lock Haven University": {
        "accept": 82, "satMedian": 1050, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Lock Haven, PA",
        "vibe": "Small river-town Pennsylvania campus with strong health science, education, and outdoor recreation",
    },
    "Millersville University": {
        "accept": 75, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Millersville, PA",
        "vibe": "Lancaster County campus with strong education, meteorology, and social work programs",
    },
    "PennWest Edinboro University": {
        "accept": 84, "satMedian": 1050, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Edinboro, PA",
        "vibe": "Northwest Pennsylvania campus known for strong art, education, and social work programs",
    },
    "PennWest University Clarion": {
        "accept": 82, "satMedian": 1050, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Clarion, PA",
        "vibe": "Rural western Pennsylvania campus with strong communication, education, and business programs",
    },
    "PennWest University,California": {
        "accept": 81, "satMedian": 1050, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "California, PA",
        "vibe": "PASSHE campus along the Monongahela with strong education, sport management, and criminal justice",
    },
    "Shippensburg University of PA": {
        "accept": 81, "satMedian": 1075, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Shippensburg, PA",
        "vibe": "South-central PA campus with strong business, public administration, and criminal justice programs",
    },
    "West Chester University": {
        "accept": 62, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "West Chester, PA",
        "vibe": "Philadelphia suburb campus with strong music, education, nursing, and a large and active student body",
    },
    # ── SAC ───────────────────────────────────────────────────────────────────
    "Carson-Newman University": {
        "accept": 73, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Jefferson City, TN",
        "vibe": "Baptist university in the East Tennessee hills with strong nursing, education, and athletics programs",
    },
    "Catawba College": {
        "accept": 53, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Salisbury, NC",
        "vibe": "Reformed Church-affiliated LAC in the Carolina Piedmont with strong business and education programs",
    },
    "Lenoir-Rhyne University": {
        "accept": 67, "satMedian": 1115, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Hickory, NC",
        "vibe": "Lutheran university in the North Carolina foothills with strong nursing, business, and education",
    },
    "Mars Hill University": {
        "accept": 62, "satMedian": 1075, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Mars Hill, NC",
        "vibe": "Baptist-affiliated mountain university near Asheville with strong education and criminal justice",
    },
    "Wingate University": {
        "accept": 65, "satMedian": 1110, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Wingate, NC",
        "vibe": "Baptist-affiliated university near Charlotte with strong pharmacy, business, and education programs",
    },
    # ── COLORADO COLLEGE (lone remaining school) ──────────────────────────────
    "Bryn Mawr College Owls": {
        "accept": 34, "satMedian": 1385, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Bryn Mawr, PA",
        "vibe": "Elite Seven Sisters college near Philadelphia; rigorous academics and a powerful women's leadership network",
    },
}

# ---------------------------------------------------------------------------
# Runtime data stores
# ---------------------------------------------------------------------------
# ── PRIMARY UNIVERSE (source of truth) ──────────────────────────────────────
# EXPLORE_SCHOOLS is the canonical 324-school swimming universe.
# Built at startup from output/lane4_snapshot_compatible.csv.
# Every endpoint that needs a school list starts here — never from TEAMS_LIST.
EXPLORE_SCHOOLS = []   # [{school, conference, conf_tier_short, meta, hasSwimData, …}]

# ── ENRICHMENT SOURCES (read-only, merged into EXPLORE_SCHOOLS at startup) ──
# lane4_swim_model.xlsx provides:
#   BENCHMARKS  — event × conference benchmarks (needed for swim scoring)
#   TEAMS/TEAMS_LIST — PSF + tier values for the 68 schools in both sources
#   CONFERENCES — used only by score_one_school (manual calculator)
# These dictionaries are never used as a school universe.  They enrich values.
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'lane4_swim_model.xlsx')
BENCHMARKS = {}    # "Conference|Event" -> {first, eighth, sixteenth, sec_per_place}
TEAMS = {}         # "Conference|School" -> enrichment record (PSF, tier, finish)
TEAMS_LIST = []    # flat list of TEAMS records — enrichment only, not the universe
CONFERENCES = {}   # conference name -> sorted list of canonical school names
NORMALIZATION_LOG = []  # records every name-normalization applied during load

# ── Out-of-universe well-known schools ───────────────────────────────────────
# Metadata for commonly searched schools outside our scored pool.
# Keys are the canonical display names (title-cased as users would type them).
# Fields match SCHOOL_META conventions; ivyLeague=True adds the Ivy badge.
OOU_SCHOOL_META = {
    # Ivy League
    "Harvard":      {"accept": 4,  "satMedian": 1580, "ivyLeague": True,  "merit": "none",     "location": "Cambridge, MA",     "vibe": "The most storied name in higher education — ultra-selective by any measure."},
    "Yale":         {"accept": 5,  "satMedian": 1570, "ivyLeague": True,  "merit": "none",     "location": "New Haven, CT",      "vibe": "World-class academics, drama, and debate — near-impossible odds."},
    "Princeton":    {"accept": 4,  "satMedian": 1580, "ivyLeague": True,  "merit": "none",     "location": "Princeton, NJ",      "vibe": "Legendary campus, no-loan financial aid, and 4% acceptance."},
    "Columbia":     {"accept": 4,  "satMedian": 1560, "ivyLeague": True,  "merit": "none",     "location": "New York, NY",       "vibe": "Ivy in the heart of Manhattan — urban energy meets academic prestige."},
    "Brown":        {"accept": 6,  "satMedian": 1555, "ivyLeague": True,  "merit": "none",     "location": "Providence, RI",     "vibe": "Open Curriculum gives students unusual freedom to design their education."},
    "Dartmouth":    {"accept": 6,  "satMedian": 1560, "ivyLeague": True,  "merit": "none",     "location": "Hanover, NH",        "vibe": "Tight-knit Ivy with a strong outdoors culture and fierce alumni loyalty."},
    "Cornell":      {"accept": 9,  "satMedian": 1510, "ivyLeague": True,  "merit": "none",     "location": "Ithaca, NY",         "vibe": "The most accessible Ivy — broad programs, engineering powerhouse."},
    "Penn":         {"accept": 7,  "satMedian": 1535, "ivyLeague": True,  "merit": "none",     "location": "Philadelphia, PA",   "vibe": "Wharton business, Penn Medicine, and strong interdisciplinary programs."},
    "UPenn":        {"accept": 7,  "satMedian": 1535, "ivyLeague": True,  "merit": "none",     "location": "Philadelphia, PA",   "vibe": "Wharton business, Penn Medicine, and strong interdisciplinary programs."},
    # Elite non-Ivy
    "MIT":          {"accept": 4,  "satMedian": 1580, "stem": True,       "merit": "none",     "location": "Cambridge, MA",      "vibe": "Global STEM leader — demanding, collaborative, and transformative."},
    "Stanford":     {"accept": 4,  "satMedian": 1560, "stem": True,       "merit": "none",     "location": "Stanford, CA",       "vibe": "Silicon Valley's university — entrepreneurship and research at scale."},
    "Duke":         {"accept": 6,  "satMedian": 1540, "merit": "none",    "location": "Durham, NC",       "vibe": "Elite academics meets ACC athletics in a beautiful residential campus."},
    "Northwestern": {"accept": 7,  "satMedian": 1530, "merit": "none",    "location": "Evanston, IL",     "vibe": "Quarter system, Big Ten athletics, and one of the strongest journalism schools."},
    "Georgetown":   {"accept": 12, "satMedian": 1470, "merit": "none",    "location": "Washington, DC",   "vibe": "Jesuit traditions, global affairs, and unmatched access to DC institutions."},
    "Notre Dame":   {"accept": 13, "satMedian": 1480, "merit": "none",    "location": "Notre Dame, IN",   "vibe": "Catholic identity, storied football, and a fiercely loyal alumni network."},
    "Vanderbilt":   {"accept": 7,  "satMedian": 1540, "merit": "moderate","location": "Nashville, TN",    "vibe": "Southern hospitality meets elite academics in a vibrant music city."},
    "Emory":        {"accept": 11, "satMedian": 1480, "hiddenIvy": True,  "merit": "moderate", "location": "Atlanta, GA",        "vibe": "Hidden Ivy with top pre-med programs and CDC proximity."},
    "Tufts":        {"accept": 11, "satMedian": 1490, "hiddenIvy": True,  "merit": "none",     "location": "Medford, MA",        "vibe": "Hidden Ivy bridging liberal arts and research, just outside Boston."},
    "Wake Forest":  {"accept": 21, "satMedian": 1400, "sat25": 1290, "sat75": 1480, "gpaMean": 3.82, "merit": "moderate","location": "Winston-Salem, NC","vibe": "Small research university with a pro-human motto and strong business school."},
    "Boston College": {"accept": 15,"satMedian": 1430,"sat25": 1360,"sat75": 1520,"gpaMean": 3.87,"merit": "none",    "location": "Chestnut Hill, MA","vibe": "Jesuit institution with strong business, nursing, and law programs."},
    "Boston University": {"accept": 19,"satMedian": 1400,"sat25": 1350,"sat75": 1500,"gpaMean": 3.73,"merit": "moderate","location": "Boston, MA",   "vibe": "Large urban research university with 300+ programs along the Charles River."},
    "Northeastern": {"accept": 7,  "satMedian": 1490,"sat25": 1440,"sat75": 1550,"gpaMean": 3.84,"stem": True,       "merit": "moderate", "location": "Boston, MA",         "vibe": "Co-op powerhouse — students graduate with up to two years of work experience."},
    "Rice":         {"accept": 9,  "satMedian": 1540, "sat25": 1490, "sat75": 1570, "gpaMean": 3.92, "stem": True,       "merit": "none",     "location": "Houston, TX",        "vibe": "Tiny but mighty — residential college system and elite engineering."},
    "University of Chicago": {"accept": 7,"satMedian": 1560,"hiddenIvy": True,"merit": "none","location": "Chicago, IL",         "vibe": "Rigorously intellectual — the life of the mind in the heart of Chicago."},
    "UChicago":     {"accept": 7,  "satMedian": 1560, "hiddenIvy": True,  "merit": "none",     "location": "Chicago, IL",        "vibe": "Rigorously intellectual — the life of the mind in the heart of Chicago."},
    "Carnegie Mellon": {"accept": 11,"satMedian": 1530,"stem": True,      "merit": "none",     "location": "Pittsburgh, PA",     "vibe": "CS and engineering titan with one of the world's top drama programs."},
    "WashU":        {"accept": 12, "satMedian": 1530, "hiddenIvy": True,  "merit": "moderate", "location": "St. Louis, MO",      "vibe": "Washington University — Hidden Ivy with generous merit aid and strong med school."},
    "Washington University": {"accept": 12,"satMedian": 1530,"hiddenIvy": True,"merit": "moderate","location": "St. Louis, MO",  "vibe": "Hidden Ivy with generous merit aid and a strong medical school."},
    "Bowdoin":      {"accept": 9,  "satMedian": 1490, "hiddenIvy": True,  "merit": "none",     "location": "Brunswick, ME",      "vibe": "Stunning Maine campus, need-blind admissions, and a powerhouse swim tradition."},
    "Middlebury":   {"accept": 13, "satMedian": 1430, "hiddenIvy": True,  "merit": "none",     "location": "Middlebury, VT",     "vibe": "World-renowned language programs and a stunning Vermont campus."},
    "Colby":        {"accept": 10, "satMedian": 1420, "hiddenIvy": True,  "merit": "none",     "location": "Waterville, ME",     "vibe": "NESCAC liberal arts with need-blind admissions and strong athletics."},
    "Colgate":      {"accept": 21, "satMedian": 1400, "hiddenIvy": True,  "merit": "none",     "location": "Hamilton, NY",       "vibe": "Strong academics in a tight-knit upstate NY community."},
}

# Case-insensitive lookup helper
def _oou_lookup(name: str) -> dict | None:
    name_l = name.lower().strip()
    for k, v in OOU_SCHOOL_META.items():
        if k.lower() == name_l:
            return v
    # partial: e.g. "Harvard University" → "Harvard"
    for k, v in OOU_SCHOOL_META.items():
        if k.lower() in name_l or name_l in k.lower():
            return v
    return None

# ── 2026 snapshot tier enrichment ────────────────────────────────────────────
# Abbreviated Excel names → full snapshot names (UAA uses short names in the Excel)
_UAA_SHORT = {
    'emory':          'emory university',
    'nyu':            'new york university',
    'chicago':        'university of chicago',
    'washingtonmo':   'washington university st louis',
    'carnegiemellon': 'carnegie mellon university',
    'casewestern':    'case western reserve universit',
    'rochester':      'university of rochester',
    'brandeis':       'brandeis university',
}

def load_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Sheet1: BENCHMARKS ──────────────────────────────────────────────────
    ws = wb['Sheet1']
    h1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col1 = {v: i + 1 for i, v in enumerate(h1) if v}

    for r in range(2, ws.max_row + 1):
        conf  = ws.cell(r, col1['Conference']).value
        event = ws.cell(r, col1['Event']).value
        if not conf or not event:
            continue
        BENCHMARKS[f"{conf}|{event}"] = {
            'first':         _float(ws.cell(r, col1['1st_sec']).value),
            'eighth':        _float(ws.cell(r, col1['8th_sec']).value),
            'sixteenth':     _float(ws.cell(r, col1['16th_sec']).value),
            'sec_per_place': _float(ws.cell(r, col1['Sec_per_place']).value),
        }

    # ── Team_Tiers: TEAMS ────────────────────────────────────────────────────
    ws2 = wb['Team_Tiers']
    h2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    col2 = {}
    for i, v in enumerate(h2):
        if v and v not in col2:   # first-occurrence wins — sheet has duplicate headers at col 14
            col2[v] = i + 1

    for r in range(2, ws2.max_row + 1):
        conf   = ws2.cell(r, col2['Conference']).value
        raw    = ws2.cell(r, col2['Team']).value
        psf    = ws2.cell(r, col2['PSF']).value
        tier   = ws2.cell(r, col2['Tier']).value
        finish = ws2.cell(r, col2['Finish']).value
        pts    = ws2.cell(r, col2['MenPoints']).value
        if not conf or not raw:
            continue

        # Apply name normalization
        normalized = False
        canonical = raw
        if raw in TEAM_NAME_MAP:
            canonical, reason = TEAM_NAME_MAP[raw]
            normalized = True
            NORMALIZATION_LOG.append({
                'raw': raw, 'canonical': canonical,
                'reason': reason, 'conference': conf, 'finish': finish,
            })

        team_rec = {
            'conference':  conf,
            'school':      canonical,
            'raw_name':    raw,
            'psf':         _float(psf) if psf is not None else 1.0,
            'tier':        tier or '',
            'finish':      finish,
            'men_points':  _float(pts),
            'normalized':  normalized,
        }
        key = f"{conf}|{canonical}"
        TEAMS[key] = team_rec
        TEAMS_LIST.append(team_rec)

        if conf not in CONFERENCES:
            CONFERENCES[conf] = []
        if canonical not in CONFERENCES[conf]:
            CONFERENCES[conf].append(canonical)

    # Sort teams within each conference by finish position
    for conf in CONFERENCES:
        CONFERENCES[conf].sort(
            key=lambda s: TEAMS.get(f"{conf}|{s}", {}).get('finish') or 99
        )

    _supplement_from_csv()
    _load_conf_tier_lookup()


def _supplement_from_csv():
    """
    Extends BENCHMARKS and TEAMS_LIST from CSV outputs for all conferences
    not already covered by the Excel model file.

    - output/all_event_anchors.csv  → adds benchmark rows to BENCHMARKS
    - output/lane4_snapshot_compatible.csv → adds team_recs to TEAMS_LIST

    Excel data always takes priority; this function never overwrites existing keys.
    Must be called after load_data() Excel loading and before _load_conf_tier_lookup().
    """
    import csv as _csv

    excel_confs = set(k.split('|')[0] for k in BENCHMARKS)

    # ── 1. Supplement BENCHMARKS from all_event_anchors.csv ─────────────────
    anchors_path = os.path.join(os.path.dirname(__file__), 'output', 'all_event_anchors.csv')
    if os.path.exists(anchors_path):
        new_bench = 0
        with open(anchors_path, newline='', encoding='utf-8') as f:
            for row in _csv.DictReader(f):
                conf  = (row.get('Conference') or '').strip()
                event = (row.get('Event') or '').strip()
                if not conf or not event:
                    continue
                key = f"{conf}|{event}"
                if key in BENCHMARKS:
                    continue  # Excel entry takes priority
                first  = _float(row.get('1st_seconds'))
                eighth = _float(row.get('8th_seconds'))
                sixteenth = _float(row.get('16th_seconds'))
                spp   = _float(row.get('Sec_per_place'))
                if first is None and eighth is None:
                    continue  # no usable benchmark data
                BENCHMARKS[key] = {
                    'first':         first,
                    'eighth':        eighth,
                    'sixteenth':     sixteenth,
                    'sec_per_place': spp,
                }
                new_bench += 1
        print(f"[supplement] Added {new_bench} benchmark rows from all_event_anchors.csv")

    # ── 2. Supplement TEAMS_LIST from lane4_snapshot_compatible.csv ─────────
    snap_path = os.path.join(os.path.dirname(__file__), 'output', 'lane4_snapshot_compatible.csv')
    if os.path.exists(snap_path):
        excel_team_keys = set(TEAMS.keys())
        new_teams = 0
        with open(snap_path, newline='', encoding='utf-8') as f:
            for row in _csv.DictReader(f):
                if (row.get('gender') or '').lower() != 'men':
                    continue
                conf = (row.get('Conference') or '').strip()
                raw  = (row.get('Team') or '').strip()
                if not conf or not raw or conf == 'Unknown':
                    continue  # 'Unknown' conference has no benchmarks — useless and harmful
                # Apply name normalization
                canonical = raw
                if raw in TEAM_NAME_MAP:
                    canonical, _ = TEAM_NAME_MAP[raw]
                key = f"{conf}|{canonical}"
                if key in excel_team_keys:
                    continue  # Excel entry takes priority
                try:
                    finish = int(row.get('Finish') or 0) or None
                except (ValueError, TypeError):
                    finish = None
                team_rec = {
                    'conference':      conf,
                    'school':          canonical,
                    'raw_name':        raw,
                    'psf':             _float(row.get('PSF')) or 1.0,
                    'tier':            row.get('Tier') or '',
                    'finish':          finish,
                    'men_points':      _float(row.get('MenPoints')),
                    'normalized':      raw != canonical,
                    'conf_tier_short': row.get('tier_short') or '',
                    'conf_tier':       row.get('final_tier') or '',
                    'conf_finish_2026': finish,
                    'conf_score_2026': row.get('MenPoints') or '',
                    'conf_power_class': row.get('PowerClass') or '',
                }
                TEAMS[key] = team_rec
                TEAMS_LIST.append(team_rec)
                excel_team_keys.add(key)
                if conf not in CONFERENCES:
                    CONFERENCES[conf] = []
                if canonical not in CONFERENCES[conf]:
                    CONFERENCES[conf].append(canonical)
                new_teams += 1
        print(f"[supplement] Added {new_teams} team records from snapshot CSV")


def _norm_key(s):
    """Lowercase, strip all non-alphanumeric for fuzzy matching."""
    import re
    return re.sub(r'[^a-z0-9]', '', s.lower().strip()) if s else ''


def _load_conf_tier_lookup():
    """
    Reads output/lane4_snapshot_compatible.csv and enriches each team_rec in
    TEAMS_LIST with 2026 conference championship data:
      conf_tier_short  — '1A' / '1B' / '2' / '3' / '4'  (empty if not in snapshot)
      conf_tier        — 'Tier 1A' … 'Tier 4'
      conf_finish_2026 — integer conference finish rank
      conf_score_2026  — team score at the 2026 championship
      conf_power_class — 'Super Powerhouse' / 'Powerhouse' / ''
    """
    import csv as _csv
    snap_path = os.path.join(os.path.dirname(__file__), 'output', 'lane4_snapshot_compatible.csv')
    if not os.path.exists(snap_path):
        return

    # Build lookup: norm_conf|norm_team → row  (men's rows preferred)
    snap_rows = []
    with open(snap_path, newline='', encoding='utf-8') as f:
        snap_rows = list(_csv.DictReader(f))

    exact = {}   # norm(conf)|norm(team) → row
    by_team = {} # norm(team) → list of rows
    for row in snap_rows:
        if row.get('gender', '').lower() != 'men':
            continue
        ck = _norm_key(row['Conference']) + '|' + _norm_key(row['Team'])
        exact[ck] = row
        by_team.setdefault(_norm_key(row['Team']), []).append(row)

    def _find(conf, school):
        # 1. exact conf|team match
        ck = _norm_key(conf) + '|' + _norm_key(school)
        if ck in exact:
            return exact[ck]
        # 2. UAA abbreviation table
        norm_school = _norm_key(school)
        full_name = _UAA_SHORT.get(norm_school)
        if full_name:
            rows = by_team.get(_norm_key(full_name), [])
            if rows:
                return rows[0]
        # 3. team-name match, same conference
        candidates = by_team.get(norm_school, [])
        same_conf = [r for r in candidates if _norm_key(r['Conference']) == _norm_key(conf)]
        if same_conf:
            return same_conf[0]
        # 4. team-name match, any conference (only if unique)
        if len(candidates) == 1:
            return candidates[0]
        return None

    for team_rec in TEAMS_LIST:
        hit = _find(team_rec['conference'], team_rec['school'])
        # Fallback: try the original raw Excel name (before TEAM_NAME_MAP normalization)
        if not hit and team_rec.get('raw_name') and team_rec['raw_name'] != team_rec['school']:
            hit = _find(team_rec['conference'], team_rec['raw_name'])
        if hit:
            try:
                finish_2026 = int(hit['Finish'])
            except (ValueError, TypeError):
                finish_2026 = None
            team_rec['conf_tier_short']  = hit.get('tier_short', '')
            team_rec['conf_tier']        = hit.get('final_tier', '')
            team_rec['conf_finish_2026'] = finish_2026
            team_rec['conf_score_2026']  = hit.get('MenPoints', '')
            team_rec['conf_power_class'] = hit.get('PowerClass', '')
        else:
            team_rec['conf_tier_short']  = ''
            team_rec['conf_tier']        = ''
            team_rec['conf_finish_2026'] = None
            team_rec['conf_score_2026']  = ''
            team_rec['conf_power_class'] = ''

    _build_explore_schools()


def _build_explore_schools():
    """
    Build EXPLORE_SCHOOLS — one record per unique school across the full
    2026 championship snapshot (~324 schools).  Every school gets the SAME
    object shape.  SCHOOL_META is an enrichment source; its absence never
    changes the code path — only the richness of the values.
    """
    import csv as _csv
    from collections import defaultdict

    EXPLORE_SCHOOLS.clear()

    snap_path = os.path.join(os.path.dirname(__file__), 'output', 'lane4_snapshot_compatible.csv')
    if not os.path.exists(snap_path):
        return

    snap_rows = []
    with open(snap_path, newline='', encoding='utf-8') as f:
        snap_rows = list(_csv.DictReader(f))

    # Build per-school, per-gender rows
    by_school = defaultdict(dict)   # school_name → {'men': row, 'women': row}
    for row in snap_rows:
        gender = row.get('gender', '').lower()
        school = row.get('Team', '').strip()
        if school in TEAM_NAME_MAP:
            school, _ = TEAM_NAME_MAP[school]   # normalize PDF team names to canonical
        if school and gender:
            by_school[school][gender] = row

    # Modeled lookup by norm of canonical name AND raw name.
    # Also add _UAA_SHORT reverse mapping so snapshot full-names (e.g. "Emory University")
    # correctly resolve to their abbreviated TEAMS_LIST entries (e.g. "Emory").
    modeled_by = {}
    for tr in TEAMS_LIST:
        modeled_by[_norm_key(tr['school'])] = tr
        raw = tr.get('raw_name', '')
        if raw:
            modeled_by[_norm_key(raw)] = tr
    # Reverse UAA_SHORT: short_norm → full_name; add full_name → team_rec
    for short_norm, full_name in _UAA_SHORT.items():
        if short_norm in modeled_by:
            modeled_by[_norm_key(full_name)] = modeled_by[short_norm]

    def _si(v):
        try:   return int(v)
        except: return None

    def _sf(v):
        try:   return float(v)
        except: return None

    tier_order = {'1A': 0, '1B': 1, '2': 2, '3': 3, '4': 4, '': 5}

    for school, gmap in by_school.items():
        men_row   = gmap.get('men')
        women_row = gmap.get('women')
        primary   = men_row or women_row
        if not primary:
            continue

        tr = modeled_by.get(_norm_key(school))

        men_ts   = men_row.get('tier_short', '')   if men_row   else ''
        women_ts = women_row.get('tier_short', '') if women_row else ''
        ts       = men_ts or women_ts

        raw_conf = primary.get('Conference', '')
        # If the snapshot CSV recorded 'Unknown' (unrecognised PDF conference),
        # fall back to the team_rec's conference so scoring and display use the
        # correct real conference name.
        display_conf = (tr['conference']
                        if tr and raw_conf in ('Unknown', '', None)
                        else raw_conf)

        entry = {
            'school':            school,
            'conference':        display_conf,
            'conf_tier_short':   ts,
            'conf_tier':         (men_row or women_row).get('final_tier', ''),
            'conf_power_class':  (men_row or women_row).get('PowerClass', ''),
            'men_finish_2026':   _si(men_row['Finish'])          if men_row   else None,
            'women_finish_2026': _si(women_row['Finish'])        if women_row else None,
            'men_score_2026':    _sf(men_row.get('MenPoints',''))   if men_row   else None,
            'women_score_2026':  _sf(women_row.get('MenPoints','')) if women_row else None,
            'men_tier_short':    men_ts,
            'women_tier_short':  women_ts,
            'gender_coverage':   sorted(gmap.keys()),
        }

        # Unified meta merge — SCHOOL_META is an enrichment source, not a gate.
        # Try direct school name first, then canonical team name.
        meta_raw = SCHOOL_META.get(school) or (SCHOOL_META.get(tr['school']) if tr else None) or {}
        entry['meta'] = {
            'accept':    meta_raw.get('accept'),
            'satMedian': meta_raw.get('satMedian'),
            'location':  meta_raw.get('location', ''),
            'hiddenIvy': meta_raw.get('hiddenIvy', False),
            'ivyLeague': meta_raw.get('ivyLeague', False),
            'stem':      meta_raw.get('stem', False),
            'merit':     meta_raw.get('merit', ''),
            'vibe':      meta_raw.get('vibe', ''),
        }
        if meta_raw.get('moonshot'):
            entry['meta']['moonshot'] = True

        if tr:
            entry['row_type']    = 'modeled_school'
            entry['psf']         = tr.get('psf', 1.0)
            entry['tier']        = tr.get('tier', '')
            entry['hasSwimData'] = True
            entry['_team_rec']   = tr   # stored so build_school_universe() re-uses the match
        else:
            entry['row_type']    = 'snapshot_only'
            entry['hasSwimData'] = False
            entry['_team_rec']   = None

        EXPLORE_SCHOOLS.append(entry)

    # Sort: tier (1A first), then men's finish, then school name
    EXPLORE_SCHOOLS.sort(key=lambda s: (
        tier_order.get(s.get('conf_tier_short', ''), 5),
        s.get('men_finish_2026') or s.get('women_finish_2026') or 99,
        s['school'],
    ))


def _float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

load_data()
_init_db()

# ---------------------------------------------------------------------------
# Scoring engine — all formulas from Swimmer_Calcs (workbook authoritative)
#
# Layer architecture (clean separation for later admissions layer):
#
#   SWIM LAYER  ──  _score_event()  →  EventScore
#                   _score_school_swim()  →  SwimResult
#
#   ADMISSION LAYER  ──  admission_chance(school, sat, gpa, adj_tier, psf)  →  AdmissionResult
#                        (takes SwimResult outputs + academic inputs; returns {label, color, …})
#
#   FULL PIPELINE  ──  build_school_universe(times, sat, gpa)  →  [SchoolResult …]
#                      Starts from EXPLORE_SCHOOLS (324-school snapshot universe).
#                      Merges swim scoring + SCHOOL_META + admission into one uniform shape.
#
# Output field names follow OUTPUT_SCHEMA.md exactly:
#   EventScore:  { event, sec, place, pts }   (+ expPts, confidence, placeLabel for tracing)
#   SchoolResult: { school, conference, tier, psf, rawPts, adjPts, adjTier,
#                   top3, allEvents, admission, meta, normalized, rawName }
# ---------------------------------------------------------------------------

ALL_EVENTS = [
    '50 Free', '100 Free', '200 Free', '500 Free',
    '1000 Free', '1650 Free',
    '100 Back', '200 Back',
    '100 Breast', '200 Breast',
    '100 Fly', '200 Fly',
    '200 IM', '400 IM',
    '50 Breast (Relay Split)',
]

# ── Primitives ─────────────────────────────────────────────────────────────

def parse_time(s):
    """'M:SS.ss' or 'SS.ss' → decimal seconds. Returns None if invalid/missing."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        if ':' in s:
            m, sec = s.split(':', 1)
            return float(m) * 60 + float(sec)
        return float(s)
    except ValueError:
        return None

def estimate_place(sec, b):
    """
    3-zone linear interpolation — workbook formula (Swimmer_Calcs):

      Zone 1  sec <= first      → 1.0          (capped; workbook IF behaviour)
      Zone 2  sec <= eighth     → 1 + (sec-1st)/(8th-1st) * 7
      Zone 3  sec <= sixteenth  → 8 + (sec-8th)/(16th-8th) * 8
      Zone 4  sec > sixteenth   → 16 + (sec-16th) / secPerPlace

    Returns a continuous float. No upper ceiling.
    """
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']
    spp       = b['sec_per_place'] or 1.0

    if sec <= first:
        return 1.0
    if sec <= eighth:
        return 1.0 + (sec - first)  / ((eighth    - first)    or 1.0) * 7
    if sec <= sixteenth:
        return 8.0 + (sec - eighth) / ((sixteenth  - eighth)   or 1.0) * 8
    return 16.0 + (sec - sixteenth) / spp

def exp_points(place):
    """MAX(0, MIN(20, 21−place)) — workbook formula. Continuous float."""
    return max(0.0, min(20.0, 21.0 - place))

def confidence_weight(place):
    """
    Bubble-zone confidence discount — from Swimmer_Calcs.
    Full weight for A/B finalists; discounted for bubble; zero below 16th.
    """
    if place <= 12: return 1.0
    if place <= 14: return 0.85
    if place <= 16: return 0.65
    return 0.0

def place_label(place):
    """Human-readable place outcome — OUTPUT_SCHEMA thresholds."""
    if place <= 1.5:  return 'Contender'
    if place <= 3.5:  return '🏅 Podium'
    if place <= 8.5:  return 'A Final'
    if place <= 16.5: return 'B Final'
    if place <= 20:   return 'Bubble'
    return 'Out of range'

def tier_label(pts):
    """
    Swim tier from adjPts — workbook thresholds (authoritative over spec).
    Called with rawPts for display and adjPts for the canonical tier.
    """
    if pts < 1:   return 'Moonshot'
    if pts < 4:   return 'Reach'
    if pts < 10:  return 'Recruitable'
    if pts < 18:  return 'Priority Recruit'
    if pts < 35:  return 'Top Recruit'
    if pts < 50:  return 'Conference Star'
    return 'High-Point Contender'

# ── Swim layer ──────────────────────────────────────────────────────────────

def _score_event(event, time_str, conf):
    """
    Score one event for one conference.
    Returns an EventScore dict or None if no benchmark / no valid time.

    EventScore (OUTPUT_SCHEMA):
      event      — event name
      sec        — time in decimal seconds
      place      — estimated decimal finish place
      pts        — confidence-weighted points (workbook: expPts × confidence)
                   NOTE: spec defines pts as integer from placeToPoints() lookup;
                   workbook uses continuous formula — workbook is authoritative.
      expPts     — points before confidence discount (for tracing)
      confidence — confidence multiplier applied (for tracing)
      placeLabel — human-readable outcome
    """
    sec = parse_time(time_str)
    if sec is None:
        return None

    bench = BENCHMARKS.get(f"{conf}|{event}")
    if bench is None:
        return None   # event not benchmarked in this conference

    place  = estimate_place(sec, bench)
    exp_pt = exp_points(place)
    cw     = confidence_weight(place)
    pts    = exp_pt * cw               # workbook weighted value → spec's `pts` field

    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),   # OUTPUT_SCHEMA field name
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': place_label(place),
    }

def _score_school_swim(team_rec, times):
    """
    Pure swim-value layer for one school.

    Input:  team_rec (from TEAMS_LIST), times dict from swimmer profile
    Output: SwimResult dict, or None if the swimmer has no scorable events here.

    SwimResult fields (OUTPUT_SCHEMA — swim-layer subset):
      school, conference, finish, tier, psf
      rawPts   — sum of top-4 pts values
      adjPts   — rawPts × psf
      adjTier  — tier label from adjPts
      top3     — up to 4 highest-scoring EventScore objects (key kept for compat)
      allEvents — all scored EventScore objects, sorted pts desc
      hasDepth  — True if swimmer has 4+ events with pts > 0
      normalized, rawName — provenance flags
    """
    conf   = team_rec['conference']
    school = team_rec['school']
    psf    = team_rec['psf']

    all_events = []
    for event, time_str in times.items():
        es = _score_event(event, time_str, conf)
        if es is not None:
            all_events.append(es)

    # Sort by pts descending; top-4 drive rawPts
    all_events.sort(key=lambda e: e['pts'], reverse=True)
    top4    = all_events[:4]
    raw_pts = round(sum(e['pts'] for e in top4), 2)

    if raw_pts == 0:
        return None   # zero-score guardrail — school excluded from results

    adj_pts  = round(raw_pts * psf, 2)
    adj_tier = tier_label(adj_pts)
    has_depth = sum(1 for e in all_events if e['pts'] > 0) >= 4

    return {
        'school':          school,
        'conference':      conf,
        'division':        CONF_DIVISION.get(conf, 'D3'),
        'finish':          team_rec['finish'],
        'tier':            team_rec['tier'],
        'psf':             psf,
        'rawPts':          raw_pts,
        'adjPts':          adj_pts,
        'adjTier':         adj_tier,
        'top3':            top4,   # key kept for downstream compat; now holds top 4
        'hasDepth':        has_depth,
        'allEvents':       all_events,
        'normalized':      team_rec['normalized'],
        'rawName':         team_rec['raw_name'],
        'confTierShort':   team_rec.get('conf_tier_short', ''),
        'confTier':        team_rec.get('conf_tier', ''),
        'confFinish2026':  team_rec.get('conf_finish_2026'),
        'confScore2026':   team_rec.get('conf_score_2026', ''),
        'confPowerClass':  team_rec.get('conf_power_class', ''),
    }

# ── Admission layer v2 — matrix model ───────────────────────────────────────
#
# Labels (6): Very Strong Chance > Strong Chance > Realistic Shot >
#             Possible > Major Reach > Moonshot
# Academic band (0-4) × Swim support band (0-4) → base label → guardrails.
# ---------------------------------------------------------------------------

_LABEL_ORDER = {
    'Moonshot': 0, 'Major Reach': 1, 'Possible': 2,
    'Realistic Shot': 3, 'Strong Chance': 4, 'Very Strong Chance': 5,
}

_LABEL_COLORS = {
    'Very Strong Chance':       '#059669',
    'Strong Chance':            '#10B981',
    'Realistic Shot':           '#2563EB',
    'Possible':                 '#3B82F6',
    'Major Reach':              '#F59E0B',
    'Moonshot':                 '#6B7280',
    'Moonshot — Apply for Fun': '#6B7280',
    'Unknown':                  '#94A3B8',
}

def _oou_admission(oou_meta: dict, sat: int | None, gpa: float | None) -> dict:
    """Compute a simplified admission label for out-of-universe schools."""
    accept    = oou_meta.get('accept', 50)
    sat_med   = oou_meta.get('satMedian', 1200)
    s = sat if sat and sat >= 400 else None
    g = gpa if gpa and gpa > 0  else None

    if accept <= 7:
        label = 'Moonshot'
    elif accept <= 15:
        if s and s >= sat_med - 30:
            label = 'Major Reach'
        else:
            label = 'Moonshot'
    elif accept <= 25:
        if s and s >= sat_med:
            label = 'Realistic Shot'
        elif s and s >= sat_med - 60:
            label = 'Major Reach'
        else:
            label = 'Moonshot'
    else:
        if s and s >= sat_med:
            label = 'Strong Chance'
        elif s and s >= sat_med - 80:
            label = 'Realistic Shot'
        else:
            label = 'Major Reach'

    return {'label': label, 'color': _LABEL_COLORS.get(label, '#94A3B8'),
            'total': None, 'acadScore': None, 'swimScore': None}

# Academic band × Swim support band → base admission label
_ADMIT_MATRIX = {
    4: {4: 'Very Strong Chance', 3: 'Strong Chance',  2: 'Strong Chance',  1: 'Realistic Shot', 0: 'Possible'},
    3: {4: 'Strong Chance',      3: 'Strong Chance',  2: 'Realistic Shot', 1: 'Possible',       0: 'Possible'},
    2: {4: 'Strong Chance',      3: 'Realistic Shot', 2: 'Possible',       1: 'Possible',       0: 'Major Reach'},
    1: {4: 'Realistic Shot',     3: 'Possible',       2: 'Possible',       1: 'Major Reach',    0: 'Moonshot'},
    0: {4: 'Major Reach',        3: 'Major Reach',    2: 'Moonshot',       1: 'Moonshot',       0: 'Moonshot'},
}

# Swim tier → base swim support band (PSF adjusts by ±1 max)
_SWIM_BASE_BAND = {
    'High-Point Contender': 4,
    'Conference Star':      4,
    'Priority Recruit':     3,
    'Top Recruit':          3,
    'Recruitable':          2,
    'Reach':                1,
    'Moonshot':             0,
}


def _adm_clamp(x, lo, hi):
    return max(lo, min(hi, x))


def _cap_label(current, max_label):
    """Return current label, or max_label if current is better than max_label."""
    return current if _LABEL_ORDER[current] <= _LABEL_ORDER[max_label] else max_label


def _selectivity_tier(accept_pct, sat_median):
    """
    Classify school selectivity from accept % (integer) and SAT median (integer).
    Returns: 'ultra_selective' | 'highly_selective' | 'selective' | 'broader_admit'
    """
    acc = (accept_pct / 100.0) if accept_pct is not None else None
    sat = sat_median
    if (acc is not None and acc <= 0.15) or (sat is not None and sat >= 1500):
        return 'ultra_selective'
    if (acc is not None and acc <= 0.30) or (sat is not None and sat >= 1400):
        return 'highly_selective'
    if (acc is not None and acc <= 0.50) or (sat is not None and sat >= 1300):
        return 'selective'
    return 'broader_admit'



def admission_chance(school, sat, gpa, adj_tier, psf):
    """
    Matrix-based admission model.  Same call signature as the old function.

    Inputs
      school   — canonical school name
      sat      — swimmer SAT composite (0 or None = not entered)
      gpa      — swimmer GPA (0 or None = not entered)
      adj_tier — swim tier string from tier_label(adjPts)
      psf      — program strength factor from TEAMS_LIST

    Returns AdmissionResult:
      label     — one of 6 labels (or 'Moonshot — Apply for Fun' / 'Unknown')
      color     — hex color for UI
      total     — None (old field kept for compatibility)
      acadScore — academic band 0-4 (debug)
      swimScore — swim support band 0-4 (debug)
    """
    meta = SCHOOL_META.get(school)
    if meta is None:
        return {'label': 'Unknown', 'color': _LABEL_COLORS['Unknown'],
                'total': None, 'acadScore': None, 'swimScore': None}

    if meta.get('moonshot'):
        return {'label': 'Moonshot — Apply for Fun',
                'color': _LABEL_COLORS['Moonshot — Apply for Fun'],
                'total': None, 'acadScore': None, 'swimScore': None}

    # Normalise inputs — treat 0 as "not entered"
    s = sat if sat and sat >= 400 else None
    g = gpa if gpa and gpa > 0  else None

    accept_pct = meta.get('accept')      # integer percent, e.g. 7 for 7%
    sat_median = meta.get('satMedian')   # integer, e.g. 1510

    # ── School academic ranges ────────────────────────────────────────────────
    # Prefer real CDS sat25/sat75 stored in SCHOOL_META; fall back to satMedian ±60.
    sat25 = meta.get('sat25') or ((sat_median - 60) if sat_median else None)
    sat75 = meta.get('sat75') or ((sat_median + 60) if sat_median else None)

    # ── Selectivity tier & floors ────────────────────────────────────────────
    sel_tier = _selectivity_tier(accept_pct, sat_median)
    gpa_off  = {'ultra_selective': 0.20, 'highly_selective': 0.25,
                'selective': 0.30,       'broader_admit':    0.35}
    sat_off  = {'ultra_selective': 60,   'highly_selective': 80,
                'selective': 100,        'broader_admit':    120}
    sat_floor = max(800, sat25 - sat_off[sel_tier]) if sat25 is not None else None
    gpa_floor = None   # no gpa25 available — gpa floor only applies via hard-stop

    # ── Academic band ────────────────────────────────────────────────────────
    gpa_below_floor = False
    sat_below_floor = False

    # Hard stop: GPA < 2.0 → band 0 regardless of SAT
    if g is not None and g < 2.0:
        acad_band = 0
        gpa_below_floor = True
        # Still check SAT floor so both_below_floor can trigger G1
        if s is not None and sat_floor is not None and s < sat_floor:
            sat_below_floor = True
    else:
        # SAT sub-score
        if s is None or sat25 is None or sat75 is None:
            sat_sub = 0
        else:
            sat_below_floor = sat_floor is not None and s < sat_floor
            if   s > sat75:            sat_sub =  2
            elif s >= sat25:           sat_sub =  1
            elif not sat_below_floor:  sat_sub =  0
            else:                      sat_sub = -2

        # GPA sub-score — compare swimmer GPA to school's real mean unweighted GPA.
        # gpaMean is sourced from published CDS or institutional data stored in SCHOOL_META.
        # When no real gpaMean is available, gpa_sub stays 0 (neutral) — no derivation.
        # +1: GPA > gpaMean + 0.10  (meaningfully above school's average)
        # -1: GPA < gpaMean - 0.15  (user-specified penalty threshold)
        #  0: within range, or no gpaMean data for this school
        gpa_mean = meta.get('gpaMean')
        if g is None or gpa_mean is None:
            gpa_sub = 0
        elif g > gpa_mean + 0.10:
            gpa_sub = 1
        elif g < gpa_mean - 0.15:
            gpa_sub = -1
        else:
            gpa_sub = 0

        raw = sat_sub + gpa_sub
        if   raw == 4:         acad_band = 4
        elif raw in (2, 3):    acad_band = 3
        elif raw in (0, 1):    acad_band = 2
        elif raw in (-1, -2):  acad_band = 1
        else:                  acad_band = 0

    both_below_floor = sat_below_floor and gpa_below_floor

    # ── Swim support band ────────────────────────────────────────────────────
    swim_base = _SWIM_BASE_BAND.get(adj_tier, 0)
    psf_mod   = 1 if psf > 1.00 else (-1 if psf <= 0.78 else 0)
    swim_band = _adm_clamp(swim_base + psf_mod, 0, 4)

    # ── Academic floor + school tier (admissions realism rules) ──────────────
    # hiddenIvy=True in SCHOOL_META marks top NESCAC/SCIAC LACs (Williams, Amherst,
    # Bowdoin, Middlebury, Pomona-Pitzer, etc.) — spec's HIGHLY_SELECTIVE tier.
    # MIT/Caltech are already caught by moonshot=True above and never reach here.
    # Ivy League schools are D1/OOU and use _oou_admission() — not this path.
    is_highly_selective = meta.get('hiddenIvy', False)

    # Academic floor: sat < satMedian-120 OR GPA < 3.4 → swim cannot assist
    sat_diff_from_med = (s - sat_median) if (s is not None and sat_median is not None) else None
    below_acad_floor  = (
        (sat_diff_from_med is not None and sat_diff_from_med <= -120) or
        (g is not None and g < 3.4)
    )

    # Apply swim constraints before matrix lookup
    if below_acad_floor:
        swim_band = 0                       # swim cannot rescue academic weakness
    elif is_highly_selective:
        swim_band = min(swim_band, 2)       # max +2 swim assist at elite D3 LACs

    # ── Base label from matrix ────────────────────────────────────────────────
    label = _ADMIT_MATRIX[acad_band][swim_band]

    # ── Guardrails (applied in order) ────────────────────────────────────────
    # G1: Both below floor → Moonshot
    if both_below_floor:
        label = 'Moonshot'
    # G2: GPA below floor (hard-stop triggered or explicit floor breach)
    elif gpa_below_floor:
        label = 'Major Reach' if swim_band == 4 else 'Moonshot'
    # G3: SAT below floor AND band 0 → cap at Major Reach
    elif sat_below_floor and acad_band == 0:
        label = _cap_label(label, 'Major Reach')

    # G4-GPA: Highly/ultra-selective + GPA < 3.5 → Moonshot regardless of SAT or swim
    if sel_tier in ('highly_selective', 'ultra_selective') and g is not None and g < 3.5:
        label = 'Moonshot'

    # G4: No swim support
    # CORE_D3 schools with strong academics (acad_band ≥ 3) can reach Strong Chance
    # without any swim support — academics alone justify it at broader-admit schools.
    # Elite LACs (hiddenIvy) and any below-floor case still cap at Possible.
    if swim_band == 0:
        if acad_band >= 3 and not is_highly_selective and not below_acad_floor:
            label = 'Strong Chance'
        else:
            label = _cap_label(label, 'Possible')

    # G5: Ultra-selective school extra caps
    if sel_tier == 'ultra_selective':
        if swim_band == 0:
            label = 'Moonshot'          # no swim support at elite schools → Extreme Reach
        elif acad_band == 1 and swim_band == 4:
            label = _cap_label(label, 'Possible')
        if acad_band == 0:
            label = 'Moonshot'
    # G6: Highly-selective school caps for band 0
    elif sel_tier == 'highly_selective':
        if acad_band == 0:
            if swim_band >= 3:
                label = _cap_label(label, 'Major Reach')  # no better than Major Reach
            else:
                label = 'Moonshot'

    # Spec floor cap: regardless of all guardrails, academic weakness ≤ Major Reach
    if below_acad_floor:
        label = _cap_label(label, 'Major Reach')

    return {
        'label':     label,
        'color':     _LABEL_COLORS.get(label, '#94A3B8'),
        'total':     None,
        'acadScore': acad_band,
        'swimScore': swim_band,
    }

# score_all_schools() removed — replaced by build_school_universe()
# which starts from the 324-school snapshot universe (EXPLORE_SCHOOLS),
# not from TEAMS_LIST.  See build_school_universe() below.


def build_school_universe(times, sat, gpa):
    """
    ONE unified school pipeline for ALL ~324 schools in EXPLORE_SCHOOLS.

    Every school goes through the SAME builder regardless of data richness.
    SCHOOL_META is an enrichment source.  Its absence changes VALUES, not CODE.

    Every result shares this schema:
      school, conference, division
      adjTier, adjPts, rawPts, psf
      top3, allEvents
      admission  { label, color, total, acadScore, swimScore }
      meta       { accept, satMedian, hiddenIvy, ivyLeague, stem, merit,
                   location, vibe, moonshot? }
      confTierShort, confTier, confFinish2026, confScore2026, confPowerClass
      hasSwimData   bool — True if this school has benchmark data in TEAMS_LIST

    Schools with swim data get full scoring.
    Schools without swim data get adjPts=0, adjTier='', empty events.
    Admission computed for all — 'Unknown' normalised to '' (value diff, not path diff).
    """
    results = []

    for es in EXPLORE_SCHOOLS:
        school_name = es['school']
        conference  = es.get('conference', '')

        # Re-use the team_rec match already resolved by _build_explore_schools()
        # (which already applies UAA short-name expansion and raw_name fallbacks).
        team_rec = es.get('_team_rec')
        has_swim = team_rec is not None

        # Unified meta — SCHOOL_META enriches, never gates
        meta_raw = (SCHOOL_META.get(school_name)
                    or (SCHOOL_META.get(team_rec['school']) if team_rec else None)
                    or {})
        meta = {
            'accept':    meta_raw.get('accept'),
            'satMedian': meta_raw.get('satMedian'),
            'hiddenIvy': meta_raw.get('hiddenIvy', False),
            'ivyLeague': meta_raw.get('ivyLeague', False),
            'stem':      meta_raw.get('stem', False),
            'merit':     meta_raw.get('merit', ''),
            'location':  meta_raw.get('location', ''),
            'vibe':      meta_raw.get('vibe', ''),
        }
        if meta_raw.get('moonshot'):
            meta['moonshot'] = True

        # Swim scoring — same call for all schools, values differ
        if has_swim:
            swim = _score_school_swim(team_rec, times)
            if swim:
                adj_tier   = swim['adjTier']
                adj_pts    = swim['adjPts']
                raw_pts    = swim['rawPts']
                top4       = swim['top3']   # dict key kept as 'top3' for compat
                all_events = swim['allEvents']
                psf        = swim['psf']
                has_depth  = swim['hasDepth']
            else:
                # Team data exists but swimmer has zero scorable events here
                adj_tier = ''
                adj_pts = raw_pts = 0.0
                top4 = all_events = []
                has_depth = False
                psf = team_rec.get('psf', 1.0)
        else:
            adj_tier = ''
            adj_pts = raw_pts = 0.0
            top4 = all_events = []
            has_depth = False
            psf = 1.0

        # Admission — same function for all schools.
        # SCHOOL_META keys may differ from snapshot CSV names (UAA abbreviated forms,
        # 30-char truncations, casing).  team_rec['school'] has already had TEAM_NAME_MAP
        # applied during load_data(), so it carries the correct canonical key.
        # Try snapshot name first; fall back to team_rec canonical if not found.
        _sm_key = (school_name
                   if SCHOOL_META.get(school_name)
                   else (team_rec['school']
                         if team_rec and SCHOOL_META.get(team_rec['school'])
                         else school_name))
        adm = admission_chance(_sm_key, sat, gpa, adj_tier, psf)
        # 'Unknown' means no SCHOOL_META entry — normalise to empty (value diff, not path diff)
        if adm.get('label') == 'Unknown':
            adm = {'label': '', 'color': '#94A3B8',
                   'total': None, 'acadScore': None, 'swimScore': None}

        results.append({
            'school':         school_name,
            'conference':     conference,
            'division':       CONF_DIVISION.get(conference, 'D3'),
            'adjTier':        adj_tier,
            'adjPts':         float(adj_pts),
            'rawPts':         float(raw_pts),
            'psf':            float(psf),
            'top3':           top4,
            'hasDepth':       has_depth,
            'allEvents':      all_events,
            'admission':      adm,
            'meta':           meta,
            'confTierShort':  es.get('conf_tier_short', ''),
            'confTier':       es.get('conf_tier', ''),
            'confFinish2026': es.get('men_finish_2026') or es.get('women_finish_2026'),
            'confScore2026':  es.get('men_score_2026') or es.get('women_score_2026'),
            'confPowerClass': es.get('conf_power_class', ''),
            'hasSwimData':    has_swim,
        })

    # Sort: swim-fit score descending, then alphabetical
    results.sort(key=lambda r: (-r['adjPts'], r['school']))
    return results


def score_one_school(times, conference, school):
    """
    Score arbitrary times at one specific school — for the manual calculator.
    Uses the TEAMS enrichment records directly (PSF + benchmarks).
    """
    team_key  = f"{conference}|{school}"
    team_rec  = TEAMS.get(team_key)
    if team_rec is None:
        return {'error': f'School "{school}" not found in {conference}'}

    # Swim layer — also collect unscored events for display
    scored, unscored = [], []
    for event, time_str in times.items():
        sec = parse_time(time_str)
        if sec is None:
            continue
        es = _score_event(event, time_str, conference)
        if es is not None:
            scored.append({**es, 'time': time_str, 'benchmarked': True})
        else:
            unscored.append({'event': event, 'time': time_str,
                             'sec': sec, 'benchmarked': False})

    scored.sort(key=lambda e: e['pts'], reverse=True)
    top4    = scored[:4]
    raw_pts = round(sum(e['pts'] for e in top4), 2)
    psf     = team_rec['psf']
    adj_pts = round(raw_pts * psf, 2)
    adj_tier = tier_label(adj_pts)
    adm     = admission_chance(school, JAMES['sat'], JAMES['gpa'], adj_tier, psf)
    meta_raw = SCHOOL_META.get(school, {})

    return {
        'school':     school,
        'conference': conference,
        'tier':       team_rec['tier'],
        'psf':        psf,
        'rawPts':     raw_pts,
        'adjPts':     adj_pts,
        'adjTier':    adj_tier,
        'top3':       [e['event'] for e in top4],   # key kept; now holds top 4
        'events':     scored + unscored,
        'admission':  adm,
        'meta': {
            'accept':    meta_raw.get('accept'),
            'satMedian': meta_raw.get('satMedian'),
            'hiddenIvy': meta_raw.get('hiddenIvy', False),
            'stem':      meta_raw.get('stem', False),
            'merit':     meta_raw.get('merit', ''),
            'location':  meta_raw.get('location', ''),
            'vibe':      meta_raw.get('vibe', ''),
        },
        'normalized': team_rec['normalized'],
        'rawName':    team_rec['raw_name'],
    }

# ---------------------------------------------------------------------------
# Claude helpers
# ---------------------------------------------------------------------------

def _get_anthropic():
    """Return an Anthropic client, or None if no key is configured."""
    key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not key:
        return None
    try:
        import anthropic
        return anthropic.Anthropic(api_key=key)
    except Exception:
        return None

def _similarity_sort(candidates, target):
    """
    Sort candidates by institutional similarity to a target school.
    Used when the user searches by school name — produces the pool that
    Claude picks the 5 'similar schools' from.

    Purely institutional dimensions — the swimmer's adjPts / swim score
    plays NO role here.  The recruiting badge still appears on each card
    from the normal scoring pipeline; it just doesn't shape which schools
    are considered 'similar'.

    Dimensions and weights:
      1. Division match         (0 if same, +4 penalty if different)
      2. Academic selectivity   (|tier distance| × 3, 4-bucket scale)
      3. Conference tier        (|tier distance| × 1, only when target has one)
    """
    SEL_ORDER  = {'ultra_selective': 0, 'highly_selective': 1,
                  'selective': 2, 'broader_admit': 3}
    TIER_ORDER = {'1A': 0, '1B': 1, '2': 2, '3': 3, '4': 4, '': 99}

    t_meta   = target.get('meta') or {}
    t_div    = target.get('division', '')
    t_sel    = _selectivity_tier(t_meta.get('accept'), t_meta.get('satMedian'))
    t_sel_n  = SEL_ORDER.get(t_sel, 3)
    t_tier   = target.get('confTierShort', '')
    t_tier_n = TIER_ORDER.get(t_tier, 99)
    has_tier = t_tier_n != 99  # skip tier distance when target has no conf-tier data

    def _score(r):
        r_meta   = r.get('meta') or {}
        r_div    = r.get('division', '')
        r_sel    = _selectivity_tier(r_meta.get('accept'), r_meta.get('satMedian'))
        r_sel_n  = SEL_ORDER.get(r_sel, 3)
        r_tier   = r.get('confTierShort', '')
        r_tier_n = TIER_ORDER.get(r_tier, 99)

        div_penalty = 0 if r_div == t_div else 4
        sel_dist    = abs(r_sel_n - t_sel_n) * 3
        tier_dist   = abs(r_tier_n - t_tier_n) if (has_tier and r_tier_n != 99) else 0
        return div_penalty + sel_dist + tier_dist

    return sorted(candidates, key=_score)


def _pre_sort(results, query, eliminated, my_list):
    """
    Re-sort top-35 slice based on query intent.
    Excludes eliminated schools and my-list schools before sorting.
    Returns top 35 from the resulting list.
    """
    excl = set(eliminated) | set(my_list)
    pool = [r for r in results if r['school'] not in excl]

    q = query.lower()

    # Unified strength score that works for both swim-scored and non-swim schools.
    # Swim schools have adjPts (real scored range); non-swim schools get a proxy
    # from confTierShort so they can compete in the pool instead of all sinking to 0.
    _TIER_PROXY = {'1A': 120, '1B': 95, '2': 70, '3': 45, '4': 20}
    def _strength(r):
        pts = r.get('adjPts') or 0
        return pts if pts > 0 else _TIER_PROXY.get(r.get('confTierShort', ''), 0)

    # Geographic region → set of 2-letter state abbreviations found in meta.location
    _GEO: dict[str, set[str]] = {
        'west coast':       {'CA', 'OR', 'WA'},
        'california':       {'CA'},
        'pacific northwest':{'OR', 'WA'},
        'southwest':        {'CA', 'AZ', 'NM', 'NV', 'UT', 'CO'},
        'mountain':         {'CO', 'UT', 'WY', 'ID', 'MT', 'NV', 'AZ', 'NM'},
        'new england':      {'MA', 'ME', 'NH', 'VT', 'RI', 'CT'},
        'northeast':        {'NY', 'NJ', 'CT', 'MA', 'ME', 'NH', 'VT', 'RI', 'PA', 'MD', 'DE'},
        'east coast':       {'NY', 'NJ', 'CT', 'MA', 'ME', 'NH', 'VT', 'RI', 'MD', 'DE',
                             'VA', 'NC', 'SC', 'GA', 'FL', 'DC'},
        'southeast':        {'VA', 'NC', 'SC', 'GA', 'FL', 'TN', 'AL', 'MS', 'LA', 'AR',
                             'KY', 'WV'},
        'south':            {'TX', 'LA', 'MS', 'AL', 'GA', 'FL', 'SC', 'NC', 'TN',
                             'KY', 'AR', 'OK', 'VA'},
        'texas':            {'TX'},
        'midwest':          {'OH', 'IN', 'IL', 'MI', 'WI', 'MN', 'IA', 'MO',
                             'ND', 'SD', 'NE', 'KS'},
        'florida':          {'FL'},
        'new york':         {'NY'},
    }

    geo_states: set[str] | None = None
    for kw, states in _GEO.items():
        if kw in q:
            geo_states = states
            break

    def _state(r) -> str:
        loc = r['meta'].get('location', '')
        return loc.rsplit(', ', 1)[-1].strip() if ', ' in loc else ''

    if geo_states:
        pool.sort(key=lambda r: (0 if _state(r) in geo_states else 1, -_strength(r)))
    elif any(k in q for k in ('prestig', 'best school', 'academic', 'selective')):
        pool.sort(key=lambda r: (r['meta'].get('accept') or 999))
    elif any(k in q for k in ('stem', 'engineer', 'tech', 'med', 'science')):
        pool.sort(key=lambda r: (0 if r['meta'].get('stem') else 1, -_strength(r)))
    elif any(k in q for k in ('money', 'cost', 'afford', 'save', 'merit', 'scholarship')):
        rank = {'high': 0, 'moderate': 1, 'none': 2, '': 3}
        pool.sort(key=lambda r: rank.get(r['meta'].get('merit', ''), 3))
    elif any(k in q for k in ('star', 'podium', 'win', 'lead', 'competi')):
        pool.sort(key=lambda r: -_strength(r))
    elif any(k in q for k in ('fun', 'social', 'happy', 'vibe', 'culture')):
        pool.sort(key=lambda r: -(r['meta'].get('accept') or 0))
    elif 'hidden ivy' in q or 'ivy' in q:
        pool.sort(key=lambda r: (0 if r['meta'].get('hiddenIvy') or r['meta'].get('ivyLeague') else 1))
    elif any(k in q for k in ('d1', 'division 1', 'division one', 'big school', 'large')):
        pool.sort(key=lambda r: (0 if r.get('division') == 'D1' else 1, -_strength(r)))
    elif any(k in q for k in ('d3', 'division 3', 'division iii', 'small school', 'small college')):
        pool.sort(key=lambda r: (0 if r.get('division') == 'D3' else 1, -_strength(r)))
    else:
        # Default: balanced sort using unified strength so non-swim schools
        # are not all buried at adjPts=0 behind every swim-scored D3 school.
        pool.sort(key=lambda r: -_strength(r))

    return pool[:35]

def _program_strength_desc(r):
    """Plain-language program strength label — never uses the word 'tier'."""
    ts = r.get('confTierShort', '')
    if ts == '1A': return 'Super Powerhouse'
    if ts == '1B': return 'Powerhouse'
    if ts == '2':  return 'Strong'
    if ts == '3':  return 'Mid-pack'
    if ts == '4':  return 'Developing'
    return r.get('adjTier', '')  # fallback to internal score label

def _build_school_line(i, r):
    """Format one numbered line for the Claude search prompt."""
    vibe = (r['meta'].get('vibe') or '')[:60]
    return (
        f"{i+1}. {r['school']} ({r['conference']}): "
        f"programStrength={_program_strength_desc(r)}, admission={r['admission']['label']}, "
        f"hiddenIvy={str(r['meta'].get('hiddenIvy', False)).lower()}, "
        f"stem={str(r['meta'].get('stem', False)).lower()}, "
        f"merit={r['meta'].get('merit', '')}, "
        f"accept={r['meta'].get('accept', '?')}%, "
        f"vibe=\"{vibe}\""
    )

# ─────────────────────────────────────────────────────────────────────────────
# AI-FIRST SEARCH PIPELINE — Steps 1-5 helpers
# ─────────────────────────────────────────────────────────────────────────────

_ADM_LABEL_SCORE: dict = {
    'Very Strong Chance':       100,
    'Strong Chance':             80,
    'Realistic Shot':            60,
    'Possible':                  40,
    'Major Reach':               15,
    'Moonshot':                   5,
    'Moonshot — Apply for Fun':   2,
    'Unknown':                   25,
}


def _classify_query_mode(query: str) -> str:
    """Step 1 — Classify query: GUIDED / CONSTRAINED / OBJECTIVE / EXPLORATORY.

    Priority order: GUIDED > CONSTRAINED > OBJECTIVE > EXPLORATORY.

    GUIDED   — explicit personalization ('for me', 'where should I', etc.)
    CONSTRAINED — personal-fit gate ('where I can get in', 'near me', etc.)
    OBJECTIVE — category-ranking truth query ('best STEM schools', 'top biology')
                with no personal qualifiers → LLM order preserved, no fit reranking.
    EXPLORATORY — open-ended browsing; light LLM + fit-aware ranking.
    """
    q = query.lower().strip()

    guided = [
        'for me', 'for my', 'for a swimmer like',
        'where should i', 'what should i', 'should i look',
        'where can i swim', 'best schools for swimming',
        'good colleges for swimming', 'good schools for swimming',
        'where i should', 'help me find', 'find me schools',
        'find schools for me', 'recommend for me',
    ]
    if any(s in q for s in guided):
        return 'GUIDED'

    constrained = [
        'where i can', 'i can get in', 'can get into',
        'i can make', 'near me', 'close to home', 'i could get into',
    ]
    if any(c in q for c in constrained):
        return 'CONSTRAINED'

    # OBJECTIVE: any "best / top / strongest / ..." query that has NO personal
    # qualifiers.  These must return a nationally defensible category truth.
    personal_qualifiers = [
        'for me', 'for my', 'where i', 'near me', 'i can', 'i could',
        'help me', 'for a swimmer', 'for the swimmer', 'fit for',
    ]
    obj_triggers = [
        'best ', 'top ', 'strongest ', 'hardest ', 'most prestigious',
        'most selective', 'most respected', 'most competitive',
        'leading ', 'premier ', 'elite ', 'top-tier',
        'nationally ranked', 'world class', 'globally ranked',
        'us news', 'top 10', 'top 20', 'top 25', 'top 50',
        'best colleges in', 'best universities', 'top universities',
    ]
    has_obj_trigger = any(s in q for s in obj_triggers)
    has_personal    = any(s in q for s in personal_qualifiers)
    if has_obj_trigger and not has_personal:
        return 'OBJECTIVE'

    return 'EXPLORATORY'


# Admissions label → numeric score (used for sorting and filtering)
_ADM_LABEL_SCORE = {
    'Very Strong Chance': 100,
    'Strong Chance':       80,
    'Realistic Shot':      60,
    'Possible':            40,
    'Major Reach':         15,
    'Moonshot':             5,
    'Moonshot — Apply for Fun': 2,
    'Unknown':             25,
}

# adjTier values that mean the swimmer is NOT competitive at this school
_SWIM_NOT_COMPETITIVE = {'Reach', 'Moonshot'}

# adjTier values that mean swim is literally impossible / should be hidden
_ADM_IMPOSSIBLE = {'Moonshot', 'Moonshot — Apply for Fun'}


def _detect_query_intent(query: str) -> dict:
    """Detect what filtering rules apply to a personalized query.

    Returns a dict of boolean flags:
      is_swim     — query is about swimming / contributing in the pool
      is_adm      — query is about where the student can get in
      is_personal — query is clearly "for me" / personalized
    """
    q = query.lower()
    is_swim = any(s in q for s in [
        'swim', 'pool', 'stroke', 'relay', 'contribute', 'compete in',
        'make the team', 'recruitable', 'roster', 'athletic fit',
        'where can i swim', 'where should i swim',
    ])
    is_adm = any(s in q for s in [
        'get in', 'admissible', 'realistic', 'where i can get',
        'where i can apply', 'i could get into', 'can get into',
        'i can get in', 'where i can',
    ])
    is_personal = any(s in q for s in [
        'for me', 'for my', 'where should i', 'help me',
        'recommend', 'find me', 'i should', 'my list',
        'where i', 'i can', 'i could',
    ])
    return {'is_swim': is_swim, 'is_adm': is_adm, 'is_personal': is_personal}


def _hard_filter(candidates: list, intent: dict) -> tuple:
    """Step 2 — Remove schools that contradict our scoring labels.

    SWIM queries: keep only schools where the swimmer is competitive
                  (adjTier not in _SWIM_NOT_COMPETITIVE AND hasSwimData).
    ADMISSIONS queries: keep only schools that are not impossible
                        (label not in _ADM_IMPOSSIBLE, i.e. not pure Moonshot).
    GENERAL 'for me' queries: remove only schools impossible on BOTH axes.
    No filtering at all for non-personal queries (OBJECTIVE / EXPLORATORY).

    Returns (kept_list, removed_debug_list).
    """
    kept, removed = [], []

    for r in candidates:
        adm_label = r.get('admission', {}).get('label', 'Unknown')
        adj_tier  = r.get('adjTier', '')
        has_swim  = r.get('hasSwimData', False)

        adm_bad  = adm_label in _ADM_IMPOSSIBLE
        swim_bad = (adj_tier in _SWIM_NOT_COMPETITIVE) or (not has_swim)

        reason = None

        if intent['is_swim']:
            if not has_swim:
                reason = f'no swim data'
            elif adj_tier in _SWIM_NOT_COMPETITIVE:
                reason = f'not competitive — {adj_tier}'

        elif intent['is_adm']:
            if adm_bad:
                reason = f'not admissible — {adm_label}'

        elif intent['is_personal']:
            # General "for me": only cut schools impossible on both axes
            if adm_bad and swim_bad:
                reason = f'impossible on both axes ({adm_label} / {adj_tier or "no swim data"})'

        if reason:
            removed.append({'school': r['school'], 'reason': reason})
        else:
            kept.append(r)

    return kept, removed


def _build_student_context(name, gpa, sat, act, times, vibe, other_prefs) -> str:
    """Build a concise student context string for GUIDED/CONSTRAINED prompts."""
    parts = [f"Student: {name or 'the swimmer'}"]
    if gpa:
        parts.append(f"GPA {gpa:.1f}")
    if sat:
        parts.append(f"SAT {sat}")
    elif act:
        parts.append(f"ACT {act}")
    if times:
        parts.append(f"events: {', '.join(list(times.keys())[:3])}")
    if vibe:
        skip = {'Not sure yet', 'Genuinely want to be well-rounded', '', None}
        prefs = [v for v in vibe.values() if v not in skip]
        if prefs:
            parts.append(f"preferences: {'; '.join(prefs[:4])}")
    if other_prefs and str(other_prefs).strip():
        parts.append(f"notes: {str(other_prefs).strip()[:200]}")
    return ', '.join(parts)


def _build_candidate_prompt(query: str, mode: str, student_ctx: str) -> tuple:
    """Step 2 — Build system + user prompts for candidate school generation.

    OBJECTIVE mode gets a fully separate prompt that forbids personalization and
    instructs the model to rank by genuine category truth only.  The resulting
    list order is preserved exactly in the final response — no fit reranking.

    GUIDED / CONSTRAINED get student context and normal pool-generation rules.
    EXPLORATORY gets no student context but normal pool-generation rules.
    """
    if mode == 'OBJECTIVE':
        system = (
            "You are an expert U.S. college counselor answering a category-ranking question.\n\n"
            "Rules — follow these EXACTLY:\n"
            "- Rank by the TRUE strength of each school for the asked category\n"
            "- Do NOT personalize — no student profile exists for this query\n"
            "- Do NOT consider admissions chances, athletics, swim teams, or finances\n"
            "- Do NOT consider any fit factors whatsoever\n"
            "- Return the strongest, most nationally-defensible schools for the category\n"
            "- Put the genuinely strongest schools FIRST (your order will be preserved)\n"
            "- Return ONLY valid JSON — no markdown, no extra text\n"
            "- 'schools' must contain exactly 12 full school name strings, strongest first\n"
            "- 'answer' is 1-2 sentences describing what makes these schools the best\n"
            'Format: {"answer": "...", "schools": ["Full School Name", ...]}'
        )
        user_lines = [
            f'Category-ranking query: "{query}"',
            "\nReturn the 12 strongest U.S. schools for this category, best first.",
            "Do NOT personalize. Do NOT consider swim teams or admissions fit.",
            "\nReturn JSON only.",
        ]
        return system, '\n'.join(user_lines)

    # GUIDED / CONSTRAINED / EXPLORATORY — pool generation (existing behaviour)
    system = (
        "You are an expert U.S. college counselor generating a candidate list of colleges. "
        "Your ONLY job is to produce a strong pool of schools relevant to the search query.\n\n"
        "Rules:\n"
        "- Focus on academic/program strength and institutional fit to the query\n"
        "- Do NOT evaluate admissions likelihood, athletic fit, or finances\n"
        "- Do NOT explain choices or write narratives\n"
        "- Return ONLY valid JSON — no markdown, no extra text\n"
        "- 'schools' must contain 12 to 18 full school name strings\n"
        "- 'answer' is 1-2 plain-English sentences interpreting the query\n"
        "- Include quality range — not just elite schools\n"
        "- Honor explicit constraints exactly (Ivy League, NESCAC, Midwest, D3, pre-med, etc.)\n"
        'Format: {"answer": "...", "schools": ["Full School Name", ...]}'
    )
    user_lines = [f'Search: "{query}"']
    if mode in ('GUIDED', 'CONSTRAINED') and student_ctx:
        user_lines.append(
            f"\nStudent context (use only to interpret the question — "
            f"do NOT filter for admissibility or swim fit):\n{student_ctx}"
        )
    else:
        user_lines.append("\n(General query — do not personalize.)")
    user_lines.append("\nReturn JSON only.")
    return system, '\n'.join(user_lines)


def _parse_candidate_names(text: str) -> tuple:
    """Parse LLM JSON → (answer str, list of school name strings)."""
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text).strip()
    m = re.search(r'\{[\s\S]*\}', text)
    if not m:
        raise ValueError('No JSON in candidate response')
    parsed = json.loads(m.group())
    answer = str(parsed.get('answer', '')).strip()
    names  = [str(s).strip() for s in parsed.get('schools', []) if str(s).strip()]
    if not names:
        raise ValueError('Empty candidate list returned by AI')
    return answer, names


_CAND_STOP = frozenset({
    'university', 'college', 'institute', 'school', 'of', 'the', 'at',
    'and', 'in', 'a', 'tech', 'for',
})


def _cname_norm(s: str) -> str:
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def _cname_toks(s: str) -> frozenset:
    return frozenset(t for t in _cname_norm(s).split()
                     if t not in _CAND_STOP and len(t) > 1)


def _map_to_universe(candidate_names: list, all_results: list) -> list:
    """
    Step 3 — Fuzzy-map LLM-generated school names → Lane4 school records.

    Three-tier matching:
      1. Exact normalized match
      2. Substring match (normalized name contained in/containing each other)
      3. Key-token Jaccard similarity ≥ 0.50
    Ignores candidates that don't match confidently; never fabricates schools.
    """
    by_norm = {_cname_norm(r['school']): r for r in all_results}
    mapped, seen = [], set()

    for cand in candidate_names:
        cand = cand.strip()
        if not cand:
            continue
        record = None

        # 1. Exact normalized match
        record = by_norm.get(_cname_norm(cand))

        # 2. Substring match
        if not record:
            c_n = _cname_norm(cand)
            for s_n, r in by_norm.items():
                if c_n and (c_n in s_n or s_n in c_n):
                    record = r
                    break

        # 3. Key-token Jaccard ≥ 0.50
        if not record:
            c_t = _cname_toks(cand)
            best, best_r = 0.0, None
            for r in all_results:
                r_t = _cname_toks(r['school'])
                if not c_t or not r_t:
                    continue
                jac = len(c_t & r_t) / len(c_t | r_t)
                if jac > best:
                    best, best_r = jac, r
            if best >= 0.50:
                record = best_r

        if record and record['school'] not in seen:
            seen.add(record['school'])
            mapped.append(record)

    return mapped


def _search_rank_score(r: dict, mode: str) -> float:
    """
    Step 5 — Composite ranking score (higher = shown first).

    GUIDED / CONSTRAINED: admissions truth dominates (0.60) + swim fit (0.40).
      Swim fit uses the scorer's actual adjPts when available:
        adjPts > 0          → swimmer earns points there, use actual score
        adjPts = 0 + has swim data → swimmer is below the roster bar → swim_n = 0
        adjPts = 0 + no swim data  → use a soft program-strength proxy (×0.25)
      → Michigan sinks when the swimmer would not make the roster.
      → MIT sinks when the student is not admissible.

    OBJECTIVE / EXPLORATORY: program strength leads (0.65) + admissions (0.35).
      Uses full program-strength proxy so results reflect genuine quality,
      not personalised fit for this swimmer.
    """
    _ADM_S = {
        'Very Strong Chance': 100, 'Strong Chance': 80,
        'Realistic Shot': 60,     'Possible': 40,
        'Major Reach': 15,        'Moonshot': 5,
        'Moonshot — Apply for Fun': 2, 'Unknown': 25,
    }
    adm_s    = _ADM_S.get(r.get('admission', {}).get('label', 'Unknown'), 25)
    adj      = r.get('adjPts') or 0
    has_swim = r.get('hasSwimData', False)
    _PROX    = {'1A': 130, '1B': 100, '2': 72, '3': 46, '4': 22}

    if mode in ('GUIDED', 'CONSTRAINED'):
        # Only credit swim fit when the scorer has confirmed the swimmer earns
        # points at this school.  Both "below the bar" (adjPts=0, hasSwimData=True)
        # and "no data" (hasSwimData=False) get swim_n=0 — if we can't confirm
        # relevance, we don't reward it.  Ranking falls back to admissions truth.
        swim_proxy = adj if adj > 0 else 0
        swim_n = min(swim_proxy / 130.0, 1.0) * 100
        # Tie-breaker: within the same admissions band, prefer broader-admit
        # schools (higher accept rate → easier reach → more honest for this kid)
        accept_bonus = (r.get('meta') or {}).get('accept') or 50
        accept_bonus = min(accept_bonus / 100.0, 1.0) * 3   # 0-3 pt tiebreak
        return adm_s * 0.60 + swim_n * 0.40 + accept_bonus

    # OBJECTIVE / EXPLORATORY — program strength, not personalised
    swim_proxy = adj if adj > 0 else _PROX.get(r.get('confTierShort', ''), 0)
    swim_n     = min(swim_proxy / 130.0, 1.0) * 100
    return swim_n * 0.65 + adm_s * 0.35


def _parse_search_response(text, sorted_35):
    """
    Parse Claude's JSON search response.
    Returns list of enriched SchoolResult dicts (with aiWhy), or raises ValueError.
    """
    # Strip markdown fences
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    text = text.strip()

    match = re.search(r'\{[\s\S]*\}', text)
    if not match:
        raise ValueError('No JSON object found in response')

    parsed = json.loads(match.group())
    answer  = parsed.get('answer', '')
    picks   = parsed.get('schools', [])

    schools = []
    for pick in picks:
        idx = pick.get('number')
        if idx is None:
            continue
        idx = int(idx) - 1
        if idx < 0 or idx >= len(sorted_35):
            continue
        r = dict(sorted_35[idx])
        r['aiWhy'] = pick.get('why', '')
        schools.append(r)

    if not schools:
        raise ValueError('No valid school picks in response')

    return answer, schools

def _build_top3_text(top3):
    """'1650 Free: Contender; 500 Free: 🏅 Podium' style string."""
    return '; '.join(f"{e['event']}: {place_label(e['place'])}" for e in top3)

def _build_vibe_lines(vibe, other_prefs=''):
    """Format answered vibe questions for deep dive prompt."""
    labels = {
        'swimGoal': 'Swim environment goal',
        'campus':   'Ideal campus feel',
        'friday':   'Friday night preference',
        'academic': 'Academic priority',
        'compete':  'Competition mindset',
        'location': 'Location preference',
        'career':   'Career interest',
    }
    lines = []
    if vibe:
        for k, v in vibe.items():
            if v:
                lines.append(f"  - {labels.get(k, k)}: {v}")
    if other_prefs and other_prefs.strip():
        lines.append(f"  - Additional preferences: {other_prefs.strip()}")
    return '\n'.join(lines)

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/meta', methods=['GET'])
def meta():
    return jsonify({
        'conferences':       sorted(CONFERENCES.keys()),
        'teams':             CONFERENCES,
        'events':            ALL_EVENTS,
        'normalizationLog':  NORMALIZATION_LOG,
    })

@app.route('/api/schools', methods=['GET'])
@login_required
def api_schools():
    """Return the unified 324-school explore dataset (modeled + snapshot_only)."""
    modeled   = sum(1 for s in EXPLORE_SCHOOLS if s.get('hasSwimData'))
    no_swim   = sum(1 for s in EXPLORE_SCHOOLS if not s.get('hasSwimData'))
    # Strip internal Python-only field before serialising
    schools_out = [{k: v for k, v in s.items() if k != '_team_rec'} for s in EXPLORE_SCHOOLS]
    return jsonify({
        'schools':       schools_out,
        'total':         len(schools_out),
        'withSwimData':  modeled,
        'conferenceOnly': no_swim,
    })

@app.route('/api/score-all', methods=['GET', 'POST'])
def score_all():
    """Score against all ~324 programs. POST body may include profile overrides."""
    if request.method == 'POST':
        body    = request.json or {}
        times   = body.get('times', JAMES['times'])
        sat     = int(body.get('sat',  JAMES['sat']))
        gpa     = float(body.get('gpa', JAMES['gpa']))
        profile = body if body else JAMES
    else:
        times, sat, gpa = JAMES['times'], JAMES['sat'], JAMES['gpa']
        profile = JAMES
    results = build_school_universe(times, sat, gpa)
    return jsonify({
        'profile':       profile,
        'totalSchools':  len(EXPLORE_SCHOOLS),
        'scoredSchools': sum(1 for r in results if r['adjPts'] > 0),
        'results':       results,
    })

@app.route('/api/search', methods=['POST'])
def search():
    """
    Natural-language search: re-sort top-35 by query intent, call Claude,
    return 6 annotated SchoolResult objects.

    If the query exactly (or partially) matches a school name, that school is
    returned first as a directMatch, and Claude picks 5 related schools.
    Otherwise Claude picks 6 schools from the pre-sorted pool.

    Body: { query, eliminated?, myList? }
    Response: { answer, schools, directMatch? } or { error, fallback? }
    """
    data       = request.json or {}
    query      = data.get('query', '').strip()
    eliminated = data.get('eliminated', [])
    my_list    = data.get('myList', [])
    prof_ovr   = data.get('profile', {})

    if not query:
        return jsonify({'error': 'Query is required'}), 400

    times         = prof_ovr.get('times') or JAMES['times']
    sat           = int(prof_ovr.get('sat')  or JAMES['sat'])
    gpa           = float(prof_ovr.get('gpa') or JAMES['gpa'])
    act_score     = prof_ovr.get('actScore', JAMES.get('actScore', 0)) or 0
    ap_count      = prof_ovr.get('apCount',  JAMES.get('apCount',  0)) or 0
    swimmer_name  = prof_ovr.get('name') or JAMES['name']
    # ONE unified pool — all ~324 schools through the same builder
    all_results = build_school_universe(times, sat, gpa)

    # ── Direct school-name match (single pass over unified 324-school universe) ──
    q_lower      = query.lower()
    direct_match = None

    # Exact match first, then prefix/partial match — both from unified universe
    for r in all_results:
        if r['school'].lower() == q_lower:
            direct_match = r
            break
    if not direct_match:
        for r in all_results:
            if q_lower in r['school'].lower():
                direct_match = r
                break

    # Out-of-universe stub — ONLY for school names truly not in the 324.
    # This handles schools like Harvard, Stanford that are outside D3 swimming.
    if not direct_match:
        _desc_words = {
            'find','show','good','best','near','with','help','looking','want',
            'suggest','recommend','schools','colleges','programs','like','similar',
            'strong','competitive','academic','research','liberal','division','d3',
            'private','public','small','large','northeast','south','west','midwest',
            # Question / intent words — prevent "where should I swim" from being
            # treated as a school name stub
            'where','should','would','could','can','what','which','how','why',
            'when','swim','swimming','get','into','for','me','my','i',
        }
        _words = q_lower.split()
        _looks_like_name = (
            1 <= len(_words) <= 5 and
            not any(w in _desc_words for w in _words) and
            not q_lower.endswith('?')
        )
        if _looks_like_name:
            display_name = ' '.join(
                w.upper() if w == w.upper() and len(w) > 1 else w.title()
                for w in query.split()
            )
            oou_meta = _oou_lookup(display_name) or {}
            oou_adm  = (_oou_admission(oou_meta, sat, gpa)
                        if oou_meta else {'label': '', 'color': '#94A3B8', 'total': None})
            direct_match = {
                'school':         display_name,
                'conference':     '',
                'division':       '',
                'adjTier':        '',
                'psf':            1.0,
                'admission':      oou_adm,
                'top3':           [],
                'hasDepth':       False,
                'allEvents':      [],
                'meta':           oou_meta,
                'confTierShort':  '',
                'confTier':       '',
                'confFinish2026': None,
                'confScore2026':  None,
                'confPowerClass': '',
                'hasSwimData':    False,
                'outOfUniverse':  True,
            }

    excl_names = set(eliminated) | set(my_list)
    if direct_match:
        excl_names.add(direct_match['school'])

    candidates = [r for r in all_results if r['school'] not in excl_names]
    if direct_match:
        # Sort by institutional similarity (division + selectivity + conf tier),
        # NOT by swimmer adjPts — so "similar schools" reflects what the school
        # is like, independently of how well this swimmer happens to score there.
        pool = _similarity_sort(candidates, direct_match)[:35]
    else:
        pool = candidates[:35]

    client = _get_anthropic()
    if not client:
        fallback = ([{**direct_match, 'directMatch': True}] if direct_match else []) + pool[:5 if direct_match else 6]
        return jsonify({
            'error': 'AI search is not configured',
            'detail': 'ANTHROPIC_API_KEY is missing or invalid',
            'fallback': fallback[:6],
            'directMatch': bool(direct_match),
        }), 503

    # ── NON-DIRECT: AI-FIRST PIPELINE (Steps 1-5) ────────────────────────────
    # Queries that are not a direct school-name lookup use a generative
    # candidate-pool approach: Claude proposes schools, Lane4 scores & ranks.
    if not direct_match:
        vibe_answers  = data.get('vibeAnswers', {}) or {}
        other_prefs_s = data.get('otherPrefs', '') or ''
        mode          = _classify_query_mode(query)
        student_ctx   = (
            _build_student_context(
                swimmer_name, gpa, sat, act_score, times,
                vibe_answers, other_prefs_s,
            )
            if mode in ('GUIDED', 'CONSTRAINED') else ''
        )
        cand_sys, cand_usr = _build_candidate_prompt(query, mode, student_ctx)

        try:
            resp = client.messages.create(
                model='claude-sonnet-4-6',
                max_tokens=700,
                system=cand_sys,
                messages=[{'role': 'user', 'content': cand_usr}],
            )
            answer, candidate_names = _parse_candidate_names(resp.content[0].text)

            # Step 3: map to universe
            excl_set  = set(eliminated) | set(my_list)
            avail     = [r for r in all_results if r['school'] not in excl_set]
            candidates = _map_to_universe(candidate_names, avail)

            # Narrow pool fallback: if mapping collapses to < 3, use _pre_sort
            if len(candidates) < 3:
                fallback = _pre_sort(all_results, query, eliminated, my_list)[:6]
                return jsonify({
                    'error': 'Candidate mapping too narrow',
                    'fallback': fallback,
                    'directMatch': False,
                }), 200

            # ── STEP 2: hard filter + sort ─────────────────────────────────
            raw_candidates = [r['school'] for r in candidates]
            removed_debug  = []

            if mode == 'OBJECTIVE':
                # Preserve LLM category-truth order exactly — no filter, no sort.
                filtered = candidates

            else:
                # Detect what filtering rules apply based on query wording.
                intent = _detect_query_intent(query)

                # Hard-remove schools that contradict our labels.
                filtered, removed_debug = _hard_filter(candidates, intent)

                # Sort remaining schools by admissions fit (strongest first).
                # Admissions is the primary and only ranking signal — swim is
                # already enforced via the hard filter above.
                filtered.sort(
                    key=lambda r: -_ADM_LABEL_SCORE.get(
                        r.get('admission', {}).get('label', 'Unknown'), 25
                    )
                )

            want_more = any(w in query.lower() for w in (
                'more schools', 'more options', 'show me more', 'give me more',
            ))
            limit   = 12 if want_more else 6
            schools = filtered[:limit]

            # Auto-generate aiWhy from existing score labels (no extra Claude call)
            for r in schools:
                adm_lbl  = r.get('admission', {}).get('label', '')
                swim_lbl = r.get('adjTier', '')
                parts    = [p for p in [adm_lbl, swim_lbl] if p and p not in ('Unknown', '')]
                r['aiWhy'] = ' · '.join(parts[:2])

            # ── DEBUG (visible in browser DevTools → Network) ──────────────
            intent_debug = _detect_query_intent(query) if mode != 'OBJECTIVE' else {}
            debug_kept = [
                {
                    'school': r['school'],
                    'admissions': r.get('admission', {}).get('label', 'Unknown'),
                    'swimTier':   r.get('adjTier', 'n/a'),
                    'hasSwimData': r.get('hasSwimData', False),
                }
                for r in schools
            ]

            return jsonify({
                'answer':      answer,
                'schools':     schools,
                'directMatch': False,
                'searchMode':  mode,
                '_debug': {
                    'queryMode':     mode,
                    'intent':        intent_debug,
                    'rawCandidates': raw_candidates,
                    'removed':       removed_debug,
                    'kept':          debug_kept,
                    'finalTop6':     [r['school'] for r in schools[:6]],
                },
            })

        except json.JSONDecodeError as e:
            fallback = _pre_sort(all_results, query, eliminated, my_list)[:6]
            return jsonify({
                'error': 'AI returned malformed JSON',
                'detail': str(e),
                'fallback': fallback,
                'directMatch': False,
            }), 200

        except (ValueError, Exception) as e:
            fallback = _pre_sort(all_results, query, eliminated, my_list)[:6]
            return jsonify({
                'error': 'Search failed',
                'detail': str(e),
                'fallback': fallback,
                'directMatch': False,
            }), 200

    # ── DIRECT MATCH PATH — unchanged (similarity sort + Claude picks 5) ─────
    system_prompt = (
        "You are Lane4. Respond ONLY with a valid JSON object. "
        "No markdown. No explanation. Start with { end with }. "
        "Keep 'why' fields under 15 words each. Keep 'answer' under 30 words."
    )

    school_lines = '\n'.join(_build_school_line(i, r) for i, r in enumerate(pool))
    user_prompt  = (
        f'The user searched for "{direct_match["school"]}" '
        f'({direct_match["conference"] or "independent"}, {direct_match.get("division","")}, '
        f'program strength: {_program_strength_desc(direct_match)}).\n\n'
        f"{swimmer_name}: GPA {gpa}, SAT {sat}" + (f", ACT {act_score}" if act_score else "") + ".\n\n"
        f"This list is already sorted by institutional similarity to {direct_match['school']} "
        f"(same division, selectivity, and conference tier). "
        "Pick the 5 schools that are most genuinely similar — consider academic culture, "
        "school size, mission, and athletic program level. Ignore the swimmer's times when judging similarity. "
        "Return ONLY JSON.\n\n"
        f"{school_lines}\n\n"
        'JSON format:\n{"answer":"1-2 sentences why these schools are similar to '
        f'{direct_match["school"]}' + '","schools":[{"number":1,"why":"under 15 words"}]}'
    )

    try:
        resp = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=800,
            system=system_prompt,
            messages=[{'role': 'user', 'content': user_prompt}],
        )
        raw_text = resp.content[0].text
        answer, ai_schools = _parse_search_response(raw_text, pool)

        _dm_conf  = direct_match.get('conference', '')
        _dm_div   = direct_match.get('division', '')
        _dm_str   = _program_strength_desc(direct_match)
        _dm_parts = [p for p in [_dm_conf, _dm_div, _dm_str] if p]
        _dm_why   = ' · '.join(_dm_parts) if _dm_parts else 'Matched from our database'
        dm        = {**direct_match, 'directMatch': True, 'aiWhy': _dm_why}
        schools   = [dm] + ai_schools

        return jsonify({'answer': answer, 'schools': schools, 'directMatch': True})

    except json.JSONDecodeError as e:
        fallback = [{**direct_match, 'directMatch': True}] + pool[:5]
        return jsonify({
            'error': 'AI returned malformed JSON',
            'detail': str(e),
            'fallback': fallback[:6],
            'directMatch': True,
        }), 200

    except ValueError as e:
        fallback = [{**direct_match, 'directMatch': True}] + pool[:5]
        return jsonify({'error': str(e), 'fallback': fallback[:6], 'directMatch': True}), 200

    except Exception as e:
        fallback = [{**direct_match, 'directMatch': True}] + pool[:5]
        return jsonify({
            'error': 'Search failed',
            'detail': str(e),
            'fallback': fallback[:6],
            'directMatch': True,
        }), 200


def _act_to_sat(act):
    """Rough ACT composite → SAT composite conversion."""
    table = {
        36: 1590, 35: 1540, 34: 1500, 33: 1460, 32: 1430, 31: 1400,
        30: 1370, 29: 1340, 28: 1310, 27: 1280, 26: 1240, 25: 1210,
        24: 1180, 23: 1140, 22: 1110, 21: 1080, 20: 1040, 19: 1010,
        18:  970, 17:  930, 16:  890, 15:  850,
    }
    return table.get(int(act), round((int(act) / 36) * 1200 + 400))


def _estimate_merit_block(merit_level, sat, gpa, sat_median, accept):
    """
    Compute deterministic merit estimates for The Money Conversation.
    Returns a dict with coa, merit, net, note, has_merit.
    """
    # COA estimate from acceptance rate
    if accept <= 20:
        coa = 72000
    elif accept <= 35:
        coa = 67000
    elif accept <= 55:
        coa = 61000
    elif accept <= 70:
        coa = 56000
    else:
        coa = 51000
    coa_str = f"~${coa:,}"

    if merit_level == 'none':
        return {
            'coa':       coa_str,
            'merit':     'None',
            'net':       f"~${coa:,} before need-based aid",
            'note':      "This school does not offer merit scholarships—aid here is need-based.",
            'has_merit': False,
        }

    max_merit = 25000 if merit_level == 'high' else 15000
    sat_diff  = (sat or 1000) - (sat_median or 1200)
    elite_gpa = gpa >= 3.9

    if sat_diff < -50:
        lo_pct, hi_pct = 0.0, 0.0
    elif sat_diff <= 50 and not elite_gpa:
        lo_pct, hi_pct = 0.25, 0.40
    elif sat_diff <= 120 and not elite_gpa:
        lo_pct, hi_pct = 0.50, 0.65
    else:
        lo_pct, hi_pct = 0.65, 0.75

    if lo_pct == 0.0:
        return {
            'coa':       coa_str,
            'merit':     '$0',
            'net':       coa_str,
            'note':      "Your current academic numbers are below this school's typical merit threshold.",
            'has_merit': False,
        }

    lo = round(max_merit * lo_pct / 1000) * 1000
    hi = round(max_merit * hi_pct / 1000) * 1000
    return {
        'coa':       coa_str,
        'merit':     f"~${lo:,}–${hi:,}",
        'net':       f"~${max(0, coa - hi):,}–${max(0, coa - lo):,}",
        'note':      "Strong students like you are often considered for merit here.",
        'has_merit': True,
    }


@app.route('/api/deep-dive', methods=['POST'])
def deep_dive():
    """
    Generate the 8-section deep dive narrative for one school.

    Body: { school }   (school must match a key in score_all results)
    Response: { sections: [{title, body}] } or { error }
    """
    data     = request.json or {}
    school   = data.get('school', '').strip()
    prof_ovr = data.get('profile', {})

    if not school:
        return jsonify({'error': 'school is required'}), 400

    times = prof_ovr.get('times') or JAMES['times']
    sat   = int(prof_ovr.get('sat')  or JAMES['sat'])
    gpa   = float(prof_ovr.get('gpa') or JAMES['gpa'])
    swimmer_name = prof_ovr.get('name') or JAMES['name']
    math_sat         = prof_ovr.get('mathSat',          JAMES.get('mathSat', ''))
    sat_projected    = prof_ovr.get('satProjected',     JAMES.get('satProjected', ''))
    math_sat_proj    = prof_ovr.get('mathSatProjected', JAMES.get('mathSatProjected', ''))
    act_score        = prof_ovr.get('actScore',         JAMES.get('actScore', 0)) or 0
    ap_count         = prof_ovr.get('apCount',          JAMES.get('apCount',  0)) or 0
    grad_year        = prof_ovr.get('gradYear',         '2026')
    # ONE unified pool — all ~324 schools through the same builder
    all_results = build_school_universe(times, sat, gpa)
    result = next((r for r in all_results if r['school'] == school), None)
    is_oou = False

    if result is None:
        # School not in the D3 universe — check out-of-universe well-known schools
        oou_meta_found = _oou_lookup(school)
        if oou_meta_found:
            oou_adm = _oou_admission(oou_meta_found, sat, gpa)
            result = {
                'school':         school,
                'conference':     '',
                'division':       '',
                'adjTier':        '',
                'psf':            1.0,
                'admission':      oou_adm,
                'top3':           [],
                'hasDepth':       False,
                'allEvents':      [],
                'meta':           oou_meta_found,
                'confTierShort':  '',
                'confTier':       '',
                'confFinish2026': None,
                'confScore2026':  None,
                'confPowerClass': '',
                'hasSwimData':    False,
                'outOfUniverse':  True,
            }
            is_oou = True
        else:
            return jsonify({'error': f'School "{school}" not found'}), 404

    client = _get_anthropic()
    if not client:
        return jsonify({
            'error': 'AI deep dive is not configured',
            'detail': 'ANTHROPIC_API_KEY is missing or invalid',
        }), 503

    meta = result['meta']
    top3_text  = _build_top3_text(result['top3'])
    vibe_answers = data.get('vibeAnswers') or prof_ovr.get('vibe') or {}
    other_prefs  = data.get('otherPrefs', '')
    vibe_lines   = _build_vibe_lines(vibe_answers, other_prefs)

    _merit_sat = sat or (_act_to_sat(act_score) if act_score else 0)
    money = _estimate_merit_block(
        merit_level = meta.get('merit', 'moderate'),
        sat         = _merit_sat,
        gpa         = gpa,
        sat_median  = meta.get('satMedian', 0),
        accept      = meta.get('accept', 50),
    )
    money_block = (
        f"MONEY DATA — use these exact figures, do not invent different numbers:\n"
        f"Estimated COA: {money['coa']}\n"
        f"Estimated Merit: {money['merit']}"
        + (" (based on your academics)" if money['has_merit'] else "") + "\n"
        f"Estimated Net: {money['net']}\n"
        f"Merit note: {money['note']}"
    )

    vibe_block = ''
    if vibe_lines:
        vibe_block = (
            f"\n{swimmer_name.upper()}'S PERSONALITY & PREFERENCES "
            f"(use these to personalize Campus Life and tone):\n{vibe_lines}\n"
        )

    hidden_ivy_note = '\nThis is a Hidden Ivy — academically elite, employer-respected, without the brand tax.' if meta.get('hiddenIvy') else ''
    stem_note       = '\nStrong STEM programs.' if meta.get('stem') else ''

    sat_detail = f"SAT {sat}" if sat else ""
    if sat and math_sat:
        sat_detail += f" (math {math_sat})"
    if act_score:
        sat_detail += (", " if sat_detail else "") + f"ACT {act_score}"
    ap_detail = f", {ap_count} projected APs" if ap_count else ""

    # Determine academic direction for optional section
    career_raw   = (vibe_answers.get('career')   or '').strip()
    academic_raw = (vibe_answers.get('academic')  or '').strip()
    _generic_career   = career_raw   in ('', 'Not sure yet')
    _generic_academic = academic_raw in ('', 'Genuinely want to be well-rounded')
    if not _generic_career or not _generic_academic:
        _parts = [p for p in [career_raw, academic_raw]
                  if p and p not in ('Not sure yet', 'Genuinely want to be well-rounded')]
        academic_direction = ' / '.join(_parts) if _parts else None
    else:
        academic_direction = None

    # Admission comparison block
    sat_median   = meta.get('satMedian', 0)
    sat25        = meta.get('sat25', 0)
    sat75        = meta.get('sat75', 0)
    gpa_mean     = meta.get('gpaMean', 0)
    accept_rate  = meta.get('accept', 0)
    adm_swimmer  = f"GPA {gpa} unweighted"
    if sat:
        adm_swimmer += f", SAT {sat}"
    if act_score:
        adm_swimmer += f", ACT {act_score}"
    adm_school_parts = [f"~{accept_rate}% acceptance rate"]
    if sat_median:
        adm_school_parts.append(f"SAT median ~{sat_median}")
    if sat25 and sat75:
        adm_school_parts.append(f"SAT range ~{sat25}-{sat75}")
    if gpa_mean:
        adm_school_parts.append(f"GPA average ~{gpa_mean}")
    admission_comparison = (
        f"ADMISSION COMPARISON (use to write 'Are You Admissible?' — do not invent different numbers):\n"
        f"Swimmer: {adm_swimmer}\n"
        f"School: {', '.join(adm_school_parts)}\n"
        f"Admission outlook: {result['admission']['label']}"
    )

    prog_strength = _program_strength_desc(result)
    conf_tier_short = result.get('confTierShort', '')
    super_powerhouse_note = (
        f"\nIMPORTANT: {result['school']} is a Super Powerhouse — they dominate their conference "
        f"and recruit well above what most peer schools in {result['conference']} can attract. "
        "In 'In the Pool', call this out directly and tell the swimmer to look closely "
        "at the current roster and committed recruits before assuming a spot."
    ) if conf_tier_short == '1A' else ''

    system_prompt = (
        LANE4_DEEP_DIVE_PROMPT + "\n\n"
        "Lane4 Technical Vocabulary (always apply):\n"
        "- Never use the word 'tier' — describe programs as 'Super Powerhouse', 'Powerhouse', "
        "'dominant in conference', 'competitive', etc.\n"
        "- 'Hidden Ivy' = academically elite and employer-respected without the Stanford rejection "
        "rate. Use naturally when applicable.\n"
        "- Never use the words 'profile', 'good school', 'strong fit', or 'also'.\n"
        "- No em dashes anywhere in the output.\n"
        "- Respond using markdown sections starting with ## for each section title.\n"
        "- 2-3 sentences per section. Short paragraphs. Strong declarative sentences."
    )

    ivy_note = '\nThis is an Ivy League school — need-based aid only, no merit scholarships.' if meta.get('ivyLeague') else ''

    # Build optional academic section instruction
    if academic_direction:
        acad_section_instr = (
            f"## If You're Serious About {academic_direction}\n"
            f"Name the specific program or department at {result['school']}. "
            "Explain why it's strong there. Be specific — research access, outcomes, reputation. "
            "Do not write this section generically.\n"
        )
    else:
        acad_section_instr = (
            "[SKIP the optional academic section — no academic direction provided. "
            "Do not include any section about majors or academic programs.]\n"
        )

    if is_oou:
        user_prompt = (
            f"Write a deep dive for {swimmer_name} considering {result['school']}.\n\n"
            f"SWIMMER: {swimmer_name}, Class of {grad_year}, GPA {gpa} unweighted, "
            f"{sat_detail}{ap_detail}."
            f"{vibe_block}\n"
            f"SCHOOL: {result['school']}\n"
            f"School vibe: {meta.get('vibe', '')}\n"
            f"Location: {meta.get('location', '')}\n"
            f"{admission_comparison}\n"
            f"{money_block}\n"
            f"{hidden_ivy_note}{ivy_note}{stem_note}\n\n"
            "NOTE: This school is not in our swim recruiting database. The swimmer is comparing it "
            "against D3 options — be honest about what choosing this school means for swim.\n\n"
            "Write exactly these sections in this order. 2-3 sentences per section.\n\n"
            "## Bottom Line\n"
            "2-3 sentences. School value + academic/personal fit + overall verdict.\n"
            "## What {school} Is Known For\n".replace('{school}', result['school'])
            + f"{acad_section_instr}"
            "## Are You Admissible?\n"
            "Use the ADMISSION COMPARISON above. Compare swimmer numbers to school numbers. "
            "Plain-English read: in range, above, slightly below, or real reach.\n"
            "## What It Costs\n"
            "Use EXACTLY the MONEY DATA figures above. Do not change the numbers. "
            "Cover COA, merit or no merit, net cost, aid philosophy.\n"
            "## Campus Life\n"
            "What do four years here actually feel like? Size, energy, setting, social scene.\n"
            "## How It Compares to Your D3 Options\n"
            "Be honest — what does choosing this school mean for continuing to swim competitively?"
        )
    else:
        user_prompt = (
            f"Write a deep dive for {swimmer_name} considering {result['school']}.\n\n"
            f"SWIMMER: {swimmer_name}, Class of {grad_year}, GPA {gpa} unweighted, "
            f"{sat_detail}{ap_detail}."
            f"{vibe_block}\n"
            f"SWIM DATA AT {result['school'].upper()} ({result['conference']}):\n"
            f"Top events: {top3_text}\n"
            f"Program strength: {prog_strength}\n"
            f"{super_powerhouse_note}\n"
            f"School vibe: {meta.get('vibe', '')}\n"
            f"Location: {meta.get('location', '')}\n"
            f"{admission_comparison}\n"
            f"{money_block}\n"
            f"{hidden_ivy_note}{ivy_note}{stem_note}\n\n"
            "Write exactly these sections in this order. "
            "Swim fit is explained ONCE in 'In the Pool' — do not repeat it in other sections. "
            "Weave in personality preferences naturally where relevant. "
            "Use 'Hidden Ivy' naturally if applicable. 2-3 sentences per section.\n\n"
            "## Bottom Line\n"
            "2-3 sentences. Swim reality + school value + overall verdict. No hedging.\n"
            "## In the Pool\n"
            "Where this swimmer lands on the team. What that means. Trajectory if they hold or drop time. "
            "Sound like a coach talking plainly. No internal metrics.\n"
            "## Coach Interest — What to Expect\n"
            "Likely level of recruiting engagement. Will they respond quickly? Is this swimmer a priority? "
            "What moves the needle: time drops, roster gaps, academic strength, event needs.\n"
            f"## What {result['school']} Is Known For\n"
            "School identity. Make it feel important and real. Prestige and seriousness when deserved.\n"
            f"{acad_section_instr}"
            "## Are You Admissible?\n"
            "Use the ADMISSION COMPARISON above. Compare swimmer numbers to school numbers directly. "
            "Plain-English read: in range, above, slightly below, or real reach. "
            "If highly selective, say so. If swim support helps, mention it briefly.\n"
            "## What It Costs\n"
            "Use EXACTLY the MONEY DATA figures above. Do not change the numbers. "
            "Cover COA, merit or no merit, net cost, aid philosophy. Practical family language.\n"
            "## Campus Life\n"
            "What do four years here actually feel like? Size, energy, setting, social scene, "
            "what kind of student thrives. No brochure copy."
        )

    try:
        resp = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1200,
            system=system_prompt,
            messages=[{'role': 'user', 'content': user_prompt}],
        )
        raw = resp.content[0].text

        # Split on section headers — per OUTPUT_SCHEMA response parsing spec
        parts  = re.split(r'^## ', raw, flags=re.MULTILINE)
        sections = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            lines = part.split('\n', 1)
            title = lines[0].strip()
            body  = lines[1].strip() if len(lines) > 1 else ''
            if title and not title.startswith('#') and not title.startswith('---'):
                sections.append({'title': title, 'body': body})

        if not sections:
            return jsonify({'error': 'AI returned empty deep dive', 'raw': raw}), 200

        return jsonify({
            'school':    result['school'],
            'sections':  sections,
            'admission': result['admission'],
            'adjTier':   result['adjTier'],
            'meta':      meta,
        })

    except Exception as e:
        return jsonify({'error': 'Deep dive failed', 'detail': str(e)}), 200


@app.route('/api/coach-email', methods=['POST'])
def coach_email():
    """
    Generate deterministic coach email for one school. No AI call.

    Body: { school, profile? }
    Response: { subject, body }
    """
    data     = request.json or {}
    school   = data.get('school', '').strip()
    prof_ovr = data.get('profile', {})

    if not school:
        return jsonify({'error': 'school is required'}), 400

    times         = prof_ovr.get('times') or JAMES['times']
    sat           = int(prof_ovr.get('sat')  or JAMES['sat'])
    gpa           = float(prof_ovr.get('gpa') or JAMES['gpa'])
    act_score     = prof_ovr.get('actScore', JAMES.get('actScore', 0)) or 0
    ap_count      = prof_ovr.get('apCount',  JAMES.get('apCount',  0)) or 0
    swimmer_name  = prof_ovr.get('name') or JAMES['name']
    grad_year     = prof_ovr.get('gradYear',     '2026')

    all_results = build_school_universe(times, sat, gpa)
    result = next((r for r in all_results if r['school'] == school), None)

    if result is None:
        return jsonify({'error': f'School "{school}" not found'}), 404

    meta  = result['meta']
    top3  = result['top3']
    best  = top3[0] if top3 else None

    if best is None:
        return jsonify({'error': 'No scored events found for this school'}), 400

    # Performance descriptor
    if best['place'] <= 1.5:
        perf = f"projected to win the {best['event']}"
    elif best['place'] <= 3.5:
        perf = f"projected to podium in the {best['event']}"
    else:
        perf = f"projected as a conference A finalist in the {best['event']}"

    second     = f" I also project to score in the {top3[1]['event']}." if len(top3) > 1 else ''
    stem_note  = ' Your programs in engineering and CS align directly with my academic direction.' if meta.get('stem') else ''
    merit_note = " I've also been looking closely at your merit scholarship opportunities." if meta.get('merit') == 'high' else ''

    # Build times summary from actual swimmer times
    time_entries = list(times.items())
    if time_entries:
        times_text = ', '.join(f"{t} in the {e.lower()}" for e, t in time_entries[:3])
    else:
        times_text = 'competitive times across multiple events'

    # Determine grad year class label
    class_label = f"Class of {grad_year}"

    subject = f"Prospective Student-Athlete Inquiry — {class_label} | Competitive Swimmer"
    body = (
        f"Dear Coach,\n\n"
        f"My name is {swimmer_name} and I'm a student in the {class_label} with strong interest "
        f"in {result['school']}'s swim program.\n\n"
        f"At the {result['conference']} conference level, I'm {perf}.{second} "
        f"My current bests include {times_text}.\n\n"
        f"Academically I carry a {gpa} GPA"
        + (f", {sat} SAT" if sat else "")
        + (f" / {act_score} ACT" if act_score else "")
        + (f", with {ap_count} APs projected" if ap_count else "")
        + f".{stem_note}{merit_note}\n\n"
        f"I'd love to connect about your program. Would you have time for a brief call or campus visit?\n\n"
        f"Thank you,\n{swimmer_name}"
    )

    return jsonify({'subject': subject, 'body': body})


@app.route('/api/health', methods=['GET'])
def health():
    key_ok = bool(os.environ.get('ANTHROPIC_API_KEY', '').strip())
    swim_data_schools = sum(1 for s in EXPLORE_SCHOOLS if s.get('hasSwimData'))
    return jsonify({
        'status':              'ok',
        # ── Primary universe (source of truth) ───
        'universeSource':      'output/lane4_snapshot_compatible.csv',
        'totalSchools':        len(EXPLORE_SCHOOLS),
        'schoolsWithSwimData': swim_data_schools,
        'conferenceOnlySchools': len(EXPLORE_SCHOOLS) - swim_data_schools,
        # ── Enrichment sources ────────────────────
        'enrichmentSource':    'data/lane4_swim_model.xlsx',
        'benchmarks':          len(BENCHMARKS),
        'enrichmentRecords':   len(TEAMS_LIST),
        'admissionRecords':    len(SCHOOL_META),
        'normalized':          len(NORMALIZATION_LOG),
        'anthropicKey':        key_ok,
    })

@app.route('/snapshot', methods=['GET'])
def download_snapshot():
    """Serve the latest Lane4 team-tier snapshot CSV for download."""
    return send_from_directory('output', 'lane4_snapshot.csv', as_attachment=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SCHOOL IMAGE MANIFEST — on-demand lazy fetch
# static/school_images.json is a persistent cache.
# /api/school_image/<school> fetches one school at a time (Wikipedia REST API)
# and caches the result so the next render is instant.
# ═══════════════════════════════════════════════════════════════════════════════

_IMG_MANIFEST_PATH = os.path.join('static', 'school_images.json')
_IMG_MANIFEST_LOCK = threading.Lock()

# Filename fragments that mean the image is a logo/seal, not a campus photo
_IMG_BAD_TOKENS = [
    'seal', 'logo', 'coat_of_arms', 'arms_of', 'crest', 'flag_of',
    'wordmark', 'insignia', 'monogram', 'mascot', 'badge', 'shield',
    'patch', '_mark.', 'icon', 'vector',
]

# Unsplash fallback images (stable, no API key required)
_IMG_HERO_FB = 'https://images.unsplash.com/photo-1562774053-701939374585?w=1200&q=80'
_IMG_SL_FB   = 'https://images.unsplash.com/photo-1523050854058-8df90110c9f1?w=1200&q=80'
_IMG_SWIM_FB = 'https://images.unsplash.com/photo-1519315901367-f34ff9154487?w=1200&q=80'


def _img_load_manifest():
    try:
        with open(_IMG_MANIFEST_PATH) as f:
            return json.load(f)
    except Exception:
        return {}


def _img_save_entry(school, entry):
    with _IMG_MANIFEST_LOCK:
        try:
            manifest = _img_load_manifest()
            manifest[school] = entry
            tmp = _IMG_MANIFEST_PATH + '.tmp'
            with open(tmp, 'w') as f:
                json.dump(manifest, f)
            os.replace(tmp, _IMG_MANIFEST_PATH)   # atomic on Linux
        except Exception:
            pass


def _img_is_bad(url):
    """Return True if the URL looks like a logo, seal, or SVG — not a campus photo."""
    low = (url or '').lower()
    if low.endswith('.svg') or '.svg.' in low:
        return True
    fn = low.split('/')[-1]
    return any(tok in fn for tok in _IMG_BAD_TOKENS)


_IMG_UA = 'Lane4Recruit/1.0 (lane4.app; on-demand image fetch)'
_IMG_WIKI_API = 'https://en.wikipedia.org/w/api.php'


def _img_wiki_title(school_name):
    """Resolve school name → Wikipedia page title via search."""
    params = {
        'action': 'query', 'list': 'search',
        'srsearch': school_name, 'srlimit': 3, 'format': 'json',
    }
    qs  = urllib.parse.urlencode(params)
    req = urllib.request.Request(f'{_IMG_WIKI_API}?{qs}', headers={'User-Agent': _IMG_UA})
    try:
        with urllib.request.urlopen(req, timeout=8) as r:
            data    = json.loads(r.read().decode())
            results = data.get('query', {}).get('search', [])
            return results[0]['title'] if results else None
    except Exception:
        return None


def _img_wiki_page_images(page_title, width=1000):
    """
    Fetch up to 20 images from a Wikipedia page (generator=images).
    Returns list of (file_title, url) tuples.
    Uses a single API call (no per-image URL resolution needed — iiprop=url).
    """
    params = {
        'action': 'query', 'titles': page_title,
        'generator': 'images', 'gimlimit': '30',
        'prop': 'imageinfo', 'iiprop': 'url',
        'iiurlwidth': str(width), 'redirects': '', 'format': 'json',
    }
    qs  = urllib.parse.urlencode(params)
    req = urllib.request.Request(f'{_IMG_WIKI_API}?{qs}', headers={'User-Agent': _IMG_UA})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            data   = json.loads(r.read().decode())
            images = []
            for page in data.get('query', {}).get('pages', {}).values():
                info = page.get('imageinfo', [{}])
                url  = (info[0].get('thumburl') or info[0].get('url', '')) if info else ''
                if url:
                    images.append((page.get('title', ''), url))
            return images
    except Exception:
        return []


def _img_score(file_title, tokens):
    """Score a Wikipedia file title for campus-photo quality. Returns -999 if disqualified."""
    fn = file_title.lower()
    # Must be jpg/png
    if not (fn.endswith('.jpg') or fn.endswith('.jpeg') or fn.endswith('.png')):
        return -999
    # Disqualify logos, seals, SVGs
    if any(tok in fn for tok in _IMG_BAD_TOKENS):
        return -999
    s = 0
    for tok in tokens:
        if tok and tok in fn: s += 8
    for kw in ['campus', 'aerial', 'quad', 'quadrangle', 'grounds', 'entrance',
               'courtyard', 'panoram', 'building', 'chapel', 'tower', 'hall',
               'library', 'fountain', 'arch', 'gate', 'square', 'green']:
        if kw in fn: s += 5
    for kw in ['portrait', 'headshot', 'rally', 'game', 'protest']:
        if kw in fn: s -= 8
    return s


def _img_normalize_name(school_name):
    """
    Convert inverted school names to standard Wikipedia titles.
    e.g. 'California, University of, Berkeley' → 'University of California Berkeley'
         'Georgia, University of'              → 'University of Georgia'
    """
    import re
    # "X, University of[, subname]" → "University of X subname"
    m = re.match(r'^(.+?),\s*University of(?:,\s*(.+))?$', school_name, re.I)
    if m:
        place = m.group(1).strip()
        sub   = (' ' + m.group(2).strip()) if m.group(2) else ''
        return f'University of {place}{sub}'
    # "X, College of" → "College of X"
    m2 = re.match(r'^(.+?),\s*College of(.*)$', school_name, re.I)
    if m2:
        return f'College of {m2.group(1).strip()}{m2.group(2)}'.strip()
    # Drop trailing ", The"
    return re.sub(r',\s*The\s*$', '', school_name, flags=re.I).strip()


def _img_wiki_rest(school_name):
    """
    Fetch a campus photo for one school.
    Strategy:
      1. Wikipedia REST summary using normalised name → check thumbnail/originalimage.
      2. If REST gives a seal/logo or returns 404, search Wikipedia for the page.
      3. Scan all page images via generator=images, score, pick best campus photo.
      4. Optional: retry with 'university' appended for schools with short names.
    Returns a URL string or None (→ caller uses fallback).
    """
    stop = {'university', 'college', 'of', 'the', 'and', 'at', 'institute',
            'technology', 'polytechnic', 'state', 'a', 'an', 'in', 'for'}
    clean   = school_name.replace(',', '').replace('-', ' ')
    tokens  = [t.lower() for t in clean.split() if t.lower() not in stop and len(t) > 2][:5]
    norm    = _img_normalize_name(school_name)

    page_title = None

    # ── Step 1: REST summary (try normalised name, then original) ──────────
    for attempt in [norm, school_name]:
        safe     = urllib.parse.quote(attempt.replace(' ', '_'), safe='():,')
        rest_url = f'https://en.wikipedia.org/api/rest_v1/page/summary/{safe}'
        req      = urllib.request.Request(rest_url, headers={'User-Agent': _IMG_UA})
        try:
            with urllib.request.urlopen(req, timeout=8) as r:
                data  = json.loads(r.read().decode())
                page_title = data.get('title')
                for key in ['originalimage', 'thumbnail']:
                    candidate = data.get(key, {}).get('source', '')
                    if candidate and not _img_is_bad(candidate):
                        return candidate   # ← good photo found immediately
                break  # found the page, just no good thumbnail
        except urllib.error.HTTPError as e:
            if e.code == 429:
                return None   # rate-limited — abort
        except Exception:
            pass

    # ── Step 2: Search API if REST gave nothing useful ─────────────────────
    if not page_title:
        page_title = _img_wiki_title(norm) or _img_wiki_title(school_name) or norm
        time.sleep(0.3)

    # ── Step 3: Full page image scan ───────────────────────────────────────
    images = _img_wiki_page_images(page_title, width=1200)
    if images:
        candidates = [(t, u) for t, u in images if not _img_is_bad(u)]
        if candidates:
            best_t, best_u = max(candidates, key=lambda x: _img_score(x[0], tokens))
            if _img_score(best_t, tokens) > 0:
                return best_u

    # ── Step 4: Retry with 'University' appended (for short names) ─────────
    lname = school_name.lower()
    if 'university' not in lname and 'college' not in lname and 'institute' not in lname:
        alt = _img_wiki_title(school_name + ' university')
        if alt and alt != page_title:
            time.sleep(0.3)
            images = _img_wiki_page_images(alt, width=1200)
            if images:
                candidates = [(t, u) for t, u in images if not _img_is_bad(u)]
                if candidates:
                    best_t, best_u = max(candidates, key=lambda x: _img_score(x[0], tokens))
                    if _img_score(best_t, tokens) > 0:
                        return best_u

    return None


def _img_harvest_background():
    """
    Background thread: fetch real campus photos for all schools still on fallback.
    Runs once at startup, one school every ~1s, saves atomically after each hit.
    """
    time.sleep(15)   # let the app fully start before hitting Wikipedia
    manifest = _img_load_manifest()
    to_fetch = [s for s in sorted(SCHOOL_META.keys())
                if manifest.get(s, {}).get('hero_is_fallback', True)]
    print(f'[img-harvest] Starting: {len(to_fetch)} schools need real photos', flush=True)
    ok = fail = 0
    for school in to_fetch:
        try:
            url = _img_wiki_rest(school)
            if url:
                ex = manifest.get(school, {})
                entry = {
                    'hero': url, 'student_life': url,
                    'swim': ex.get('swim') or _IMG_SWIM_FB,
                    'hero_is_fallback': False,
                    'swim_is_fallback': ex.get('swim_is_fallback', True),
                }
                _img_save_entry(school, entry)
                manifest[school] = entry   # keep local copy in sync
                ok += 1
                print(f'[img-harvest] ✓ {school}', flush=True)
            else:
                fail += 1
        except Exception as exc:
            print(f'[img-harvest] error {school}: {exc}', flush=True)
            fail += 1
        time.sleep(1.0)   # polite — ~1 req/sec
    print(f'[img-harvest] Done. ok={ok} fail={fail}', flush=True)


@app.route('/api/school_image/<path:school_name>')
def api_school_image(school_name):
    """
    On-demand image lookup for one school.
    1. Return cached real image instantly if we already have one.
    2. Otherwise hit Wikipedia REST API (one lightweight call).
    3. Cache result (hit or miss) so repeat renders are instant.
    4. Always return a URL — never an error.
    """
    manifest = _img_load_manifest()
    entry    = manifest.get(school_name, {})

    # Already have a real (non-fallback) image cached
    if entry.get('hero') and not entry.get('hero_is_fallback'):
        return jsonify({'url': entry['hero'], 'is_fallback': False})

    # Try Wikipedia REST API
    hero_url = _img_wiki_rest(school_name)

    if hero_url:
        new_entry = {
            'hero':             hero_url,
            'student_life':     hero_url,
            'swim':             entry.get('swim') or _IMG_SWIM_FB,
            'hero_is_fallback': False,
            'swim_is_fallback': entry.get('swim_is_fallback', True),
        }
        _img_save_entry(school_name, new_entry)
        return jsonify({'url': hero_url, 'is_fallback': False})

    # Mark as failed so we don't retry on every page load (retry after 24h via timestamp)
    failed_entry = {**entry,
                    'hero':             _IMG_HERO_FB,
                    'student_life':     _IMG_SL_FB,
                    'hero_is_fallback': True,
                    'fetch_failed_at':  time.time()}
    _img_save_entry(school_name, failed_entry)
    return jsonify({'url': _IMG_HERO_FB, 'is_fallback': True})


if __name__ == '__main__':
    # Background image harvest — fills in real campus photos for all fallback schools
    _harvest_thread = threading.Thread(target=_img_harvest_background, daemon=True)
    _harvest_thread.start()
    app.run(host='0.0.0.0', port=5000, debug=True)
