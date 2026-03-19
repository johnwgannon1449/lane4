#!/usr/bin/env python3
"""
Lane4 Experimental Scoring Model — Calibration Experiment
==========================================================
Tests a 5th/14th-anchor curve against the production 8th/16th model.

DO NOT import from or modify production code.
DO NOT affect any production outputs.
This is a standalone, isolated experiment.

Design Principles:
  - Production: 3 zones using 1st / 8th / 16th anchors, confidence 0.65 at 14-16th
  - Experimental: 3 zones using derived 5th / 14th anchors, confidence 0.35 at 14-16th
  - Premium zone (1–5): time → place with steeper slope → more place improvement per second
  - Main zone (5–14): broader meaningful scoring band
  - Extension (14+): low-value, sharp confidence dropoff — reduces false optimism
  - Points formula identical in both (21 - place) → comparable rawPts / adjPts scale
  - Top-3 aggregation unchanged (per experiment spec)
"""

import os, sys, csv, json, re
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

# ── Paths ───────────────────────────────────────────────────────────────────
EXCEL_PATH  = ROOT / 'data' / 'lane4_swim_model.xlsx'
ANCHOR_CSV  = ROOT / 'output' / 'all_event_anchors.csv'
OUT_CSV     = ROOT / 'experiments' / 'exp_vs_prod_comparison.csv'
OUT_SUMMARY = ROOT / 'experiments' / 'exp_vs_prod_summary.txt'

# ── Test configuration ───────────────────────────────────────────────────────
TEST_SCHOOLS = [
    'Emory University',
    'University of Chicago',
    'Johns Hopkins University',
    'Williams College',
    'Franklin & Marshall College',
    'Gettysburg College',
    'Macalester College',
    'Whitman College',
]

# Representative D3-competitive swimmer
# Times chosen to land in the 4th–14th range across the test conferences
SAMPLE_SWIMMER = {
    '50 Free':  '22.1',
    '100 Free': '47.8',
    '200 Free': '1:46.5',
    '500 Free': '4:52.0',
    '100 Fly':  '53.5',
    '200 IM':   '2:02.0',
    '400 IM':   '4:22.0',
}

# ── Shared helpers ───────────────────────────────────────────────────────────

def _float(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def parse_time(t):
    """'MM:SS.ss' or 'SS.ss' → decimal seconds. Returns None on failure."""
    if not t or str(t).strip() == '':
        return None
    t = str(t).strip()
    m = re.match(r'^(\d+):(\d{2}(?:\.\d+)?)$', t)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))
    try:
        return float(t)
    except ValueError:
        return None

# ── Data loading (independent of production) ─────────────────────────────────

def load_benchmarks():
    """Load benchmarks from all_event_anchors.csv (supplement) + Excel (primary)."""
    benchmarks = {}

    # Primary: Excel Sheet1
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb['Sheet1']
    h  = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {v: i + 1 for i, v in enumerate(h) if v}
    for r in range(2, ws.max_row + 1):
        conf  = ws.cell(r, col['Conference']).value
        event = ws.cell(r, col['Event']).value
        if not conf or not event:
            continue
        benchmarks[f"{conf}|{event}"] = {
            'first':         _float(ws.cell(r, col['1st_sec']).value),
            'eighth':        _float(ws.cell(r, col['8th_sec']).value),
            'sixteenth':     _float(ws.cell(r, col['16th_sec']).value),
            'sec_per_place': _float(ws.cell(r, col['Sec_per_place']).value),
        }

    # Supplement: CSV anchors
    if ANCHOR_CSV.exists():
        with open(ANCHOR_CSV, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                conf  = row.get('Conference', '').strip()
                event = row.get('Event', '').strip()
                key   = f"{conf}|{event}"
                if key not in benchmarks and conf and event:
                    benchmarks[key] = {
                        'first':         _float(row.get('1st_seconds')),
                        'eighth':        _float(row.get('8th_seconds')),
                        'sixteenth':     _float(row.get('16th_seconds')),
                        'sec_per_place': _float(row.get('Sec_per_place')),
                    }
    return benchmarks


def load_teams():
    """Load team records (school, conference, psf) from Excel Team_Tiers."""
    teams = {}
    wb  = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws  = wb['Team_Tiers']
    h   = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {}
    for i, v in enumerate(h):
        if v and v not in col:
            col[v] = i + 1

    for r in range(2, ws.max_row + 1):
        conf   = ws.cell(r, col['Conference']).value
        school = ws.cell(r, col['Team']).value
        psf    = ws.cell(r, col['PSF']).value
        tier   = ws.cell(r, col['Tier']).value
        if not conf or not school:
            continue
        key = f"{conf}|{school}"
        if key not in teams:
            teams[key] = {
                'school':     school,
                'conference': conf,
                'psf':        _float(psf) if psf is not None else 1.0,
                'tier':       tier or '',
            }
    return teams

# ── PRODUCTION scoring (exact copy of production logic) ─────────────────────

def prod_estimate_place(sec, b):
    """Production: 3-zone linear interpolation on 1st / 8th / 16th anchors."""
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']
    spp       = b['sec_per_place'] or 1.0

    if sec <= first:
        return 1.0
    if sec <= eighth:
        return 1.0 + (sec - first) / ((eighth - first) or 1.0) * 7
    if sec <= sixteenth:
        return 8.0 + (sec - eighth) / ((sixteenth - eighth) or 1.0) * 8
    return 16.0 + (sec - sixteenth) / spp

def prod_exp_points(place):
    return max(0.0, min(20.0, 21.0 - place))

def prod_confidence(place):
    if place <= 12: return 1.0
    if place <= 14: return 0.85
    if place <= 16: return 0.65
    return 0.0

def prod_tier(pts):
    if pts < 1:  return 'Moonshot'
    if pts < 4:  return 'Reach'
    if pts < 10: return 'Recruitable'
    if pts < 18: return 'Priority Recruit'
    if pts < 35: return 'Top Recruit'
    if pts < 50: return 'Conference Star'
    return 'High-Point Contender'

def prod_place_label(place):
    if place <= 1.5:  return 'Contender'
    if place <= 3.5:  return 'Podium'
    if place <= 8.5:  return 'A Final'
    if place <= 16.5: return 'B Final'
    if place <= 20:   return 'Bubble'
    return 'Out of range'

def score_event_prod(event, time_str, conf, benchmarks):
    sec = parse_time(time_str)
    if sec is None:
        return None
    bench = benchmarks.get(f"{conf}|{event}")
    if bench is None:
        return None
    place  = prod_estimate_place(sec, bench)
    exp_pt = prod_exp_points(place)
    cw     = prod_confidence(place)
    pts    = exp_pt * cw
    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': prod_place_label(place),
    }

def score_school_prod(team_rec, times, benchmarks):
    conf = team_rec['conference']
    psf  = team_rec['psf']
    events = []
    for ev, t in times.items():
        es = score_event_prod(ev, t, conf, benchmarks)
        if es:
            events.append(es)
    events.sort(key=lambda e: e['pts'], reverse=True)
    top3    = events[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)
    if raw_pts == 0:
        return None
    adj_pts = round(raw_pts * psf, 2)
    return {
        'rawPts': raw_pts,
        'adjPts': adj_pts,
        'adjTier': prod_tier(adj_pts),
        'top3':   top3,
        'allEvents': events,
    }

# ── EXPERIMENTAL scoring (5th/14th anchors, modified confidence) ─────────────

def exp_estimate_place(sec, b):
    """
    Experimental: 3-zone interpolation using derived 5th and 14th anchors.

    Anchors derived from production 1st/8th/16th data:
      fifth      = first + (4/7) * (eighth - first)
      fourteenth = eighth + (6/8) * (sixteenth - eighth)

    Zones:
      [1, 5]  — premium: 4 places over (first→fifth) time span
      [5, 14] — main scoring: 9 places over (fifth→fourteenth) time span
      [14+]   — extension: extrapolated using 5-14 slope (no cliff)
    """
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']

    fifth      = first  + (4.0 / 7.0) * (eighth - first)
    fourteenth = eighth + (6.0 / 8.0) * (sixteenth - eighth)
    spp_ext    = max((fourteenth - fifth) / 9.0, 0.001)

    if sec <= first:
        return 1.0
    if sec <= fifth:
        return 1.0 + (sec - first) / max(fifth - first, 0.001) * 4.0
    if sec <= fourteenth:
        return 5.0 + (sec - fifth) / max(fourteenth - fifth, 0.001) * 9.0
    return 14.0 + (sec - fourteenth) / spp_ext

def exp_exp_points(place):
    """Same formula as production — keeps comparable scale."""
    return max(0.0, min(20.0, 21.0 - place))

def exp_confidence(place):
    """
    Experimental confidence weighting.
    Full weight extended to 14 (vs 12 in production).
    Sharp drop from 14–16 (0.35x vs 0.65x in production).
    Beyond 16: zero (same as production).

    Rationale: 14-16 swimmers are low-value extensions, not real recruits.
    Reducing their credit prevents false optimism at strong programs.
    """
    if place <= 14: return 1.0
    if place <= 16: return 0.35
    return 0.0

def exp_place_label(place):
    """Labels aligned to experimental zone definitions."""
    if place <= 1.5:  return 'Contender'
    if place <= 3.5:  return 'Podium'
    if place <= 5.5:  return 'Premium (1–5)'
    if place <= 14.5: return 'Main Scoring (6–14)'
    if place <= 16.5: return 'Extension (15–16)'
    return 'Out of range'

def score_event_exp(event, time_str, conf, benchmarks):
    sec = parse_time(time_str)
    if sec is None:
        return None
    bench = benchmarks.get(f"{conf}|{event}")
    if bench is None:
        return None
    place  = exp_estimate_place(sec, bench)
    exp_pt = exp_exp_points(place)
    cw     = exp_confidence(place)
    pts    = exp_pt * cw
    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': exp_place_label(place),
    }

def score_school_exp(team_rec, times, benchmarks):
    conf = team_rec['conference']
    psf  = team_rec['psf']
    events = []
    for ev, t in times.items():
        es = score_event_exp(ev, t, conf, benchmarks)
        if es:
            events.append(es)
    events.sort(key=lambda e: e['pts'], reverse=True)
    top3    = events[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)
    if raw_pts == 0:
        return None
    adj_pts = round(raw_pts * psf, 2)
    return {
        'rawPts': raw_pts,
        'adjPts': adj_pts,
        'adjTier': prod_tier(adj_pts),
        'top3':   top3,
        'allEvents': events,
    }

# ── Runner ───────────────────────────────────────────────────────────────────

def fmt_top3(top3):
    return ' | '.join(
        f"{e['event']} pl={e['place']:.1f} pts={e['pts']:.1f}"
        for e in top3
    )

def run():
    print("Loading data...")
    benchmarks = load_benchmarks()
    teams = load_teams()
    print(f"  {len(benchmarks)} benchmark entries, {len(teams)} team records\n")

    rows = []
    summary_lines = []

    for school in TEST_SCHOOLS:
        # Find the canonical team record
        team_rec = None
        for key, t in teams.items():
            if t['school'] == school:
                team_rec = t
                break
        if team_rec is None:
            print(f"  WARNING: {school!r} not found in teams data — skipping")
            continue

        conf = team_rec['conference']
        psf  = team_rec['psf']

        prod = score_school_prod(team_rec, SAMPLE_SWIMMER, benchmarks)
        exp  = score_school_exp (team_rec, SAMPLE_SWIMMER, benchmarks)

        if not prod and not exp:
            print(f"  {school}: no scorable events in {conf}")
            continue

        p_raw  = prod['rawPts']  if prod else 0
        p_adj  = prod['adjPts']  if prod else 0
        p_tier = prod['adjTier'] if prod else '—'
        p_top3 = fmt_top3(prod['top3']) if prod else '—'

        e_raw  = exp['rawPts']  if exp else 0
        e_adj  = exp['adjPts']  if exp else 0
        e_tier = exp['adjTier'] if exp else '—'
        e_top3 = fmt_top3(exp['top3']) if exp else '—'

        # Compute significant place shifts
        place_shifts = []
        if prod and exp:
            all_ev_prod = {e['event']: e for e in prod['allEvents']}
            all_ev_exp  = {e['event']: e for e in exp['allEvents']}
            for ev in all_ev_prod:
                if ev in all_ev_exp:
                    dp = all_ev_exp[ev]['place'] - all_ev_prod[ev]['place']
                    dp_pts = all_ev_exp[ev]['pts'] - all_ev_prod[ev]['pts']
                    if abs(dp) >= 0.5 or abs(dp_pts) >= 0.5:
                        place_shifts.append(
                            f"{ev}: prod pl={all_ev_prod[ev]['place']:.1f}/pts={all_ev_prod[ev]['pts']:.1f}"
                            f" → exp pl={all_ev_exp[ev]['place']:.1f}/pts={all_ev_exp[ev]['pts']:.1f}"
                            f" (Δplace={dp:+.1f}, Δpts={dp_pts:+.1f})"
                        )

        delta_raw = round(e_raw - p_raw, 2)
        delta_adj = round(e_adj - p_adj, 2)
        tier_change = 'same' if p_tier == e_tier else f"{p_tier} → {e_tier}"

        rows.append({
            'school':            school,
            'conference':        conf,
            'psf':               psf,
            'prod_rawPts':       p_raw,
            'exp_rawPts':        e_raw,
            'delta_rawPts':      delta_raw,
            'prod_adjPts':       p_adj,
            'exp_adjPts':        e_adj,
            'delta_adjPts':      delta_adj,
            'prod_tier':         p_tier,
            'exp_tier':          e_tier,
            'tier_change':       tier_change,
            'prod_top3':         p_top3,
            'exp_top3':          e_top3,
            'place_shifts':      '; '.join(place_shifts),
        })

        print(f"{school} ({conf}, PSF={psf})")
        print(f"  PROD  rawPts={p_raw:5.2f}  adjPts={p_adj:5.2f}  tier={p_tier}")
        print(f"  EXP   rawPts={e_raw:5.2f}  adjPts={e_adj:5.2f}  tier={e_tier}")
        print(f"  Δraw={delta_raw:+.2f}  Δadj={delta_adj:+.2f}  tier: {tier_change}")
        if place_shifts:
            for s in place_shifts:
                print(f"    {s}")
        print()

    # ── Write CSV ─────────────────────────────────────────────────────────────
    if rows:
        fieldnames = list(rows[0].keys())
        with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f"CSV written: {OUT_CSV}")

    # ── Write summary text ────────────────────────────────────────────────────
    harder   = [r for r in rows if r['delta_adjPts'] < -1.0]
    similar  = [r for r in rows if abs(r['delta_adjPts']) <= 1.0]
    easier   = [r for r in rows if r['delta_adjPts'] > 1.0]
    tier_chg = [r for r in rows if r['tier_change'] != 'same']

    lines = [
        "═" * 60,
        "EXPERIMENTAL vs PRODUCTION — SUMMARY",
        "═" * 60,
        "",
        f"Swimmer: representative D3-competitive profile",
        f"Events: {', '.join(SAMPLE_SWIMMER.keys())}",
        "",
        "Model differences:",
        "  Production:    3 zones (1st / 8th / 16th anchors)",
        "                 confidence: 1.0 (≤12), 0.85 (12–14), 0.65 (14–16), 0 (16+)",
        "  Experimental:  3 zones (1st / 5th / 14th anchors, derived)",
        "                 confidence: 1.0 (≤14), 0.35 (14–16), 0 (16+)",
        "",
        "─" * 60,
        f"Schools HARDER under experimental (adjPts dropped >1.0):",
    ]
    for r in harder:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f})")
    if not harder:
        lines.append("  (none)")

    lines += [
        "",
        f"Schools SIMILAR (adjPts within ±1.0):",
    ]
    for r in similar:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f})")
    if not similar:
        lines.append("  (none)")

    lines += [
        "",
        f"Schools EASIER under experimental (adjPts gained >1.0):",
    ]
    for r in easier:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f})")
    if not easier:
        lines.append("  (none)")

    lines += [
        "",
        f"Tier changes:",
    ]
    for r in tier_chg:
        lines.append(f"  {r['school']}: {r['tier_change']}")
    if not tier_chg:
        lines.append("  (no tier changes for this swimmer profile)")

    lines += [
        "",
        "─" * 60,
        "Interpretation:",
        "  If experimental shows harder scores at strong programs (UAA, NESCAC),",
        "  that aligns with recruiting reality — these programs are selective and",
        "  the 14–16 zone should not generate false optimism.",
        "  Tier changes from 'Top Recruit' → lower signal that the model is",
        "  correctly reducing borderline scores at top-end programs.",
        "═" * 60,
    ]

    summary = '\n'.join(lines)
    print(summary)
    with open(OUT_SUMMARY, 'w', encoding='utf-8') as f:
        f.write(summary + '\n')
    print(f"\nSummary written: {OUT_SUMMARY}")


if __name__ == '__main__':
    run()
