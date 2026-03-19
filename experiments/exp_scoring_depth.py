#!/usr/bin/env python3
"""
Lane4 Experimental Scoring — Event Depth Test
==============================================
Tests whether including additional events (beyond top 3) improves
recruiting realism.

Three variants:
  PRODUCTION  — top 3 events, no weighting (control)
  EXP_TOP4    — top 4 events, weights 1.0 / 1.0 / 1.0 / 0.70
  EXP_TOP5    — top 5 events, weights 1.0 / 1.0 / 1.0 / 0.70 / 0.50

DO NOT import from or modify production code.
DO NOT affect any production outputs.
"""

import csv, re, sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
EXCEL_PATH  = ROOT / 'data' / 'lane4_swim_model.xlsx'
ANCHOR_CSV  = ROOT / 'output' / 'all_event_anchors.csv'
OUT_CSV     = ROOT / 'output' / 'scoring_depth_experiment.csv'
OUT_SUMMARY = ROOT / 'experiments' / 'exp_depth_summary.txt'

# ── Variant weights (position-indexed, 0-based) ───────────────────────────
WEIGHTS = {
    'PROD':    [1.00, 1.00, 1.00],
    'EXP_TOP4': [1.00, 1.00, 1.00, 0.70],
    'EXP_TOP5': [1.00, 1.00, 1.00, 0.70, 0.50],
}

# ── Test schools ──────────────────────────────────────────────────────────
TEST_SCHOOLS = [
    'Emory University',
    'Johns Hopkins University',
    'University of Chicago',
    'Williams College',
    'Franklin & Marshall College',
    'Gettysburg College',
    'Macalester College',
    'Whitman College',
]

# ── Sample swimmer ────────────────────────────────────────────────────────
SAMPLE_SWIMMER = {
    '50 Free':  '22.1',
    '100 Free': '47.8',
    '200 Free': '1:46.5',
    '500 Free': '4:52.0',
    '100 Fly':  '53.5',
    '200 IM':   '2:02.0',
    '400 IM':   '4:22.0',
}

# ── Helpers ───────────────────────────────────────────────────────────────

def _float(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def parse_time(t):
    if not t or str(t).strip() == '':
        return None
    t = str(t).strip()
    m = re.match(r'^(\d+):(\d{2}(?:\.\d+)?)$', t)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))
    try:
        return float(t)
    except ValueError:
        return None

# ── Data loaders ──────────────────────────────────────────────────────────

def load_benchmarks():
    benchmarks = {}
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb['Sheet1']
    h  = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {v: i + 1 for i, v in enumerate(h) if v}
    for r in range(2, ws.max_row + 1):
        conf  = ws.cell(r, col['Conference']).value
        event = ws.cell(r, col['Event']).value
        if not conf or not event:
            continue
        benchmarks[f"{conf}|{event}"] = {
            'first':         _float(ws.cell(r, col['1st_sec']).value),
            'eighth':        _float(ws.cell(r, col['8th_sec']).value),
            'sixteenth':     _float(ws.cell(r, col['16th_sec']).value),
            'sec_per_place': _float(ws.cell(r, col['Sec_per_place']).value),
        }
    if ANCHOR_CSV.exists():
        with open(ANCHOR_CSV, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                conf  = row.get('Conference', '').strip()
                event = row.get('Event', '').strip()
                key   = f"{conf}|{event}"
                if key not in benchmarks and conf and event:
                    benchmarks[key] = {
                        'first':         _float(row.get('1st_seconds')),
                        'eighth':        _float(row.get('8th_seconds')),
                        'sixteenth':     _float(row.get('16th_seconds')),
                        'sec_per_place': _float(row.get('Sec_per_place')),
                    }
    return benchmarks


def load_teams():
    teams = {}
    wb  = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws  = wb['Team_Tiers']
    h   = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {}
    for i, v in enumerate(h):
        if v and v not in col:
            col[v] = i + 1
    for r in range(2, ws.max_row + 1):
        conf   = ws.cell(r, col['Conference']).value
        school = ws.cell(r, col['Team']).value
        psf    = ws.cell(r, col['PSF']).value
        tier   = ws.cell(r, col['Tier']).value
        if not conf or not school:
            continue
        key = f"{conf}|{school}"
        if key not in teams:
            teams[key] = {
                'school':     school,
                'conference': conf,
                'psf':        _float(psf) if psf is not None else 1.0,
                'tier':       tier or '',
            }
    return teams

# ── Production scoring (unchanged) ────────────────────────────────────────

def estimate_place(sec, b):
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']
    spp       = b['sec_per_place'] or 1.0
    if sec <= first:
        return 1.0
    if sec <= eighth:
        return 1.0 + (sec - first) / ((eighth - first) or 1.0) * 7
    if sec <= sixteenth:
        return 8.0 + (sec - eighth) / ((sixteenth - eighth) or 1.0) * 8
    return 16.0 + (sec - sixteenth) / spp

def exp_points(place):
    return max(0.0, min(20.0, 21.0 - place))

def confidence(place):
    if place <= 12: return 1.0
    if place <= 14: return 0.85
    if place <= 16: return 0.65
    return 0.0

def tier_label(pts):
    if pts < 1:  return 'Moonshot'
    if pts < 4:  return 'Reach'
    if pts < 10: return 'Recruitable'
    if pts < 18: return 'Priority Recruit'
    if pts < 35: return 'Top Recruit'
    if pts < 50: return 'Conference Star'
    return 'High-Point Contender'

def score_all_events(team_rec, times, benchmarks):
    """Score every event, sorted by pts descending. Returns full list."""
    conf = team_rec['conference']
    events = []
    for ev, t in times.items():
        sec = parse_time(t)
        if sec is None:
            continue
        bench = benchmarks.get(f"{conf}|{ev}")
        if bench is None:
            continue
        place  = estimate_place(sec, bench)
        ep     = exp_points(place)
        cw     = confidence(place)
        pts    = ep * cw
        events.append({
            'event': ev,
            'sec':   round(sec, 3),
            'place': round(place, 2),
            'pts':   round(pts, 2),
        })
    events.sort(key=lambda e: e['pts'], reverse=True)
    return events


def apply_variant(events, psf, weights):
    """Apply positional weights, sum, apply PSF."""
    raw = 0.0
    used = []
    for i, w in enumerate(weights):
        if i >= len(events):
            break
        pts = events[i]['pts'] * w
        raw += pts
        used.append(f"{events[i]['event']}×{w:.0%}={pts:.1f}")
    raw = round(raw, 2)
    adj = round(raw * psf, 2)
    return raw, adj, used

# ── Runner ────────────────────────────────────────────────────────────────

def run():
    print("Loading data...")
    benchmarks = load_benchmarks()
    teams      = load_teams()
    print(f"  {len(benchmarks)} benchmarks, {len(teams)} team records\n")

    rows         = []
    missing      = []

    for school in TEST_SCHOOLS:
        team_rec = None
        for t in teams.values():
            if t['school'] == school:
                team_rec = t
                break
        if team_rec is None:
            print(f"  WARNING: {school!r} not in teams data — skipping")
            missing.append(school)
            continue

        conf = team_rec['conference']
        psf  = team_rec['psf']

        all_events = score_all_events(team_rec, SAMPLE_SWIMMER, benchmarks)
        if not all_events:
            print(f"  {school}: no scorable events in {conf}")
            continue

        results = {}
        for variant, weights in WEIGHTS.items():
            raw, adj, used = apply_variant(all_events, psf, weights)
            results[variant] = {
                'raw': raw,
                'adj': adj,
                'tier': tier_label(adj),
                'events': used,
            }

        prod = results['PROD']
        top4 = results['EXP_TOP4']
        top5 = results['EXP_TOP5']

        print(f"{school} ({conf}, PSF={psf})")
        print(f"  PROD  adj={prod['adj']:5.2f}  tier={prod['tier']}")
        print(f"  TOP4  adj={top4['adj']:5.2f}  tier={top4['tier']}  Δ={top4['adj']-prod['adj']:+.2f}")
        print(f"  TOP5  adj={top5['adj']:5.2f}  tier={top5['tier']}  Δ={top5['adj']-prod['adj']:+.2f}")
        print(f"  top5 events: {', '.join(e['event'] for e in all_events[:5])}")
        print()

        rows.append({
            'school':          school,
            'conference':      conf,
            'psf':             psf,
            'prod_adjPts':     prod['adj'],
            'prod_tier':       prod['tier'],
            'top4_adjPts':     top4['adj'],
            'top4_tier':       top4['tier'],
            'top4_delta':      round(top4['adj'] - prod['adj'], 2),
            'top5_adjPts':     top5['adj'],
            'top5_tier':       top5['tier'],
            'top5_delta':      round(top5['adj'] - prod['adj'], 2),
            'prod_events':     ' | '.join(e['event'] for e in all_events[:3]),
            'top4_events':     ' | '.join(e['event'] for e in all_events[:4]),
            'top5_events':     ' | '.join(e['event'] for e in all_events[:5]),
            'top5_pts_detail': ' | '.join(top5['events']),
        })

    # ── CSV ────────────────────────────────────────────────────────────────
    if rows:
        fieldnames = list(rows[0].keys())
        with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f"CSV written → {OUT_CSV}")

    # ── Summary ────────────────────────────────────────────────────────────
    tier_changes_4 = [r for r in rows if r['top4_tier'] != r['prod_tier']]
    tier_changes_5 = [r for r in rows if r['top5_tier'] != r['prod_tier']]
    meaningful_4   = [r for r in rows if r['top4_delta'] > 2.0]
    meaningful_5   = [r for r in rows if r['top5_delta'] > 2.0]

    lines = [
        "═" * 64,
        "EVENT DEPTH EXPERIMENT — SUMMARY",
        "═" * 64,
        "",
        f"Swimmer: representative D3-competitive profile",
        f"Events scored: {', '.join(SAMPLE_SWIMMER.keys())}",
        f"Schools tested: {len(rows)} (skipped: {', '.join(missing) if missing else 'none'})",
        "",
        "Weights:  PROD=top3@100%  |  TOP4=+event4@70%  |  TOP5=+event5@50%",
        "",
        "─" * 64,
        "Per-school results:",
    ]
    for r in rows:
        t4_note = f"  ← tier: {r['prod_tier']} → {r['top4_tier']}" if r['top4_tier'] != r['prod_tier'] else ''
        t5_note = f"  ← tier: {r['prod_tier']} → {r['top5_tier']}" if r['top5_tier'] != r['prod_tier'] else ''
        lines.append(
            f"  {r['school']} ({r['conference']}, PSF={r['psf']})"
        )
        lines.append(
            f"    PROD={r['prod_adjPts']:.2f}  TOP4={r['top4_adjPts']:.2f} (Δ{r['top4_delta']:+.2f}){t4_note}  TOP5={r['top5_adjPts']:.2f} (Δ{r['top5_delta']:+.2f}){t5_note}"
        )

    lines += [
        "",
        "─" * 64,
        f"Tier changes vs PROD — TOP4: {len(tier_changes_4)} school(s)",
    ]
    for r in tier_changes_4:
        lines.append(f"  {r['school']}: {r['prod_tier']} → {r['top4_tier']}")
    if not tier_changes_4:
        lines.append("  (none)")

    lines += [f"Tier changes vs PROD — TOP5: {len(tier_changes_5)} school(s)"]
    for r in tier_changes_5:
        lines.append(f"  {r['school']}: {r['prod_tier']} → {r['top5_tier']}")
    if not tier_changes_5:
        lines.append("  (none)")

    lines += [
        "",
        "─" * 64,
        "Assessment:",
    ]

    # Auto-generate interpretation
    if not meaningful_4 and not meaningful_5:
        lines.append("  Depth adds little for this swimmer profile — all events beyond top 3 contribute")
        lines.append("  less than 2 adj pts even at 70% weight. The top-3 model captures this")
        lines.append("  swimmer's recruiting value well; depth expansion is not material.")
    else:
        if meaningful_5:
            lines.append(f"  TOP5 produces meaningful gains (>2 adj pts) at {len(meaningful_5)} school(s):")
            for r in meaningful_5:
                ev4 = r['top4_events'].split('|')[-1].strip() if '|' in r['top4_events'] else '—'
                lines.append(f"    {r['school']}: +{r['top5_delta']:.2f} adj pts  (4th event: {ev4})")
        if not tier_changes_5:
            lines.append("  No tier changes — depth adjusts points but doesn't flip any recruiting label.")
        else:
            lines.append("  Tier changes present — review whether they align with recruiting reality.")

    lines += [
        "",
        "─" * 64,
        "Data notes:",
    ]
    if missing:
        lines.append(f"  Skipped schools (not in Team_Tiers XLSX): {', '.join(missing)}")
        lines.append("  These schools exist in the app's live data but not in the Excel reference sheet.")
    lines += [
        "  All variants use production anchor benchmarks and PSF values unchanged.",
        "═" * 64,
    ]

    summary = '\n'.join(lines)
    print(summary)
    with open(OUT_SUMMARY, 'w', encoding='utf-8') as f:
        f.write(summary + '\n')
    print(f"\nSummary written → {OUT_SUMMARY}")


if __name__ == '__main__':
    run()
