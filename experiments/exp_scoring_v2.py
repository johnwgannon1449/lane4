#!/usr/bin/env python3
"""
Lane4 Experimental Scoring Model — v2 (TRUE 5th/14th Anchors)
==============================================================
Validates the 5th/14th confidence model against TRUE place data from parsed PDFs
instead of the derived anchors used in v1.

Source-of-truth strategy:
  PDF-parsed conferences (UAA, NESCAC, MIAC):
      TRUE 5th / 14th times are read directly from output/event_results_full.csv,
      which is the lean per-place export added in the v2 harvester patch.
  XLSX-only conferences (Centennial, NWC):
      5th / 14th remain DERIVED from 1st/8th/16th anchors (same as v1):
          fifth      = first + (4/7) * (eighth - first)
          fourteenth = eighth + (6/8) * (sixteenth - eighth)

DO NOT import from or modify production code.
DO NOT affect any production outputs.
This is a standalone, isolated experiment.
"""

import os, sys, csv, re
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
EXCEL_PATH      = ROOT / 'data' / 'lane4_swim_model.xlsx'
ANCHOR_CSV      = ROOT / 'output' / 'all_event_anchors.csv'
FULL_PLACES_CSV = ROOT / 'output' / 'event_results_full.csv'
OUT_CSV         = ROOT / 'experiments' / 'exp_v2_vs_prod_comparison.csv'
OUT_SUMMARY     = ROOT / 'experiments' / 'exp_v2_vs_prod_summary.txt'

# Conferences whose 5th/14th come from TRUE PDF data (not derived)
PDF_PARSED_CONFS = frozenset({"UAA", "NESCAC", "MIAC"})

# ── Test configuration ────────────────────────────────────────────────────────
TEST_SCHOOLS = [
    # UAA — PDF-parsed
    'Emory University',
    'University of Chicago',
    # NESCAC — PDF-parsed
    'Williams College',
    # MIAC — PDF-parsed
    'Macalester College',
    # Centennial — XLSX-only (derived)
    'Johns Hopkins University',
    'Franklin & Marshall College',
    'Gettysburg College',
    # NWC — XLSX-only (derived)
    'Whitman College',
]

# Gender of the sample swimmer — must match the gender column in event_results_full.csv
# to ensure TRUE anchors are taken from the correct sex's event data.
SAMPLE_GENDER = 'men'

# Representative D3-competitive swimmer
SAMPLE_SWIMMER = {
    '50 Free':  '22.1',
    '100 Free': '47.8',
    '200 Free': '1:46.5',
    '500 Free': '4:52.0',
    '100 Fly':  '53.5',
    '200 IM':   '2:02.0',
    '400 IM':   '4:22.0',
}

# ── Shared helpers ────────────────────────────────────────────────────────────

def _float(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def parse_time(t):
    """'MM:SS.ss' or 'SS.ss' → decimal seconds. Returns None on failure."""
    if not t or str(t).strip() == '':
        return None
    t = str(t).strip()
    m = re.match(r'^(\d+):(\d{2}(?:\.\d+)?)$', t)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))
    try:
        return float(t)
    except ValueError:
        return None

# ── Data loading (independent of production) ──────────────────────────────────

def load_benchmarks():
    """
    Load 1st/8th/16th anchor benchmarks from Excel + CSV.
    Returns dict keyed by 'Conference|Event'.
    """
    benchmarks = {}

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb['Sheet1']
    h  = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {v: i + 1 for i, v in enumerate(h) if v}
    for r in range(2, ws.max_row + 1):
        conf  = ws.cell(r, col['Conference']).value
        event = ws.cell(r, col['Event']).value
        if not conf or not event:
            continue
        benchmarks[f"{conf}|{event}"] = {
            'first':         _float(ws.cell(r, col['1st_sec']).value),
            'eighth':        _float(ws.cell(r, col['8th_sec']).value),
            'sixteenth':     _float(ws.cell(r, col['16th_sec']).value),
            'sec_per_place': _float(ws.cell(r, col['Sec_per_place']).value),
        }

    if ANCHOR_CSV.exists():
        with open(ANCHOR_CSV, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                conf  = row.get('Conference', '').strip()
                event = row.get('Event', '').strip()
                key   = f"{conf}|{event}"
                if key not in benchmarks and conf and event:
                    benchmarks[key] = {
                        'first':         _float(row.get('1st_seconds')),
                        'eighth':        _float(row.get('8th_seconds')),
                        'sixteenth':     _float(row.get('16th_seconds')),
                        'sec_per_place': _float(row.get('Sec_per_place')),
                    }
    return benchmarks


def load_full_places():
    """
    Load per-place data from event_results_full.csv.
    Returns dict: (Conference, Event) → {place_int → seconds_float}.
    Only populated for PDF-parsed conferences (UAA, NESCAC, MIAC).
    Returns empty dict if the file doesn't exist yet.
    """
    places: dict[tuple[str, str], dict[int, float]] = {}
    if not FULL_PLACES_CSV.exists():
        print(f"  WARNING: {FULL_PLACES_CSV} not found — v2 will fall back to derived anchors for ALL conferences")
        return places
    with open(FULL_PLACES_CSV, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            conf  = row.get('Conference', '').strip()
            event = row.get('Event', '').strip()
            place_raw = row.get('Place', '').strip()
            time_raw  = row.get('Time_Sec', '').strip()
            gender = row.get('Gender', '').strip()
            if not conf or not event or not place_raw or not time_raw:
                continue
            try:
                place_int = int(place_raw)
                time_sec  = float(time_raw)
            except ValueError:
                continue
            key = (conf, event, gender)
            if key not in places:
                places[key] = {}
            # Keep the best (fastest) time if somehow duplicate places exist
            if place_int not in places[key] or time_sec < places[key][place_int]:
                places[key][place_int] = time_sec
    return places


def build_exp_anchors(benchmarks, full_places):
    """
    Build the experimental 5th/14th anchor dict for every Conference|Event.

    For PDF-parsed conferences (UAA, NESCAC, MIAC):
        TRUE 5th  = full_places[(conf, event)][5]
        TRUE 14th = full_places[(conf, event)][14]
        Fall back to derived if a specific place is missing.

    For all others:
        DERIVED from 1st/8th/16th:
            fifth      = first + (4/7) * (eighth - first)
            fourteenth = eighth + (6/8) * (sixteenth - eighth)
    """
    exp_anchors: dict[str, dict] = {}

    for key_str, b in benchmarks.items():
        conf, event = key_str.split('|', 1)

        first     = b['first']
        eighth    = b['eighth']
        sixteenth = b['sixteenth']

        # Derived fallbacks
        fifth_derived      = first  + (4.0 / 7.0) * (eighth    - first)
        fourteenth_derived = eighth + (6.0 / 8.0) * (sixteenth - eighth)

        fifth      = fifth_derived
        fourteenth = fourteenth_derived
        source_5   = 'derived'
        source_14  = 'derived'

        if conf in PDF_PARSED_CONFS:
            # Lock to SAMPLE_SWIMMER gender (men) — the Excel 1st/8th/16th
            # benchmarks are men's data. Mixing women's 5th/14th with men's
            # 1st creates inverted zones (women's 5th > men's 14th for
            # events like 400 IM). Only use TRUE anchors when men's p5 AND
            # men's p14 are both present from the same PDF parse.
            p_dict = full_places.get((conf, event, SAMPLE_GENDER), {})
            if 5 in p_dict and 14 in p_dict:
                fifth      = p_dict[5]
                fourteenth = p_dict[14]
                source_5   = f'true_pdf({SAMPLE_GENDER})'
                source_14  = f'true_pdf({SAMPLE_GENDER})'

        # Safety clamp: if anchors are non-monotonic (fifth >= fourteenth),
        # the data is unreliable — fall back to derived values for both.
        if fifth >= fourteenth:
            fifth      = fifth_derived
            fourteenth = fourteenth_derived
            source_5   = 'derived_fallback(monotonicity)'
            source_14  = 'derived_fallback(monotonicity)'

        exp_anchors[key_str] = {
            'first':          first,
            'eighth':         eighth,
            'sixteenth':      sixteenth,
            'fifth':          fifth,
            'fourteenth':     fourteenth,
            'sec_per_place':  b['sec_per_place'],
            'source_5':       source_5,
            'source_14':      source_14,
        }

    return exp_anchors


def load_teams():
    """Load team records from Excel Team_Tiers."""
    teams = {}
    wb  = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws  = wb['Team_Tiers']
    h   = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {}
    for i, v in enumerate(h):
        if v and v not in col:
            col[v] = i + 1
    for r in range(2, ws.max_row + 1):
        conf   = ws.cell(r, col['Conference']).value
        school = ws.cell(r, col['Team']).value
        psf    = ws.cell(r, col['PSF']).value
        tier   = ws.cell(r, col['Tier']).value
        if not conf or not school:
            continue
        key = f"{conf}|{school}"
        if key not in teams:
            teams[key] = {
                'school':     school,
                'conference': conf,
                'psf':        _float(psf) if psf is not None else 1.0,
                'tier':       tier or '',
            }
    return teams

# ── PRODUCTION scoring (exact copy of production logic) ──────────────────────

def prod_estimate_place(sec, b):
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']
    spp       = b['sec_per_place'] or 1.0
    if sec <= first:
        return 1.0
    if sec <= eighth:
        return 1.0 + (sec - first) / ((eighth - first) or 1.0) * 7
    if sec <= sixteenth:
        return 8.0 + (sec - eighth) / ((sixteenth - eighth) or 1.0) * 8
    return 16.0 + (sec - sixteenth) / spp

def prod_exp_points(place):
    return max(0.0, min(20.0, 21.0 - place))

def prod_confidence(place):
    if place <= 12: return 1.0
    if place <= 14: return 0.85
    if place <= 16: return 0.65
    return 0.0

def prod_tier(pts):
    if pts < 1:  return 'Moonshot'
    if pts < 4:  return 'Reach'
    if pts < 10: return 'Recruitable'
    if pts < 18: return 'Priority Recruit'
    if pts < 35: return 'Top Recruit'
    if pts < 50: return 'Conference Star'
    return 'High-Point Contender'

def prod_place_label(place):
    if place <= 1.5:  return 'Contender'
    if place <= 3.5:  return 'Podium'
    if place <= 8.5:  return 'A Final'
    if place <= 16.5: return 'B Final'
    if place <= 20:   return 'Bubble'
    return 'Out of range'

def score_event_prod(event, time_str, conf, benchmarks):
    sec = parse_time(time_str)
    if sec is None:
        return None
    bench = benchmarks.get(f"{conf}|{event}")
    if bench is None:
        return None
    place  = prod_estimate_place(sec, bench)
    exp_pt = prod_exp_points(place)
    cw     = prod_confidence(place)
    pts    = exp_pt * cw
    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': prod_place_label(place),
    }

def score_school_prod(team_rec, times, benchmarks):
    conf = team_rec['conference']
    psf  = team_rec['psf']
    events = []
    for ev, t in times.items():
        es = score_event_prod(ev, t, conf, benchmarks)
        if es:
            events.append(es)
    events.sort(key=lambda e: e['pts'], reverse=True)
    top3    = events[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)
    if raw_pts == 0:
        return None
    adj_pts = round(raw_pts * psf, 2)
    return {
        'rawPts':   raw_pts,
        'adjPts':   adj_pts,
        'adjTier':  prod_tier(adj_pts),
        'top3':     top3,
        'allEvents': events,
    }

# ── EXPERIMENTAL v2 scoring (TRUE 5th/14th where available) ──────────────────

def exp_v2_estimate_place(sec, ea):
    """
    3-zone interpolation using the exp anchor dict (which holds true or derived
    fifth/fourteenth as appropriate for the conference).

    Zones:
      [1, 5]   — premium
      [5, 14]  — main scoring
      [14+]    — extension using 5-to-14 slope
    """
    first      = ea['first']
    fifth      = ea['fifth']
    fourteenth = ea['fourteenth']
    spp_ext    = max((fourteenth - fifth) / 9.0, 0.001)

    if sec <= first:
        return 1.0
    if sec <= fifth:
        return 1.0 + (sec - first) / max(fifth - first, 0.001) * 4.0
    if sec <= fourteenth:
        return 5.0 + (sec - fifth) / max(fourteenth - fifth, 0.001) * 9.0
    return 14.0 + (sec - fourteenth) / spp_ext

def exp_confidence(place):
    if place <= 14: return 1.0
    if place <= 16: return 0.35
    return 0.0

def exp_place_label(place):
    if place <= 1.5:  return 'Contender'
    if place <= 3.5:  return 'Podium'
    if place <= 5.5:  return 'Premium (1–5)'
    if place <= 14.5: return 'Main Scoring (6–14)'
    if place <= 16.5: return 'Extension (15–16)'
    return 'Out of range'

def score_event_exp_v2(event, time_str, conf, exp_anchors):
    sec = parse_time(time_str)
    if sec is None:
        return None
    ea = exp_anchors.get(f"{conf}|{event}")
    if ea is None:
        return None
    place  = exp_v2_estimate_place(sec, ea)
    exp_pt = max(0.0, min(20.0, 21.0 - place))
    cw     = exp_confidence(place)
    pts    = exp_pt * cw
    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': exp_place_label(place),
        'source_5':   ea['source_5'],
        'source_14':  ea['source_14'],
        'fifth_sec':  round(ea['fifth'], 3),
        'fourteen_sec': round(ea['fourteenth'], 3),
    }

def score_school_exp_v2(team_rec, times, exp_anchors):
    conf = team_rec['conference']
    psf  = team_rec['psf']
    events = []
    for ev, t in times.items():
        es = score_event_exp_v2(ev, t, conf, exp_anchors)
        if es:
            events.append(es)
    events.sort(key=lambda e: e['pts'], reverse=True)
    top3    = events[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)
    if raw_pts == 0:
        return None
    adj_pts = round(raw_pts * psf, 2)
    return {
        'rawPts':    raw_pts,
        'adjPts':    adj_pts,
        'adjTier':   prod_tier(adj_pts),
        'top3':      top3,
        'allEvents': events,
    }

# ── Runner ────────────────────────────────────────────────────────────────────

def fmt_top3(top3):
    return ' | '.join(
        f"{e['event']} pl={e['place']:.1f} pts={e['pts']:.1f}"
        for e in top3
    )

def run():
    print("Loading data...")
    benchmarks  = load_benchmarks()
    full_places = load_full_places()
    exp_anchors = build_exp_anchors(benchmarks, full_places)
    teams       = load_teams()

    # Report which conferences have TRUE vs derived anchors
    true_count    = sum(1 for ea in exp_anchors.values() if 'true_pdf' in ea.get('source_5', ''))
    derived_count = sum(1 for ea in exp_anchors.values() if ea.get('source_5', '') == 'derived')
    print(f"  {len(benchmarks)} benchmark entries, {len(teams)} team records")
    print(f"  Exp anchors: {true_count} TRUE (PDF-parsed), {derived_count} DERIVED\n")

    rows = []

    for school in TEST_SCHOOLS:
        team_rec = None
        for key, t in teams.items():
            if t['school'] == school:
                team_rec = t
                break
        if team_rec is None:
            print(f"  WARNING: {school!r} not found in teams data — skipping")
            continue

        conf = team_rec['conference']
        psf  = team_rec['psf']
        anchor_type = 'TRUE (PDF)' if conf in PDF_PARSED_CONFS else 'DERIVED'

        prod = score_school_prod(team_rec, SAMPLE_SWIMMER, benchmarks)
        exp  = score_school_exp_v2(team_rec, SAMPLE_SWIMMER, exp_anchors)

        if not prod and not exp:
            print(f"  {school}: no scorable events in {conf}")
            continue

        p_raw  = prod['rawPts']  if prod else 0
        p_adj  = prod['adjPts']  if prod else 0
        p_tier = prod['adjTier'] if prod else '—'
        p_top3 = fmt_top3(prod['top3']) if prod else '—'

        e_raw  = exp['rawPts']  if exp else 0
        e_adj  = exp['adjPts']  if exp else 0
        e_tier = exp['adjTier'] if exp else '—'
        e_top3 = fmt_top3(exp['top3']) if exp else '—'

        # Significant place/point shifts
        place_shifts = []
        anchor_sources = []
        if prod and exp:
            all_ev_prod = {e['event']: e for e in prod['allEvents']}
            all_ev_exp  = {e['event']: e for e in exp['allEvents']}
            for ev in all_ev_prod:
                if ev in all_ev_exp:
                    dp     = all_ev_exp[ev]['place']  - all_ev_prod[ev]['place']
                    dp_pts = all_ev_exp[ev]['pts']    - all_ev_prod[ev]['pts']
                    s5     = all_ev_exp[ev].get('source_5', '?')
                    s14    = all_ev_exp[ev].get('source_14', '?')
                    anchor_sources.append(f"{ev}:5th={s5},14th={s14}")
                    if abs(dp) >= 0.5 or abs(dp_pts) >= 0.5:
                        place_shifts.append(
                            f"{ev}: prod pl={all_ev_prod[ev]['place']:.1f}/pts={all_ev_prod[ev]['pts']:.1f}"
                            f" → exp pl={all_ev_exp[ev]['place']:.1f}/pts={all_ev_exp[ev]['pts']:.1f}"
                            f" (Δplace={dp:+.1f}, Δpts={dp_pts:+.1f}) [{s5}/{s14}]"
                        )

        delta_raw  = round(e_raw - p_raw, 2)
        delta_adj  = round(e_adj - p_adj, 2)
        tier_change = 'same' if p_tier == e_tier else f"{p_tier} → {e_tier}"

        rows.append({
            'school':          school,
            'conference':      conf,
            'anchor_type':     anchor_type,
            'psf':             psf,
            'prod_rawPts':     p_raw,
            'exp_rawPts':      e_raw,
            'delta_rawPts':    delta_raw,
            'prod_adjPts':     p_adj,
            'exp_adjPts':      e_adj,
            'delta_adjPts':    delta_adj,
            'prod_tier':       p_tier,
            'exp_tier':        e_tier,
            'tier_change':     tier_change,
            'prod_top3':       p_top3,
            'exp_top3':        e_top3,
            'place_shifts':    '; '.join(place_shifts),
            'anchor_sources':  '; '.join(anchor_sources),
        })

        print(f"{school} ({conf}, PSF={psf}, anchors={anchor_type})")
        print(f"  PROD  rawPts={p_raw:5.2f}  adjPts={p_adj:5.2f}  tier={p_tier}")
        print(f"  EXP   rawPts={e_raw:5.2f}  adjPts={e_adj:5.2f}  tier={e_tier}")
        print(f"  Δraw={delta_raw:+.2f}  Δadj={delta_adj:+.2f}  tier: {tier_change}")
        if place_shifts:
            for s in place_shifts:
                print(f"    {s}")
        print()

    # ── Write CSV ──────────────────────────────────────────────────────────────
    if rows:
        fieldnames = list(rows[0].keys())
        with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f"CSV written: {OUT_CSV}")

    # ── Write summary text ─────────────────────────────────────────────────────
    harder   = [r for r in rows if r['delta_adjPts'] < -1.0]
    similar  = [r for r in rows if abs(r['delta_adjPts']) <= 1.0]
    easier   = [r for r in rows if r['delta_adjPts'] > 1.0]
    tier_chg = [r for r in rows if r['tier_change'] != 'same']

    true_rows    = [r for r in rows if r['anchor_type'] == 'TRUE (PDF)']
    derived_rows = [r for r in rows if r['anchor_type'] == 'DERIVED']

    lines = [
        "═" * 64,
        "EXPERIMENTAL v2 vs PRODUCTION — SUMMARY",
        "TRUE 5th/14th anchors for PDF-parsed conferences",
        "═" * 64,
        "",
        f"Swimmer: representative D3-competitive profile",
        f"Events: {', '.join(SAMPLE_SWIMMER.keys())}",
        "",
        "Model differences:",
        "  Production:      3 zones (1st / 8th / 16th anchors)",
        "                   confidence: 1.0 (≤12), 0.85 (12–14), 0.65 (14–16), 0 (16+)",
        "  Experimental v2: 3 zones (1st / 5th / 14th anchors)",
        "                   confidence: 1.0 (≤14), 0.35 (14–16), 0 (16+)",
        "                   5th/14th: TRUE from PDF for UAA/NESCAC/MIAC",
        "                             DERIVED from 1st/8th/16th for Centennial/NWC",
        "",
        "─" * 64,
        "TRUE-anchor schools (UAA, NESCAC, MIAC):",
    ]
    for r in true_rows:
        chg = f"  tier: {r['tier_change']}" if r['tier_change'] != 'same' else ''
        lines.append(f"  {r['school']} ({r['conference']}): prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f}){chg}")

    lines += [
        "",
        "DERIVED-anchor schools (Centennial, NWC):",
    ]
    for r in derived_rows:
        chg = f"  tier: {r['tier_change']}" if r['tier_change'] != 'same' else ''
        lines.append(f"  {r['school']} ({r['conference']}): prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f}){chg}")

    lines += [
        "",
        "─" * 64,
        "Schools HARDER under experimental (adjPts dropped >1.0):",
    ]
    for r in harder:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f}) [{r['anchor_type']}]")
    if not harder:
        lines.append("  (none)")

    lines += ["", "Schools SIMILAR (adjPts within ±1.0):"]
    for r in similar:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f}) [{r['anchor_type']}]")
    if not similar:
        lines.append("  (none)")

    lines += ["", "Schools EASIER under experimental (adjPts gained >1.0):"]
    for r in easier:
        lines.append(f"  {r['school']}: prod={r['prod_adjPts']:.2f} → exp={r['exp_adjPts']:.2f}  ({r['delta_adjPts']:+.2f}) [{r['anchor_type']}]")
    if not easier:
        lines.append("  (none)")

    lines += ["", "Tier changes:"]
    for r in tier_chg:
        lines.append(f"  {r['school']}: {r['tier_change']} [{r['anchor_type']}]")
    if not tier_chg:
        lines.append("  (no tier changes for this swimmer profile)")

    lines += [
        "",
        "─" * 64,
        "Key question: do TRUE anchors differ materially from DERIVED?",
        "  If Δadj for TRUE-anchor schools ≈ Δadj for DERIVED schools → model is",
        "  robust; derived anchors are good enough and no harvester change is needed.",
        "  If TRUE-anchor schools show larger/smaller shifts → the 4/7 interpolation",
        "  is wrong and we need conference-specific calibration.",
        "═" * 64,
    ]

    summary = '\n'.join(lines)
    print(summary)
    with open(OUT_SUMMARY, 'w', encoding='utf-8') as f:
        f.write(summary + '\n')
    print(f"\nSummary written: {OUT_SUMMARY}")


if __name__ == '__main__':
    run()
